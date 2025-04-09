from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect
from decimal import Decimal
from django.contrib.auth.models import User
from django.db import IntegrityError, DataError
from django.contrib.staticfiles import finders
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.cell import MergedCell
import string
from django.utils import timezone
from django.http import FileResponse
from reportlab.lib.colors import HexColor
from .utils import get_browser_identifier
from django.contrib.sessions.models import Session
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, TableStyle, PageBreak, Image, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from site_panel.models import UserSession
from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.lib import utils
from openpyxl.drawing.image import Image as ExcelImage
import random
from decimal import Decimal
from django.utils import timezone
from django.utils.timezone import now
from reportlab.lib.pagesizes import A3
import smtplib
import pandas as pd
import os
from django.contrib.auth.decorators import user_passes_test
import re
from django.utils.crypto import get_random_string
import json
from django.conf import settings
import openpyxl
from io import BytesIO
from openpyxl.utils import get_column_letter
from django.views.decorators.csrf import csrf_exempt
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.db.models import Q
from .models import User, ActivitySales, ActivityPromotions, ActivityFinancial, ActivityPurchase, ActivityPaidOut, employee, compensation, Timesheet, PayDailyEntry
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.views.decorators.http import require_GET, require_POST
from django.http import JsonResponse
from django.contrib.auth.hashers import check_password
from django.db import connection, IntegrityError
from django.http import HttpResponse
from datetime import datetime, timedelta
from dateutil import parser
from django.core.paginator import Paginator, EmptyPage
from django.db.utils import DatabaseError  
import logging
from reportlab.lib.pagesizes import letter
from django.utils.timezone import make_aware
from reportlab.pdfgen import canvas
logger = logging.getLogger(__name__)

def admin_generate_reset_code():
    return ''.join(random.choices(string.digits, k=6))

def send_reset_email(user_email, reset_code):
    msg = MIMEMultipart()
    msg['From'] = 'jbansal640ab@gmail.com'
    msg['To'] = user_email
    msg['Subject'] = 'Password Reset Code'

    body = f'Your password reset code is: {reset_code}'
    msg.attach(MIMEText(body, 'plain'))

    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login('jbansal640ab@gmail.com', 'pfpk zcim csxs vjzi')
        text = msg.as_string()
        server.sendmail('jbansal640ab@gmail.com', user_email, text)

@require_POST
def admin_send_reset_email(request):
    email = request.POST.get('email')
    request.session['email'] = email

    if User.objects.filter(email=email).exists():
        reset_code = admin_generate_reset_code()
        send_reset_email(email, reset_code)
        request.session['reset_code'] = reset_code

        return JsonResponse({'success': True, 'message': 'OTP sent on your mail!'})
    
    else:
        return JsonResponse({"success": False, "message": "Email not found."})

def admin_check_pass(request):
    if request.method == 'POST':
        password = request.POST.get('password')
        password_regex = r"^(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*#?&])[A-Za-z\d@$!%*#?&]{8,}$"
        
        # Check if password matches the regex pattern
        if re.match(password_regex, password):
            return JsonResponse({'success': True, 'message': 'Password is valid!'})
        else:
            return JsonResponse({'success': False, 'message': 'Password must start with a capital letter, contain at least one digit, one special character, and be at least 8 characters long.'})
    return JsonResponse({'success': False, 'message': 'Invalid request method!'})

def admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.username == 'admin':
                browser_id = get_browser_identifier(request)

                try:
                    user_session = UserSession.objects.get(user=user)
                    
                    if user_session.session_key:
                        # Check if the session is still valid
                        print(browser_id)
                
                        print(user_session.browser_identifier)
                        try:
                            session = Session.objects.get(session_key=user_session.session_key)

                            # Check if the user is trying to log in from a different browser
                            if user_session.browser_identifier and user_session.browser_identifier != browser_id:
                                return render(request, 'site_panel/admin_login.html', {'error': 'User is already logged in from another browser.'})
                        except Session.DoesNotExist:
                            # If the session is invalid, clear it
                            user_session.session_key = None
                            user_session.save()
                except UserSession.DoesNotExist:
                    # Create a new UserSession object if it doesn't exist
                    user_session = UserSession.objects.create(user=user)

                login(request, user)

                # Force session to save so that session_key is generated
                request.session.save()
                
                # Now request.session.session_key should exist
                user_session.browser_identifier = browser_id
                user_session.session_key = request.session.session_key  # Update session key
                user_session.save()

            
                return redirect('/site_panel/admin/dashboard/')

            else:
                return render(request, 'site_panel/admin_login.html', {'error': 'You are not authorized to access this site.'})
        return render(request, 'site_panel/admin_login.html', {'error': 'Invalid username or password'})

    return render(request, 'site_panel/admin_login.html')

def admin_logout(request):
    if request.user.is_authenticated:
        request.session.flush()  # Remove session data completely
        logout(request)  # Logs the user out
    return redirect('/site_panel/admin')

@require_POST
def admin_verify_reset_code(request):
        reset_code = request.session.get('reset_code') 
        entered_code = request.POST.get('reset_code')

        if reset_code and reset_code == entered_code:
            request.session['reset_verified'] = True  # Set session flag
            return JsonResponse({"success": True, "message": "Code verified successfully."})
        else:
            return JsonResponse({"success": False, "message": "Invalid reset code."})
        
def admin_email_otp_verification(request): 
    return render(request, 'site_panel/email_otp_verification.html')

def admin_new_pass(request): 
    if request.session.get('reset_verified'):  # Check if code is verified
        return render(request, 'site_panel/new_pass.html')
    else:
        return redirect('/site_panel/reset_password/')  # Redirect to reset page if not verified

@csrf_exempt  # Temporarily disabling CSRF for testing; use proper CSRF protection in production
def change_admin_password(request):
    if request.method == "POST":
        new_password = request.POST.get("newPassword")
        confirm_password = request.POST.get("rePassword")

        # Validate if both passwords match
        if new_password != confirm_password:
            return JsonResponse({"success": False, "message": "Passwords do not match."})

        # Fetch the admin user (assuming username is 'admin')
        try:
            admin_user = User.objects.get(username="admin")  # Change if your admin username is different
            admin_user.set_password(new_password)  # Set new password
            admin_user.save()  # Save changes

            return JsonResponse({"success": True, "message": "Password changed successfully."})

        except User.DoesNotExist:
            return JsonResponse({"success": False, "message": "Admin user not found."})

    return JsonResponse({"success": False, "message": "Invalid request."})


def is_super_admin(user):
    return user.id == 1

@user_passes_test(is_super_admin, login_url='/site_panel/admin')
def admin_dashboard(request):
    return render(request, 'site_panel/admin_dashboard1.html')


@user_passes_test(is_super_admin, login_url='/site_panel/admin')
def admin_Create_site(request):
    return render(request, 'site_panel/Create_site.html')

@user_passes_test(is_super_admin, login_url='/site_panel/admin')
def admin_viewSites(request): 
    return render(request, 'site_panel/View_sites.html')

def admin_Insert_new_site(request):
    if request.method == 'POST':
        site_name = request.POST.get('site-name')
        location = request.POST.get('address')
        username_site = request.POST.get('username')
        password_site = request.POST.get('password')

        if User.objects.filter(Q(first_name=site_name) | Q(last_name=location) | Q(username=username_site)).exists():
            return JsonResponse({'success': False, 'message': 'One of the fields already exists!'})

        # Check if the password is already in use
        for user in User.objects.all():
            if user.password and check_password(password_site, user.password):
                return JsonResponse({'success': False, 'message': 'Password is already in use!'})

      
        user = User(
                    first_name=site_name,
                    last_name=location,
                    username=username_site,
                    date_joined=timezone.now(),
                    is_staff=True,
                    is_active=True,
                    is_superuser=False
                )

        user.set_password(password_site)
        user.save()

        return JsonResponse({'success': True, 'message': 'Site created successfully!'})

    return JsonResponse({'success': False, 'message': 'Invalid request method!'})

@require_GET
def admin_view_sites(request):
    users = User.objects.filter(is_staff=True).all()
    site_list = [
        {
            "id": user.id,
            "site_name": user.first_name,
            "address": user.last_name,
            "username": user.username,
            "password": "******",
            "active": user.is_active,
            "superuser": user.is_superuser
        } for user in users
    ]
    return JsonResponse(site_list, safe=False)

@csrf_exempt  
@require_POST
def admin_update_site_password(request):
    try:
        data = json.loads(request.body)
        site_id = data.get('site_id')
        new_password = data.get('new_password')
        print(f"Received site_id: {site_id}, new_password: {new_password}")
        print(site_id)

        if site_id is None or new_password is None:
            return JsonResponse({"success": False, "message": "Invalid request data"}, status=400)

        user = User.objects.get(id=site_id)
        user.set_password(new_password)
        user.save()

        return JsonResponse({"success": True, "message": "Site password updated successfully!"})
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

@csrf_exempt  # If you are handling POST requests without CSRF token in your AJAX calls
@require_POST
def admin_update_site_status(request):
    try:
        data = json.loads(request.body)
        site_id = data.get('site_id')
        new_status = data.get('new_status')

        if site_id is None or new_status is None:
            return JsonResponse({"success": False, "message": "Invalid request data"}, status=400)

        try:
            user = User.objects.get(id=site_id)
        except User.DoesNotExist:
            return JsonResponse({"success": False, "message": "User not found"}, status=404)

        user.is_active = new_status
        user.save()

        # End the user's session and delete the UserSession entry
        try:
            user_session = UserSession.objects.get(user=user)
            session_key = user_session.session_key

            # Delete the Django session
            if session_key:
                Session.objects.filter(session_key=session_key).delete()

            # Delete the UserSession entry
            user_session.delete()
        except UserSession.DoesNotExist:
            # User might not have an active session or UserSession entry
            pass
        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error ending session: {str(e)}"}, status=500)
    
        if new_status == 1:
            return JsonResponse({"success": True, "message": "Site activated successfully!!"})
        
        elif new_status == 0:

            return JsonResponse({"success": True, "message": "Site deactivated successfully!"})
        
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "message": "Invalid JSON data"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

@csrf_exempt  # If you are handling POST requests without CSRF token in your AJAX calls
@require_POST
def admin_update_site_superuser(request):
    try:
        data = json.loads(request.body)
        site_id = data.get('site_id')
        is_superuser = data.get('NewSup')

        if site_id is None or is_superuser is None:
            return JsonResponse({"success": False, "message": "Invalid request data"}, status=400)
        
        user = User.objects.get(id=site_id)
        user.is_superuser = is_superuser
        user.save()

        if is_superuser == 1:

            return JsonResponse({"success": True, "message": f"Superuser privileges granted to {user.first_name}"})
        
        elif is_superuser == 0:

            return JsonResponse({"success": True, "message": f"Superuser privileges removed from {user.first_name}"})
        
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

def site_login(request):
    if request.method == "POST":
        username_site = request.POST.get("username")
        password_site = request.POST.get("password")

        if not username_site or not password_site:
            return render(request, 'site_panel/site_login.html', {'error': 'Both username and password are required.'})

        user = authenticate(username=username_site, password=password_site)

        if user is not None:
            if (user.is_active and user.is_staff) or (user.last_name == 'all'):
                browser_id = get_browser_identifier(request)

                try:
                    user_session = UserSession.objects.get(user=user)

                    # Check if a session key is set
                    if user_session.session_key:
                        try:
                            # Fetch the session
                            session = Session.objects.get(session_key=user_session.session_key)

                            # Ensure session is still valid (not expired)
                            session_data = session.get_decoded()
                            if session.expire_date > timezone.now():
                                # Check browser mismatch
                                if user_session.browser_identifier != browser_id:
                                    return render(request, 'site_panel/site_login.html', {
                                        'error': 'User is already logged in from another browser.'
                                    })
                            else:
                                # Session expired → clear old session
                                user_session.session_key = None
                                user_session.browser_identifier = None
                                user_session.save()

                        except Session.DoesNotExist:
                            # Session not found in DB → clear session
                            user_session.session_key = None
                            user_session.browser_identifier = None
                            user_session.save()

                except UserSession.DoesNotExist:
                    user_session = UserSession.objects.create(user=user)

                # Log the user in and set session variables
                login(request, user)
                request.session.save()

                user_session.browser_identifier = browser_id
                user_session.session_key = request.session.session_key
                user_session.save()

                request.session['site_location'] = user.last_name
                request.session['user_name'] = username_site
                request.session['restaurant_name'] = user.first_name
                request.session['userLoggedIn'] = True  # ✅ Persist login state

                # ✅ Extend session duration
                request.session.set_expiry(86400)  # 1 day

                return redirect('/site_panel/site_home/')  # Use redirect instead of render

            else:
                return render(request, 'site_panel/site_login.html', {'error': 'You are not authorized to access this site.'})

        return render(request, 'site_panel/site_login.html', {'error': 'Invalid username or password'})

    return render(request, 'site_panel/site_login.html')

def check_session(request):

    last_username=request.GET.get('last_username')

    try:
        user = User.objects.get(username=last_username)
        print(user.is_active)
        
        return JsonResponse({
            'is_active': int(user.is_active)
        })
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def site_logout(request):
    if request.user.is_authenticated:
        try:
            user_session = UserSession.objects.get(user=request.user)
            user_session.session_key = None
            user_session.browser_identifier = None
            user_session.save()
        except UserSession.DoesNotExist:
            pass
    logout(request)
    return redirect('/site_panel/')

def is_site_admin(user):
    if user.is_authenticated:
        return user.is_staff or user.last_name == 'all'
    return False

def your_view(request):
    try:
        image_dir = os.path.join(settings.STATIC_ROOT, 'site_panel/images')
        if not os.path.exists(image_dir):
            raise FileNotFoundError(f"Directory {image_dir} not found")
            
        all_images = [f for f in os.listdir(image_dir) 
                     if f.endswith(('.jpg', '.png', '.jpeg'))]
        
        if not all_images:
            raise ValueError("No images found in directory")
            
        random_images = random.sample(all_images, min(8, len(all_images)))
        return render(request, 'site_home.html', {'random_images': random_images})
        
    except Exception as e:
        # Handle the error appropriately
        return render(request, 'error.html', {'error': str(e)})

@user_passes_test(is_site_admin, login_url='/site_panel/')
def site_home(request):
    # You can add any context data you want to pass to the template here

    return render(request, 'site_panel/site_home.html')

@user_passes_test(is_site_admin, login_url='/site_panel/')
def end_session(request):
    if request.user.is_authenticated:
        # Clear active session when the user logs out
        request.user.session_info.active_session = None
        request.user.session_info.save()
        
        logout(request)  # Django's logout function # This triggers the logout signal and clears active_session
    return redirect('/site_panel/')

@user_passes_test(is_site_admin, login_url='/site_panel/')
def way_to_logout(request):
    return render(request, 'site_panel/site_logout.html')

@user_passes_test(is_site_admin, login_url='/site_panel/')
def site_dashboard(request):
        return render(request, 'site_panel/site_home.html')

@user_passes_test(is_site_admin, login_url='/site_panel/')
def site_activity_tracker_dashboard(request):
    # Your SQL query
    try:
        site_username=request.session.get('user_name')
        user = User.objects.get(username=site_username)
        site_id = user.id
        

        base_query = """
            SELECT 
                s.site_id as "SiteID",
                s.activity_date AS "Date",
                s.sales_notes AS "Notes",
                u.last_name AS "Location"
            FROM 
                activity_sales s
            LEFT JOIN 
                activity_purchase p ON s.site_id = p.site_id AND s.activity_date = p.activity_date
            LEFT JOIN 
                activity_paid_out po ON s.site_id = po.site_id AND s.activity_date = po.activity_date
            LEFT JOIN 
                activity_financial f ON s.site_id = f.site_id AND s.activity_date = f.activity_date
            LEFT JOIN 
                activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
            LEFT JOIN  
                auth_user u ON s.site_id = u.id
            """
        
        query_params = []
        if user.last_name != 'all':
            base_query += " WHERE s.site_id = %s ORDER BY s.activity_date DESC"
            query_params.append(site_id)
        elif user.last_name == 'all':
            site_selected_location = request.GET.get('site_selected_location')
            if site_selected_location:
                base_query += "WHERE u.last_name = %s ORDER BY s.activity_date DESC;"
                query_params.append(site_selected_location)
            else:
                base_query += "ORDER BY u.id, s.activity_date DESC;"

        base_query += " "

        # Execute the raw SQL query
        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        except DatabaseError as db_error:
            raise ValueError(f"Database error: {str(db_error)}")


        per_page = int(request.GET.get('per_page', 1))  # Default 1 (as requested)
        page_number = int(request.GET.get('page', 1))

        paginator = Paginator(results, per_page)

        # Handle edge cases (no results OR empty page)
        if not paginator.object_list or page_number > paginator.num_pages:  # No results or page number too high
            page_obj = paginator.get_page(1) if paginator.object_list else None # Empty page if no results
            page_range = []  # Empty page range
            return render(request, 'site_panel/Site_Daily_Activity_dashboard.html', {
                'activity_data': [],  # Empty list for the template
                'page_obj': page_obj,
                'page_range': page_range,
                'per_page': per_page,
                'no_results': True if not paginator.object_list else False,  # Flag for no results
                'invalid_page': True if page_number > paginator.num_pages and paginator.object_list else False, # Flag for invalid page number
            })

        page_obj = paginator.get_page(page_number)
        page_range = paginator.page_range

        return render(request, 'site_panel/Site_Daily_Activity_dashboard.html', {
            'activity_data': page_obj.object_list,
            'page_obj': page_obj,
            'page_range': page_range,
            'per_page': per_page,
        })

    except Exception as e:
        # If request is AJAX, return JSON error response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({"success": False, "error": str(e)}, status=400)

        # Otherwise, return error page
        return render(request, 'site_panel/error.html', {"error_message": str(e)})

@user_passes_test(is_site_admin, login_url='/site_panel/')
def Site_New_Employees_Addition(request):
    return render(request, 'site_panel/Site_New_Employees_Addition.html')

@user_passes_test(is_site_admin, login_url='/site_panel/')
def Save_new_Compensation_data(request):
    employee_id = request.GET.get("e_id", None)
    print(employee_id)
    return render(request, 'site_panel/e_new_comp_create.html', {"employee_id": employee_id})

def normalize_phone(phone):
    """Removes all non-numeric characters from the phone number."""
    if phone is None:
        print("normalize_phone received None")  # Debugging log
        return None
    return re.sub(r'\D', '', phone)  # \D removes all non-digit characters

def is_valid_phone(phone):
    """
    Validates a phone number using the same regex logic as the frontend JavaScript.
    """
    if not phone:
        return False

    phone = phone.strip()

    # Regular expression for valid phone number format
    pattern = re.compile(r"^\(?(?:\+\d{1,3})\)?[-.\s]?\d{6,14}$")
    
    # Check for invalid characters (anything other than numbers, +, -, ., (), spaces)
    invalid_chars_pattern = re.compile(r"[^0-9+\-.\s()]")
    
    # Check for unbalanced brackets
    unbalanced_brackets_pattern = re.compile(r"\([^\d)]*\)|\([^\d]*$|\)[^\d(]*\(")

    if invalid_chars_pattern.search(phone):
        return "Invalid characters detected! Only numbers, +, -, ., (), and spaces are allowed."
    elif unbalanced_brackets_pattern.search(phone):
        return "Invalid format! Ensure parentheses are used correctly, e.g., (+91) 9876543210."
    elif not pattern.match(phone):
        return "Invalid format! Ensure the number starts with a country code or spacing issue (e.g., +1, +91)."
    return True  # Valid phone number

@require_POST
@user_passes_test(is_site_admin, login_url='/site_panel/')
def Save_new_Employee_data(request):
    site_username=request.session.get('user_name')
    user = User.objects.get(username=site_username)
    site_id = user.id

    form_name = request.POST.get('form_name')

    if form_name=='NewEmployee':
        errors = {}
        try:
            # Extracting form data
            # Ensure this is sent from the form
            
            s_id = site_id
            location = request.POST.get("location")
            owner = request.POST.get("Employee_Owner")  # Get the value from POST

            if owner:  # Only strip if owner is not None
                owner = owner
            else:
                owner = None  # Ensure it remains None if empty

            pattern = re.compile(r"^[A-Za-z]+$")  # Only allows alphabets, no spaces

            if owner and not pattern.match(owner):  # Validate only if owner is not None
                errors["Employee_Owner"] = "Invalid owner name! Only alphabets are allowed, no spaces."
            employee_name = request.POST.get("Employee_name")
            pattern2 = re.compile(r"^[A-Za-z]+(?:\s[A-Za-z]+)*$")

            if not pattern2.match(employee_name):
                errors["Employee_name"] = "Invalid Employee name! Only alphabets and trailing spaces aren't allowed."

            start_date = request.POST.get("Employee_Start_date")
            # Ensure the start_date is within the valid range (not a week before today)
            
            end_date = request.POST.get("Employee_End_date") or None
            phone1 = request.POST.get("phone1")  # Ensure phone1 is always a string

            phone1_validation = is_valid_phone(phone1)

            if phone1_validation is not True:
                errors["phone1"] = "Phone number isn't in the required format."

            print(f"Raw phone1 from POST: {phone1}")  # Debugging log
            normalized_phone1 = normalize_phone(phone1) if phone1 else None  # Only normalize if it's not empty
            print(f"Normalized phone1: {normalized_phone1}")  # Debugging log

            phone2 = request.POST.get('phone2') if request.POST.get('phone2') else ''
            phone2 = request.POST.get("phone2").strip()
            if phone2:
                phone2_validation = is_valid_phone(phone2)
                if phone2_validation is not True:
                    errors["phone2"] = "Phone number isn't in the required format."
            
            normalized_phone2 = normalize_phone(phone2) if phone2 else None  # Only normalize if it's not empty
            smart_serve = request.POST.get("smart_serve") or None

            if not employee_name:
                errors["Employee_name"] = "Employee name is required."

            if not start_date:
                errors["Employee_Start_date"] = "Start date is required."

            if not location:
                errors["location"] = "Location date is required."

            if not normalized_phone1:
                errors["phone1"] = "Phone number is required."
            elif len(normalized_phone1) < 10:  # Ensure it's not None before checking length
                errors["phone1"] = "Phone number must be at least 10 digits."

            if errors:
                return JsonResponse({"success": False, "errors": errors})
        
            request.session['Employee_name'] = employee_name
            request.session['Employee_startdate'] = start_date # Convert date fields
            print(f"Received start_date: {start_date}")
            print(f"Received end_date: {end_date}")
            
            if start_date:  # Check if end_date_str is not empty or None
                    try:
                        start_date=parser.parse(start_date).date()
                    except ValueError:
                        return JsonResponse({"success": False, "message": "Invalid Start Date format!"})
            print(f"Converted start_date: {start_date}")
            if end_date:  # Check if end_date_str is not empty or None
                    try:
                        end_date = parser.parse(end_date).date()
                    except ValueError:
                        return JsonResponse({"success": False, "message": "Invalid End Date format!"}) # Correctly handle date format errors
            print(f"Converted end_date: {end_date}")

            if smart_serve:
                existing_phone = employee.objects.filter(s_id=s_id, employee_name=employee_name,location=location, smart_serve=smart_serve, active_inactive=1).filter(phone1=r'^\D*' + normalized_phone1 + r'\D*$').first()
            else:
                existing_phone = employee.objects.filter(s_id=s_id, employee_name=employee_name,location=location, active_inactive=1).filter(phone1=r'^\D*' + normalized_phone1 + r'\D*$').first()

            if existing_phone:
                return JsonResponse({"success": False, "message": "Employee already exist!"})

            # Save data to the database
            try:
                    new_employee = employee.objects.create(
                    s_id=s_id,
                    location=location,
                    owner=owner,
                    employee_name=employee_name,
                    start_date=start_date,
                    end_date=end_date,
                    phone1=normalized_phone1,
                    phone2=normalized_phone2,
                    smart_serve=smart_serve,
                    )
                    new_employee.save()
                    e_id=new_employee.e_id

                    return JsonResponse({
                        "success": True,
                        "message": "Employee saved successfully!",
                        "formName": "NewEmployee",
                        "data": e_id  # Sending the newly created employee's ID
                    })

            except IntegrityError as e:
                error_message = str(e)

                # Identify field causing error
                if "unique constraint" in error_message.lower():
                        if "employee_name" in error_message.lower():
                            errors["Employee_name"] = "This employee name already exists."
                        elif "phone1" in error_message.lower():
                            errors["phone1"] = "This phone number is already registered."
                        elif "smart_serve" in error_message.lower():
                            errors["smart_serve"] = "This smart_server is already registered."
                        else:
                            errors["general"] = "A unique constraint error occurred."

                elif "foreign key" in error_message.lower():
                        if "s_id" in error_message.lower():
                            errors["general"] = "Invalid site id constraint extracted from auth_user."
                        else:
                            errors["general"] = "A foreign key constraint failed."

            except DataError as e:
                return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}})

            return JsonResponse({"success": False, "errors": errors})

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error: {str(e)}"})
        
    elif form_name=='Compensation':
        errors = {}
        try:
            employee_tb_name = request.session.get('Employee_name')
            employee_name = request.POST.get("employee_Compensation_name")

            if not employee_name:
                errors["employee_Compensation_name"] = "Employee name is required."
            pattern = re.compile(r"^[A-Za-z]+(?:\s[A-Za-z]+)*$")
            if not pattern.match(employee_name):  # Validate only if owner is not None
                errors["employee_Compensation_name"] = "Invalid Employee name! Only alphabets and trailing spaces aren't allowed."

            page_type = request.POST.get("page_type")
            e_position = request.POST.get("employee_position")
            pattern2 = re.compile(r"^[A-Za-z]+$")
            if not pattern2.match(e_position):
                errors["employee_position"] = "Invalid Employee's position! Only alphabets are allowed, no spaces."
            
            employee_compensation_ffective_Date = request.POST.get("employee_Effective_Date")
            if not employee_compensation_ffective_Date:
                errors["employee_compensation_ffective_Date"] = "Effective Date is required."
            employee_startdate = request.session.get("Employee_startdate")
            emp_id= request.POST.get("employee_id")
            if page_type=='main_page':
                if employee_name==employee_tb_name and employee_startdate==employee_compensation_ffective_Date:
                    emp=employee.objects.get(employee_name= employee_name, start_date=employee_startdate)
                    emp_id=emp.e_id
                    print(emp_id)
                else:
                    errors["employee_compensation_ffective_Date"] = "Effective Date should be equal to Employee's joining date."

            elif page_type == 'edit_page':
                try:
                    emp = employee.objects.get(employee_name=employee_name, e_id=emp_id)
                    # If get() succeeds, emp exists, and you can proceed

                except employee.DoesNotExist:
                    # Employee not found, handle error
                    errors["general"] = "Incorrect employee name or invalid ID!"
 
            if errors:
                return JsonResponse({"success": False, "errors": errors})
            
            try:

                if not emp_id:
                    errors["general"] = "Employee Id field is empty or not extracting!"
                # Extracting form data
                s_id = site_id  # Ensure this is sent from the form
                e_id= emp_id
                employee_compensation_name=employee_name
                e_position = e_position
                e_position_type = request.POST.get("employee_position_type")
                                
                if e_position_type not in ['Front of the House', 'Back of the House', 'Staff/Supervisor', 'Management/Owner'] :
                    errors["employee_position_type"] = "Invalid position type! Choose a valid option."
                
                e_rate_type = request.POST.get("employee_rate_type")
                e_rate_type = e_rate_type 
                if e_rate_type not in ['Hourly Salary', 'Daily Salary'] :
                    errors["e_rate_type"] = "Invalid rate type! Choose a valid option."
                e_rate = request.POST.get("employee_rate")
                pattern3= re.compile(r"^\d+(\.\d{1,2})?$")
                if not pattern3.match(e_rate):
                    errors["employee_rate"] = "Invalid rate! Enter a valid number with up to two decimal places."
                e_comp_end_date = request.POST.get("employee_compensation_end_date") or None

                if employee_compensation_ffective_Date:  # Check if end_date_str is not empty or None
                    try:
                        e_compensation_ffective_Date=parser.parse(employee_compensation_ffective_Date).date()
                    except ValueError:
                        return JsonResponse({"success": False, "message": "Invalid Effective Date format!"})

                if e_compensation_ffective_Date < emp.start_date :
                        errors["employee_compensation_ffective_Date"] = "Effective Date shouldn't be less than Employee's joining date."

                if not e_position:
                    errors["employee_position"] = "Employee's position is required."
                if not e_position_type:
                    errors["employee_position_type"] = "Employee's position-type is required."
                if not e_rate_type:
                    errors["employee_rate_type"] = "Employee's salary-type is required."
                if not e_rate:
                    errors["employee_rate"] = "Employee's salary rate is required."
                if errors:
                    return JsonResponse({"success": False, "errors": errors})

                # Convert date fields
                print(f"Received Effective_Date: {employee_compensation_ffective_Date}")
                print(f"Received compensation end_date: {e_comp_end_date}")
                    
                if e_comp_end_date:  # Check if end_date_str is not empty or None
                    try:
                        e_comp_end_date=parser.parse(e_comp_end_date).date()
                    except ValueError:
                        return JsonResponse({"success": False, "message": "Invalid End Date format!"})

                print(f"Converted Effective_Date: {employee_compensation_ffective_Date}")
                print(f"Converted Compensation end_date: {e_comp_end_date}")

                existing_emp_comp = compensation.objects.filter(s_id=s_id, e_id=e_id, active_inactive=1).first()
                if existing_emp_comp:
                    return JsonResponse({"success": False, "message": "Employee's compensation data already got added!"})


                # Save data to the database
                try:

                    new_employee_compensation = compensation.objects.create(
                        s_id=s_id,
                        e_id=e_id,
                        e_position=e_position,
                        e_position_type=e_position_type,
                        e_rate_type=e_rate_type,
                        e_rate=e_rate,
                        e_compensation_ffective_Date=e_compensation_ffective_Date,
                        e_comp_end_date=e_comp_end_date,
                        employee_compensation_name=employee_compensation_name,
                    )
                    new_employee_compensation.save()

                    empl_details = compensation.objects.filter(e_id=e_id, e_rate=e_rate).order_by('-c_id').values('c_id', 'e_id', 'e_rate').first()

                    if empl_details: # Check if empl_details is not None
                        # empl_details is now a dictionary (or None)
                        empl_details_list = [empl_details] #If you need a list you can convert it. But if you want a single dictionary then use empl_details only.
                    else:
                        empl_details_list = []
                        return JsonResponse({"success": False, "message": "Employee details can't be extracted from Employee table!"})

                    return JsonResponse({"success": True, "message": "Compensation data saved successfully!", "empl_details": empl_details_list})
                
                except IntegrityError as e:
                    error_message = str(e)

                # Identify field causing error
                    if "unique constraint" in error_message.lower():
                        if "employee_compensation_name" in error_message.lower():
                            errors["employee_compensation_name"] = "This employee name already exists."
                        else:
                            errors["general"] = "A unique constraint error occurred."

                    elif "foreign key" in error_message.lower():
                        if "s_id" in error_message.lower():
                            errors["general"] = "Invalid site id constraint extracted from auth_user."
                        elif "e_id" in error_message.lower():
                            errors["general"] = "Invalid employee id constraint extracted from employee table."
                        else:
                            errors["general"] = "A foreign key constraint failed."

                except DataError as e:
                    return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}})

                return JsonResponse({"success": False, "errors": errors})
            
            except Exception as e:
                return JsonResponse({"success": False, "message": f"Error: {str(e)}"})

        except Exception as e:
                return JsonResponse({"success": False, "message": f"Error: {str(e)}"})

    return JsonResponse({"success": False, "message": "Invalid request method!"})

def Update_Employee_data(request):

    try:
        form_name = request.POST.get('form_name')
        
        if form_name == 'UpdateEmployee':
            errors = {}
            active_inactive = 1

            # Extract required fields
            e_id = request.POST.get('eid')
            if not e_id:
                errors["general"] = "Employee ID is missing."

            s_id = request.POST.get('sid')
            if not s_id:
                errors["general"] = "Site ID is missing."

            # Extract optional fields
            location = request.POST.get('location')
            employee_owner = request.POST.get('Employee_Owner') if request.POST.get('Employee_Owner') else ''
            pattern = re.compile(r"^[A-Za-z]*$")  # Only allows alphabets, no spaces

            if not pattern.match(employee_owner):
                errors["Employee_Owner"] = "Invalid owner name! Only alphabets are allowed, no spaces.."
            
            employee_name = request.POST.get('Employee_name')
            pattern2 = re.compile(r"^[A-Za-z\s]+$")

            if not pattern2.match(employee_name):
                errors["Employee_name"] = "Invalid Employee name! Only alphabets and spaces are allowed."
            employee_start_date = request.POST.get('Employee_Start_date')
            employee_end_date = request.POST.get('Employee_End_date') or None
            phone1 = request.POST.get('phone1')
            phone1_validation = is_valid_phone(phone1)
            if phone1_validation is not True:
                errors["phone1"] = "Phone number isn't in the required format."
            normalized_phone1 = normalize_phone(phone1)
            phone2 = request.POST.get('phone2') if request.POST.get('phone2') else ''
            phone2 = request.POST.get("phone2").strip()
            if phone2:
                phone2_validation = is_valid_phone(phone2)
                if phone2_validation is not True:
                    errors["phone2"] = "Phone number isn't in the required format."
            normalized_phone2 = normalize_phone(phone2)
            smart_serve = request.POST.get('smart_serve') if request.POST.get('smart_serve') else None

            if not employee_name:
                errors["Employee_name"] = "Employee name is required."

            if not employee_start_date:
                errors["Employee_Start_date"] = "Start date is required."

            if not location:
                errors["location"] = "Location date is required."

            if not normalized_phone1:
                errors["phone1"] = "Phone number is required."
            elif len(normalized_phone1) < 10:  # Ensure it's not None before checking length
                errors["phone1"] = "Phone number must be at least 10 digits."

            # Validate date fields
            try:
                if employee_start_date:
                    employee_start_date = parser.parse(employee_start_date).date()
                if employee_end_date:
                    employee_end_date = parser.parse(employee_end_date).date()
            except ValueError:
                errors["date"] = "Invalid date format in start or end date."

            if employee_end_date and employee_end_date <= datetime.now().date():
                active_inactive = 0


            # If errors exist, return response
            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

            data = {
                'location': location,
                'owner': employee_owner,
                'employee_name': employee_name,
                'start_date': employee_start_date,
                'end_date': employee_end_date,
                'phone1': normalized_phone1,
                'phone2': normalized_phone2,
                'smart_serve': smart_serve,
                'active_inactive': active_inactive,
            }

            # Save data to the database
            try:
                obj, created = employee.objects.update_or_create(
                    e_id=e_id,
                    s_id=s_id,
                    defaults=data
                )

                return JsonResponse({'success': True, 'message': 'Employee record updated!' if not created else 'Employee data added successfully!'})

            except IntegrityError as e:
                error_message = str(e)

                if "unique constraint" in error_message.lower():
                    errors["general"] = "A unique constraint error occurred. Please check the data."
                elif "foreign key" in error_message.lower():
                    errors["general"] = "Foreign key constraint failed. Invalid Employee ID or Site ID."
                else:
                    errors["general"] = "Database integrity error occurred."

            except DataError as e:
                return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

            return JsonResponse({"success": False, "errors": errors}, status=400)

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)


@user_passes_test(is_site_admin, login_url='/site_panel/')
def Site_Edit_empl(request):
    """Fetch data from MySQL and render the edit page with pre-filled values."""
    if request.method == 'GET':
        empl_details_json = request.GET.get('empl_details')
        try:
                empl_details_list = json.loads(empl_details_json)  # Parse JSON (list of dictionaries)

                if isinstance(empl_details_list, list) and len(empl_details_list) > 0:
                    first_item = empl_details_list[0]  # Get the first dictionary from the list.
                    e_id = first_item.get('e_id')# Get the e_id from the dictionary.
                    print(e_id)

                    if e_id is not None: #Check if e_id is not None
                        employee_data = employee.objects.filter(e_id=e_id).first() #Now e_id is a single integer.

                        if employee_data:
                            # ... your existing code to process employee_data ...
                            # Prepare context with pre-filled data
                            context = {
                                'e_id': e_id,
                                'employee': employee_data,
                            }

                            return render(request, 'site_panel/Site_Edit_empl.html', context)
                    else:
                        return HttpResponse("Employee data not found.")
                else:
                    return HttpResponse("e_id is missing in empl_details.")
        except json.JSONDecodeError:
            return HttpResponse("Invalid empl_details JSON data.")
    else:
        return HttpResponse("empl_details data is missing.")
    

@user_passes_test(is_site_admin, login_url='/site_panel/')
def get_compensation_data(request):
    if request.method == 'POST':
        try:  
            eID = request.POST.get('eID')
            if not eID:
                return JsonResponse({'success': False, 'message': 'eID is required (POST).'})

            try:  

                request.session['compensation_e_id'] = eID

                return JsonResponse({'success': True})

            except compensation.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Compensation record not found.'})
            except Exception as e:
                print(e)
                return JsonResponse({'success': False, 'message': str(e)})

        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': str(e)})

    elif request.method == 'GET':
        eID = request.session.get('compensation_e_id')

        if not eID:
            return render(request, 'site_panel/Site_Edit_empl.html', {
                'compensation_data': [],
                'page_obj': None,
                'page_range': [],
                'per_page': 1,
                'no_results': True,
                'invalid_page': False,
            })
        try:
            with connection.cursor() as cursor:
                current_date = datetime.now().date()
                query_param = [0, current_date, eID]  # 0 = inactive, current_date, site_id
                query = """
                UPDATE compensation
                SET active_inactive = %s
                WHERE e_comp_end_date <= %s AND e_id = %s;
                """
                cursor.execute(query, query_param)
                    
        except Exception as e:
            print(f"Error updating employee status: {e}")

        try:  
            query = """
                SELECT 
                    c.c_id as "cID",
                    c.employee_compensation_name AS "Emp_Name",
                    c.e_position AS "Position",
                    c.e_position_type AS "Position_Type",
                    c.e_rate_type AS "Rate_Type",
                    c.e_rate AS "Rate",
                    c.e_compensation_ffective_Date AS "Effective_Date",
                    c.e_comp_end_date AS "End_Date"
                FROM 
                    compensation c
                WHERE
                    c.e_id = %s
                ORDER BY 
                    c.e_compensation_ffective_Date;
            """

            with connection.cursor() as cursor:
                cursor.execute(query, [eID])
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]
                
            
            per_page = int(request.GET.get('per_page', 25))  
            paginator = Paginator(results, per_page)
            page_number = int(request.GET.get('page', 1))  # Convert to int to avoid errors

            if paginator.num_pages == 0 or page_number > paginator.num_pages:
                return JsonResponse({
                    'success': True,
                    'compensation': [],
                    'page_obj': {
                        'has_next': False,
                        'has_previous': False,
                        'current_page': 1,
                        'total_pages': 1,
                    },
                    'page_range': [],
                    'per_page': per_page,
                })

            page_obj = paginator.get_page(page_number)

            return JsonResponse({
                'success': True,
                'compensation': list(page_obj.object_list),  # Convert to a list
                'page_obj': {
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                    'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
                    'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                },
                'page_range': list(paginator.page_range),
                'per_page': per_page,
            })

        except Exception as e:
            print(e)
            return JsonResponse({'success': False, 'message': str(e)})

    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})

@user_passes_test(is_site_admin, login_url='/site_panel/')
def Site_Edit_empl_comp(request):
    """Fetch data from MySQL and render the edit page with pre-filled values."""
    errors = {}

    try:
        if request.method != 'GET':
            return JsonResponse({"success": False, "message": "Invalid request method!"}, status=400)

        empl_details_json = request.GET.get('empl_details')
        if not empl_details_json:
            return JsonResponse({"success": False, "message": "Missing empl_details parameter!"}, status=400)

        try:
            empl_details_list = json.loads(empl_details_json)  # Parse JSON
            first_item = empl_details_list[0]
            c_id = first_item.get('c_id')
            effective_date = first_item.get('effectiveDate')
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "message": "Invalid JSON format for empl_details!"}, status=400)

        # Ensure the list is not empty
        if not empl_details_list:
            return JsonResponse({"success": False, "message": "empl_details list is empty!"}, status=400)

        # Validate required fields
        if not c_id:
            errors["c_id"] = "Compensation ID (c_id) is required."

        if errors:
            return JsonResponse({"success": False, "errors": errors}, status=400)

        # Fetch compensation data
        try:
            compensation_data = compensation.objects.filter(c_id = c_id).first()
            if not compensation_data:
                return JsonResponse({"success": False, "message": "Compensation data not found!"}, status=404)
        except DatabaseError as e:
            return JsonResponse({"success": False, "message": f"Database error: {str(e)}"}, status=500)

        # Prepare context with pre-filled data
        context = {
            'c_id': c_id,
            'compensation': compensation_data,
        }

        print(effective_date)
        print(compensation_data.e_compensation_ffective_Date)

        return render(request, 'site_panel/Site_Edit_empl_comp.html', context)

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)


@user_passes_test(is_site_admin, login_url='/site_panel/')
def edit_compensation_data(request):
    try:
        errors = {}
        active_inactive = 1
        # Fetch compensation_id
        c_id = request.POST.get('compensation_id')
        if not c_id:
            errors["general"] = "Compensation ID is missing."
        
        e_position = request.POST.get("employee_position")
        pattern2 = re.compile(r"^[A-Za-z]+$")
        if not pattern2.match(e_position):
            errors["employee_position"] = "Invalid Employee's position! Only alphabets are allowed, no spaces."

        e_position_type = request.POST.get("employee_position_type")
        if e_position_type not in ['Front of the House', 'Back of the House', 'Staff/Supervisor', 'Management/Owner']:
            errors["employee_position_type"] = "Invalid position type! Choose a valid option."
        # Fetch and validate end date
        employee_compensation_ffective_Date = request.POST.get("e_compensation_effective_date")
        e_comp_end_date = request.POST.get("employee_compensation_end_date") or None
        
        if employee_compensation_ffective_Date:  # Check if end_date_str is not empty or None
            try:
                e_compensation_ffective_Date=parser.parse(employee_compensation_ffective_Date).date()
            except ValueError:
                return JsonResponse({"success": False, "message": "Invalid Effective Date format!"})
        
        if e_comp_end_date:  # Check if end_date_str is not empty or None
            try:
                e_comp_end_date=parser.parse(e_comp_end_date).date()
            except ValueError:
                return JsonResponse({"success": False, "message": "Invalid End Date format!"})
        else:
            e_comp_end_date = None
        
        if e_comp_end_date and e_comp_end_date!= None and e_comp_end_date <= e_compensation_ffective_Date :
            errors["employee_compensation_end_date"] = "End date cannot be earlier than effective date."

        if e_comp_end_date and e_comp_end_date <= datetime.now().date():
            active_inactive = 0

        if not e_position:
            errors["employee_position"] = "Employee's position is required."
        if not e_position_type:
            errors["employee_position_type"] = "Employee's position-type is required."
    

        # If errors exist, return response
        if errors:
            return JsonResponse({"success": False, "errors": errors}, status=400)

        # Prepare data dictionary
        data = {
            'e_position': e_position,
            'e_position_type': e_position_type,
            'e_comp_end_date': e_comp_end_date,
            'active_inactive': active_inactive,
        }

        # Save data to the database
        try:
            obj, created = compensation.objects.update_or_create(
                c_id=c_id,
                defaults=data
            )
            return JsonResponse({'success': True, 'message': 'Compensation record updated!' if not created else 'Compensation data added successfully!'})

        except IntegrityError as e:
            error_message = str(e)

            if "unique constraint" in error_message.lower():
                errors["general"] = "A unique constraint error occurred. Please check the data."
            elif "foreign key" in error_message.lower():
                errors["general"] = "Foreign key constraint failed. Invalid compensation ID."
            else:
                errors["general"] = "Database integrity error occurred."

        except DataError as e:
            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

        return JsonResponse({"success": False, "errors": errors}, status=400)

    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)

@user_passes_test(is_site_admin, login_url='/site_panel/')
def delete_compensation_data(request):
    try:
        # Log the incoming request data
        try:
            data = json.loads(request.body.decode('utf-8'))
            logger.info("Received request data: %s", data)  # Log request body
        except json.JSONDecodeError:
            logger.error("Invalid JSON data received")
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)

        selected_rows = data.get('data', [])

        if not selected_rows:
            logger.warning("No compensation data provided")
            return JsonResponse({'success': False, 'message': 'No data provided'}, status=400)

        failed_rows = []

        for row in selected_rows:
            try:
                comp_id = row.get('cid')  # Get the employee ID safely

                if comp_id is None:
                    failed_rows.append({'comp_id': None, 'message': 'Missing compensation id'})
                    continue

                comp = compensation.objects.filter(c_id=comp_id).first()
                
                if not comp:
                    failed_rows.append({'comp_id': comp_id, 'message': 'Employee not found'})
                    continue

                # Check if end_date is None
                if comp.e_comp_end_date is None:
                    comp.active_inactive = 0
                    comp.e_comp_end_date = timezone.now().date()  
                else:
                    comp.active_inactive = 0

                comp.save()
                logger.info(f"Updated employee {comp_id} successfully.")

            except IntegrityError as e:
                logger.error(f"Database error: {e}")
                failed_rows.append({'comp_id': comp_id, 'message': f'Database error: {str(e)}'})
            except Exception as e:
                logger.error(f"Error updating employee {comp_id}: {e}")
                failed_rows.append({'comp_id': comp_id, 'message': f'Error: {str(e)}'})

        if failed_rows:
            return JsonResponse({'success': False, 'failed_rows': failed_rows}, status=400)

        return JsonResponse({'success': True, 'message': 'Employees marked as inactive'})

    except Exception as e:
        logger.critical(f"Critical error in Delete_Employee_Data: {e}", exc_info=True)
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'}, status=500)

@user_passes_test(is_site_admin, login_url='/site_panel/')
def site_payroll(request):
    try:
        site_username=request.session.get('user_name')
        
        if not site_username:
            raise ValueError("User session not found. Please log in again.")
        try:
            user = User.objects.get(username=site_username)
            site_id = user.id
        except User.DoesNotExist:
            raise ValueError("User does not exist in the database!")

        
        query = """
        WITH ranked_compensation AS (
            SELECT 
                c.e_id as EmployeeID,
                c.c_id as CompID,
                c.s_id as SiteID,
                c.employee_compensation_name as Name,
                c.e_position AS Position,
                c.e_position_type AS Position_Type,
                c.e_compensation_ffective_Date AS EffectiveDate,
                ROW_NUMBER() OVER (PARTITION BY c.e_id ORDER BY c.e_compensation_ffective_Date DESC) AS rn
                FROM
                    compensation c
                JOIN 
                    employee e ON c.e_id = e.e_id
                WHERE 
                    c.s_id = %s 
                    AND e.active_inactive = %s
        )
        SELECT 
            EmployeeID, 
            CompID, 
            SiteID, 
            Name, 
            Position, 
            Position_Type
        FROM 
            ranked_compensation
        WHERE 
            rn = %s
        ORDER BY 
            EffectiveDate;
            """
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(query, [site_id, 1, 1])
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        except DatabaseError as db_error:
            raise ValueError(f"Database error: {str(db_error)}")

        return render(request, 'site_panel/site_payroll.html', {'compensation_data': results})
    
    except Exception as e:
        # If request is AJAX, return JSON error response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({"success": False, "error": str(e)}, status=400)

        # Otherwise, return error page
        return render(request, 'site_panel/error.html', {"error_message": str(e)})
    
def get_locations(request):
    locations = User.objects.values_list("last_name", flat=True).distinct()
    locations = [location for location in locations if location.lower() != "all"]  # Exclude 'all'
    return JsonResponse({"locations": locations})

def get_working_hours(request):
    try:
        data = json.loads(request.body)
        employees = data.get("employees", [])
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        logger.info(f"Received data: {data}")
        logger.info(f"Start Date: {start_date}, End Date: {end_date}")

        response_data = {}

        for emp in employees:
            empID = emp["empID"]
            logger.info(f"Fetching work hours for empID: {empID}")

            query = """
                SELECT DAYNAME(p.day_date) AS day_name, p.regular_hrs_worked
                FROM site_panel_paydailyentry p 
                LEFT JOIN site_panel_timesheet t ON t.timesheet_id = p.timesheet_id_id
                WHERE t.week_start_date = %s
                AND t.week_end_date = %s
                AND t.e_id_id = %s;
            """

            site = employee.objects.filter(e_id=empID).values_list("s_id", flat=True).first()
            superuser = User.objects.filter(id=site).values_list("is_superuser", flat=True).first() or False


            work_hours = {"Mon": 0, "Tues": 0, "Weds": 0, "Thurs": 0, "Fri": 0, "Sat": 0, "Sun": 0}

            with connection.cursor() as cursor:
                cursor.execute(query, (start_date, end_date, empID))
                results = cursor.fetchall()

                logger.info(f"Query Results for {empID}: {results}")

                day_mapping = {
                    "Monday": "Mon", "Tuesday": "Tues", "Wednesday": "Weds",
                    "Thursday": "Thurs", "Friday": "Fri", "Saturday": "Sat", "Sunday": "Sun"
                }

                for day_name, hours in results:
                    if day_name in day_mapping:
                        work_hours[day_mapping[day_name]] = hours

            response_data[empID] = work_hours
        response_data["superuser"] = superuser
        print(response_data)

        logger.info(f"Final Response: {response_data}")
        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error in get_working_hours: {str(e)}", exc_info=True)
        return JsonResponse({"error": str(e)}, status=500)

def get_admin_payroll_data(request):
    location = request.GET.get("location")

    if not location:
        return JsonResponse({"error": "Location parameter is required"}, status=400)

    # Fetching site_id from User model (assuming last_name represents location)
    site = User.objects.filter(last_name=location).values_list("id", flat=True).first()

    if not site:
        return JsonResponse({"error": "Invalid location"}, status=404)

    query = """
        SELECT 
            c.e_id AS EmployeeID,
            c.c_id AS CompID,
            c.s_id AS SiteID,
            c.employee_compensation_name AS Name,
            c.e_position AS Position,
            c.e_position_type AS Position_Type
        FROM 
            compensation c
        JOIN 
            employee e ON c.e_id = e.e_id
        WHERE
            c.s_id = %s and e.active_inactive = %s
        ORDER BY 
            c.e_compensation_ffective_Date;
    """

    try:
        with connection.cursor() as cursor:
            cursor.execute(query, [site, 1])
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        return JsonResponse(results, safe=False)
    
    except DatabaseError as db_error:
        return JsonResponse({"error": f"Database error: {str(db_error)}"}, status=500)

@require_POST
@user_passes_test(is_site_admin, login_url='/site_panel/')
def save_site_payroll(request):
    try:
        current_time = datetime.now()  # Use timezone-aware datetime

        # Calculate deadlines with timezone awareness
        this_monday = current_time - timedelta(days=current_time.weekday())
        this_monday_2pm = datetime.combine(this_monday.date(), datetime.min.time().replace(hour=14))

        # Extract the new `week_end_date` from the request data
        new_week_end_date = request.POST.get("week_end_date")
        week_start_date = request.POST.get("week_start_date")
        
        if isinstance(new_week_end_date, str):  # Ensure it's a string before conversion
            print(f"Received payroll week end_date: {new_week_end_date}")
            new_week_end_date=parser.parse(new_week_end_date).date()
            print(f"Converted payroll week end_date: {new_week_end_date}")
        if isinstance(week_start_date, str):  # Ensure it's a string before conversion
            print(f"Received payroll week start_date: {week_start_date}")
            week_start_date=parser.parse(week_start_date).date()
            print(f"Converted payroll week start_date: {week_start_date}")

        index = 0
        while f"payroll[{index}][payroll_emp_id]" in request.POST:
            payroll_emp_id = request.POST.get(f"payroll[{index}][payroll_emp_id]")
            notes = request.POST.get(f"payroll[{index}][notes]")

            emp = employee.objects.get(e_id=payroll_emp_id)
            user = User.objects.get(id=emp.s_id)

            if new_week_end_date:
                # Deadline check (timezone-aware)
                if new_week_end_date < this_monday.date() and current_time > this_monday_2pm and user.is_superuser == 0:
                    return JsonResponse(
                        {"status": "error", "message": "Cannot update past payroll data after Monday 2 PM."}, 
                        status=403
                    )

                # Create/update timesheet with timezone-aware datetime
                submission_datetime = datetime.now()  # Fixed timezone issue
                timesheet, created = Timesheet.objects.update_or_create(
                    e_id=emp,
                    s_id=user,
                    week_start_date=week_start_date,
                    week_end_date=new_week_end_date,
                    defaults={"submission_datetime": submission_datetime, "notes": notes}
                )

                # Process daily entries
                for i in range(7):
                    day_date_str = request.POST.get(f"payroll[{index}][daily_hours][{i}][date]")
                    day_date = parser.parse(day_date_str).date()

                    # ✅ Correct compensation lookup with your field names
                    earliest_compensation = compensation.objects.filter(e_id=emp.e_id).order_by('e_compensation_ffective_Date').first()

                    if earliest_compensation.e_compensation_ffective_Date <= day_date:

                    
                        compensations = compensation.objects.filter(
                            e_id=emp.e_id,
                            e_compensation_ffective_Date__lte=day_date
                        ).filter(
                            Q(e_comp_end_date__gte=day_date) | Q(e_comp_end_date__isnull=True)
                        ).order_by('-e_compensation_ffective_Date')
                        print(f"Found compensations: {list(compensations)}")
                        comp = compensations.first()

                        if not comp:
                            return JsonResponse(
                                {
                                    "status": "error", 
                                    "message": f"No compensation record found for Employee ID {emp.e_id} on {day_date}. Ensure a valid compensation record exists."
                                }, 
                                status=400
                            )

                        # Convert hours worked safely
                        try:
                            hours_worked = int(request.POST.get(f"payroll[{index}][daily_hours][{i}][hours_worked]", 0))
                            if hours_worked > 24:
                                return JsonResponse(
                                    {
                                        "status": "error", 
                                        "message": f"Hours worked is less than 24."
                                    }, 
                                    status=400
                                )
                        except ValueError:
                            hours_worked = 0

                        total_daily_rate = hours_worked * comp.e_rate

                        # Update/Create daily entry
                        PayDailyEntry.objects.update_or_create(
                            timesheet_id=timesheet,
                            day_date=day_date,
                            defaults={
                                "c_id": comp,
                                "regular_hrs_worked": hours_worked,
                                "overtime_hr": 0,
                                "daily_hourly_rate": comp.e_rate,
                                "total_daily_rate": total_daily_rate,
                            }
                        )

                index += 1

        return JsonResponse({"status": "success", "message": "Payroll saved!"})
    
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)
       
@user_passes_test(is_site_admin, login_url='/site_panel/')
def Employees_dashboard(request):
    try:
        site_username = request.session.get('user_name')
        if not site_username:
            raise ValueError("User session not found. Please log in again.")

        try:
            user = User.objects.get(username=site_username)
            site_id = user.id
        except User.DoesNotExist:
            raise ValueError("User does not exist in the database!")

        try:
            with connection.cursor() as cursor:
                current_date = datetime.now().date()
                query_param = [0, current_date, site_id]  # 0 = inactive, current_date, site_id
                query = """
                UPDATE employee
                SET active_inactive = %s
                WHERE end_date <= %s AND s_id = %s;
                """
                cursor.execute(query, query_param)
                
        except Exception as e:
            print(f"Error updating employee status: {e}")

        base_query = """
            SELECT 
                e.e_id as "EmployeeID",
                c.e_position as "Position",
                e.location AS "Location",
                e.owner AS "Owner",
                e.employee_name AS "Ename",
                e.start_date AS "StartDate",
                e.end_date AS "EndDate",
                e.phone1 AS "Phone1",
                e.phone2 AS "Phone2",
                e.smart_serve AS "SmartServer",
                e.active_inactive AS "active_inactive"
            FROM 
                employee e
            LEFT JOIN
                compensation c ON c.e_id = e.e_id
        """

        query_params = []
        if user.last_name != 'all':
            base_query += " WHERE e.s_id = %s ORDER BY e.start_date;"
            query_params.append(site_id)
        elif user.last_name == 'all':
            site_selected_location = request.GET.get('site_selected_location')
            if site_selected_location:
                base_query += " WHERE e.location = %s ORDER BY e.s_id, e.start_date;"
                query_params.append(site_selected_location)
            else:
                base_query += " ORDER BY e.s_id, e.start_date;"

        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]
        except DatabaseError as db_error:
            raise ValueError(f"Database error: {str(db_error)}")

        per_page = int(request.GET.get('per_page', 1))
        page_number = int(request.GET.get('page', 1))

        paginator = Paginator(results, per_page)

        try:
            page_obj = paginator.get_page(page_number)
        except EmptyPage:
            page_obj = paginator.get_page(1)

        # Handle AJAX requests
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            data = {
                'employee_data': list(page_obj.object_list),
                'page': page_obj.number,
                'num_pages': paginator.num_pages,
                'per_page': per_page,
                'no_results': not results,
            }
            return JsonResponse(data)

        # Handle non-AJAX requests (render the full HTML page)
        return render(request, 'site_panel/site_employee_dashboard.html', {
            'employee_data': page_obj.object_list,
            'page_obj': page_obj,
            'page_range': paginator.page_range,
            'per_page': per_page,
            'no_results': not results,
        })

    except Exception as e:
        # Handle errors
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({"success": False, "error": str(e)}, status=400)
        return render(request, 'site_panel/error.html', {"error_message": str(e)})

@csrf_exempt
@require_POST
def update_employee_status(request):
    try:
        data = json.loads(request.body)
        emp_id = data.get('empid')
        current_status = data.get('active_inactive')
        
        
        if emp_id is None and current_status is None:
            return JsonResponse({"success": False, "message": "Invalid request data"}, status=400)

        if current_status == 1:
            emp = employee.objects.get(e_id=emp_id)
            emp.active_inactive = current_status
            emp.end_date = None
            emp.save()
            return JsonResponse({"success": True, "message": f"{emp.employee_name} employee is now active!"})
        elif current_status == 0:
            emp = employee.objects.get(e_id=emp_id)
            emp.active_inactive = current_status
            

            failed_rows = []
                
            if not emp:
                failed_rows.append({'emp_id': emp_id, 'message': 'Employee not found'})

            if failed_rows:
                return JsonResponse({'success': False, 'failed_rows': failed_rows}, status=400)
            
            if emp.end_date is None:
                emp.end_date = timezone.now().date()  

            emp.save()

            return JsonResponse({"success": True, "message": f"{emp.employee_name} employee is now inactive!"})
    except Exception as e:
        return JsonResponse({"success": False, "message": str(e)}, status=500)

@user_passes_test(is_site_admin, login_url='/site_panel/')
def Site_New_Daily_Activity_Tracker(request):
        return render(request, 'site_panel/Site_New_Daily_Activity_Tracker.html')

@require_POST  # Use the decorator for POST-only views
@user_passes_test(is_site_admin, login_url='/site_panel/')
def Edit_Site_Daily_Activity_Tracker(request):
    siteID = request.POST.get('site_id')
    site_Date = request.POST.get('site_date')
    siteDate = request.POST.get('date')
    permission = request.POST.get('permission')

    check_date = timezone.now().date() - timedelta(days=7)

    request.session['site_id'] = siteID
    request.session['site_date'] = str(siteDate)  # Ensure consistency in session storage
    request.session['permission'] = str(permission)  # Ensure consistency in session storage

    print('site ID:', siteID)
    print('site Date:', siteDate)

    try:
        # Convert siteDate (string) to a date object
        site_Date = datetime.strptime(site_Date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    return JsonResponse({'redirect_url': '/site_panel/edit_daily_activity_record/'}, status=200)

@user_passes_test(is_site_admin, login_url='/site_panel/')
def Site_Edit_Daily_Activity_Tracker(request):
    """Fetch data from MySQL and render the edit page with pre-filled values."""

    siteID = request.session.get('site_id')
    siteDate = request.session.get('site_date')
    permission = request.session.get('permission')

    if not siteID or not siteDate:
        return redirect('/site_panel/')  # Redirect to home if session data is missing
    
    try:
        siteDate = parser.parse(siteDate).date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)


    # Fetching sales data
    sales_data = ActivitySales.objects.filter(site_id=siteID, activity_date=siteDate).first()

    # Fetching promotions data
    promotions_data = ActivityPromotions.objects.filter(site_id=siteID, activity_date=siteDate).first()
    financial_data = ActivityFinancial.objects.filter(site_id=siteID, activity_date=siteDate).first()
    purchase_data = ActivityPurchase.objects.filter(site_id=siteID, activity_date=siteDate).first()
    paid_out_data = ActivityPaidOut.objects.filter(site_id=siteID, activity_date=siteDate).first()

    # Prepare context with pre-filled data
    context = {
        'site_id': siteID,
        'site_date': siteDate,
        'sales': sales_data,  
        'promotions': promotions_data, 
        'financial': financial_data,
        'purchase': purchase_data, 
        'paid_out': paid_out_data
    }

    if permission == 'allowed':
        return render(request, 'site_panel/Site_Edit_Daily_Activity_Tracker.html', context)
    
    else:
        return render(request, 'site_panel/Site_Edit_Daily_Activity_Tracker_not_allowed.html', context)

def delete_site_daily_activity_tracker(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
        selected_rows = data.get('data', [])

        if not selected_rows:
            return JsonResponse({'success': False, 'message': 'No data provided'}, status=400)

        for row in selected_rows:
            try:  # Inner try-except for each row
                site_id = int(row.get('siteID'))
                site_date_str = row.get('siteDate')
                print(f"Received date: {site_date_str}")
                site_date=parser.parse(site_date_str).date()
                print(f"Converted date: {site_date}")

                ActivitySales.objects.filter(site_id=site_id, activity_date=site_date).delete()
                ActivityPromotions.objects.filter(site_id=site_id, activity_date=site_date).delete()
                ActivityFinancial.objects.filter(site_id=site_id, activity_date=site_date).delete()
                ActivityPurchase.objects.filter(site_id=site_id, activity_date=site_date).delete()
                ActivityPaidOut.objects.filter(site_id=site_id, activity_date=site_date).delete()

            except (ValueError, TypeError) as e: # Handle data type errors
                return JsonResponse({'success': False, 'message': f'Invalid data: {str(e)} for row: {row}'}, status=400)
            except IntegrityError as e: # Handle foreign key or other database integrity errors
                return JsonResponse({'success': False, 'message': f'Database integrity error: {str(e)} for row: {row}'}, status=500)
            except Exception as e:  # Catch all other exceptions for each row
                return JsonResponse({'success': False, 'message': f'An error occurred during deletion: {str(e)} for row: {row}'}, status=500)

        return JsonResponse({'success': True})

    except json.JSONDecodeError as e:
        return JsonResponse({'success': False, 'message': f'Invalid JSON data: {str(e)}'}, status=400)
    except Exception as e:  # Catch errors during JSON parsing
        return JsonResponse({'success': False, 'message': f'An error occurred: {str(e)}'}, status=500)
        
@require_POST   
def Site_Insert_new_activity_data(request):
    global sales_backend_total, sales_tax1, sales_tax2, liquor_tax_1, promotion_total
    try:
        site_username = request.session.get('user_name')
    
        if not site_username:
            return JsonResponse({'success': False, 'message': 'User session not found!'}, status=400)

        try:
            user = User.objects.get(username=site_username)
            site_id = user.id
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'User not found!'}, status=400)

        form_name = request.POST.get('form_name')
        if form_name == 'sales':
                errors = {}
                today = datetime.today().date()

                # Extract and validate activity date
                activity_date = request.POST.get('activity_date')
                request.session['activity_date'] = activity_date

                if not activity_date:
                    errors["activity_date"] = "Activity date is required."
                else:
                    try:
                        activity_date = parser.parse(activity_date).date()
                        if activity_date and activity_date < today - timedelta(weeks=1):
                            errors["activity_date"] = "Start Date cannot be more than a week old!"
                    except (ValueError, TypeError):
                        errors["activity_date"] = "Invalid date format for activity date."

                # Extract and validate critical fields
                try:
                    customer_count = int(request.POST.get('customer_count', 0))
                    sales_total = float(request.POST.get('sales_total', 0))
                    honest_to_goodness_fees = float(request.POST.get('honest_to_goodness_fees_sales', 0))
                except ValueError:
                    errors["general"] = "Invalid number format in one of the required fields."

                if customer_count <= 0:
                    errors["customer_count"] = "Customer count must be greater than zero."

                if sales_total <= 0:
                    errors["sales_total"] = "Sales total must be greater than zero."

                # Check if the record already exists
                if ActivitySales.objects.filter(Q(site_id=site_id) & Q(activity_date=activity_date)).exists():
                    return JsonResponse({'success': False, 'message': 'The record already exists!'}, status=400)

                # Return errors if any
                if errors:
                    return JsonResponse({"success": False, "errors": errors}, status=400)

                # Save sales data
                try:
                    food = float(request.POST.get('food_sales', 0) or 0)
                    liquor = float(request.POST.get('liquor_sales', 0) or 0)
                    beer = float(request.POST.get('beer_sales', 0) or 0)
                    wine = float(request.POST.get('wine_sales', 0) or 0)
                    beverages_iced_beverages = float(request.POST.get('beverages_iced_beverages_sales', 0) or 0)
                    draft = float(request.POST.get('draft_sales', 0) or 0)
                    honest_to_goodness_fees = float(request.POST.get('honest_to_goodness_fees_sales', 0) or 0)
                    miscellaneous_income = float(request.POST.get('miscellaneous_income_sales', 0) or 0)

                    # Now perform the addition safely
                    
                    sales_backend_total = (food + liquor + beer + wine + beverages_iced_beverages +
                                        draft + honest_to_goodness_fees + miscellaneous_income)
                    
                    sales_tax1 = float(request.POST.get('sales_tax1', 0) or 0)
                    sales_tax2 = float(request.POST.get('sales_tax2', 0) or 0)
                    liquor_tax_1 = float(request.POST.get('liquor_tax_1', 0) or 0)
                    
                    activity_sales = ActivitySales(
                        site_id=site_id,  # Foreign key
                        activity_date=activity_date,
                        food= food,
                        liquor= liquor,
                        beer= beer,
                        wine= wine,
                        beverages_iced_beverages=request.POST.get('beverages_iced_beverages_sales') or 0,
                        draft= draft,
                        honest_to_goodness_fees=honest_to_goodness_fees,
                        miscellaneous_income= miscellaneous_income,
                        sales_total= sales_backend_total,
                        sales_tax1= sales_tax1,
                        sales_tax2= sales_tax2,
                        liquor_tax_1= liquor_tax_1,
                        customer_count=customer_count,
                        sales_notes=request.POST.get('notes') or "",
                    )

                    activity_sales.save()

                    return JsonResponse({
                        'success': True,
                        'message': 'Sales data added successfully!',
                        'form_name': 'sales',
                        'data': {
                            'site_id': site_id,
                            'activity_date': activity_date.strftime('%Y-%m-%d'),
                            'customer_count': customer_count,
                            'food_sales': food,
                            'liquor_sales': liquor,
                            'beer_sales': beer,
                            'wine_sales': wine,
                            'beverages_iced_beverages_sales': beverages_iced_beverages,
                            'draft_sales': draft,
                            'honest_to_goodness_fees_sales': honest_to_goodness_fees,
                            'miscellaneous_income_sales': miscellaneous_income,
                            'sales_backend_total': sales_backend_total,
                            'sales_tax1': sales_tax1,
                            'sales_tax2': sales_tax2,
                            'liquor_tax_1': liquor_tax_1,
                            'sales_notes': request.POST.get('notes', ''),
                        }
                    })


                except IntegrityError as e:
                    error_message = str(e)
                    if "unique constraint" in error_message.lower():
                        errors["general"] = "A unique constraint error occurred."
                    elif "foreign key" in error_message.lower():
                        errors["general"] = "Invalid site_id reference. Ensure the site exists."
                    else:
                        errors["general"] = "Database integrity error."

                except DataError as e:
                    errors["general"] = f"Invalid data error: {str(e)}"

                if errors:
                    return JsonResponse({"success": False, "errors": errors}, status=400)

                return JsonResponse({'success': False, 'message': 'Invalid form submission!'}, status=400)


        elif form_name == 'Promotion':
            errors = {}

            # Ensure `activity_date` is available from session
            activity_date = request.session.get('activity_date')
            if not activity_date:
                errors["promotions-general"] = "Activity date is required."

            # Validate `site_id`
            if not site_id:
                errors["promotions-general"] = "Site ID is missing or invalid."

            # Check if the record already exists
            if ActivityPromotions.objects.filter(Q(site_id=site_id) & Q(activity_date=activity_date)).exists():
                return JsonResponse({'success': False, 'message': 'The record already exists!'}, status=400)

            # Convert `promotion_total` and `lsm_2` to numeric values safely
            try:
                lsm_2 = float(request.POST.get('promotions_lsm_2', 0))
                promotion_total = float(request.POST.get('promotion_total', 0))
            except ValueError:
                errors["promotions-general"] = "Invalid numeric value in Promotion Total or LSM 2."

            # Return errors if any
            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

            # Logging for debugging
            logger.info(f"Received Promotion Data: {request.POST}")

            # Save promotion data
            try:
                
                staff_meals=float(request.POST.get('promotions_staff_meals') or 0)
                manager_meals=float(request.POST.get('promotion_manager_meals') or 0)
                coupons_disc=float(request.POST.get('promotion_coupons_disc') or 0)
                chucks_bucks=float(request.POST.get('promotion_chucks_bucks') or 0)
                manager_prom=float(request.POST.get('promotion_manager_prom') or 0)
                qsa_complaints=float(request.POST.get('promotion_qsa_complaints') or 0)
                lsm_1=float(request.POST.get('promotion_lsm_1') or 0)
                promotion_total = staff_meals + manager_meals + coupons_disc + chucks_bucks + manager_prom + qsa_complaints + lsm_1 + lsm_2

                activity_promotions = ActivityPromotions(
                    site_id=site_id,
                    activity_date=activity_date,
                    staff_meals= staff_meals,
                    manager_meals=manager_meals,
                    coupons_disc=coupons_disc,
                    chucks_bucks=chucks_bucks,
                    manager_prom=manager_prom,
                    qsa_complaints=qsa_complaints,
                    lsm_1=lsm_1,
                    lsm_2=lsm_2,
                    promotion_total=promotion_total
                )
                activity_promotions.save()
                return JsonResponse({'success': True, 
                                     'message': 'Promotion record created successfully!',
                                     'form_name': 'Promotion',
                                     'data': {
                                         'site_id': site_id,
                                         'staff_meals': staff_meals,
                                         'manager_meals': manager_meals,
                                         'coupons_disc': coupons_disc,
                                         'chucks_bucks': chucks_bucks,
                                         'manager_prom': manager_prom,
                                         'qsa_complaints': qsa_complaints,
                                         'lsm_1' : lsm_1,
                                         'lsm_2': lsm_2,
                                         'promotion_total': promotion_total,
                                     }})

            except IntegrityError as e:
                errors["promotions-general"] = "A database integrity error occurred. Check for duplicate values."
            except DataError as e:
                errors["promotions-general"] = f"Invalid data error: {str(e)}"

            # Return errors if any
            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

        
        elif form_name == 'financial':
            errors = {}

            # Extract site_id
            if not site_id:
                return JsonResponse({'success': False, 'message': 'User session not found!'}, status=400)

            # Extract and validate activity_date
            activity_date = request.session.get('activity_date')
            if not activity_date:
                errors["financial-general"] = "Activity date is required."
            else:
                try:
                    activity_date = parser.parse(activity_date).date()
                except (ValueError, TypeError):
                    errors["financial-general"] = "Invalid date format for activity date."

            # Check if record already exists
            if ActivityFinancial.objects.filter(Q(site_id=site_id) & Q(activity_date=activity_date)).exists():
                return JsonResponse({'success': False, 'message': 'The record already exists!'}, status=400)

            # Log received data
            logger.info(f"Received Financial Data: {request.POST}")

            # Extract financial data (other fields are optional, so use default values if missing)
            gift_card_purchased = float(request.POST.get('financial_gift_card_purchased') or 0)
            gift_card_redeemed = float(request.POST.get('financial_gift_card_redeemed') or 0)
            skip_the_dishes = float(request.POST.get('financial_skip_the_dishes') or 0)
            just_eat = float(request.POST.get('financial_just_eat') or 0)
            amex = float(request.POST.get('financial_amex') or 0)
            print(amex)
            din_club = float(request.POST.get('financial_din_club') or 0)
            discover = float(request.POST.get('financial_discover') or 0)
            master_card = float(request.POST.get('financial_master_card') or 0)
            visa = float(request.POST.get('financial_visa') or 0)
            dr_card = float(request.POST.get('financial_dr_card') or 0)
            total = skip_the_dishes + just_eat + amex + din_club + discover + master_card + visa + dr_card
            actual_deposit_amt = float(request.POST.get('actual_deposit_amt') or 0)
            required_deposit = sales_backend_total + sales_tax1 + sales_tax2 + liquor_tax_1 - promotion_total + gift_card_purchased - (gift_card_redeemed + skip_the_dishes + just_eat + amex + din_club + discover + master_card + visa + dr_card)

            # Validate numerical fields
            try:
                total = float(total)
                actual_deposit_amt = float(actual_deposit_amt)
                required_deposit = float(required_deposit)
            except ValueError:
                errors["financial-general"] = "Invalid number format in one of the financial fields."

            # Return errors if any
            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

            # Save financial data
            try:
                print()
                activity_financial = ActivityFinancial(
                    site_id=site_id,
                    activity_date=activity_date,
                    gift_card_purchased=gift_card_purchased,
                    gift_card_redeemed=gift_card_redeemed,
                    skip_the_dishes=skip_the_dishes,
                    just_eat=just_eat,
                    amex=amex,
                    din_club=din_club,
                    discover=discover,
                    master_card=master_card,
                    visa=visa,
                    dr_card=dr_card,
                    financial_total=total,
                    actual_deposit_amt=actual_deposit_amt,
                    required_deposit=required_deposit,
                )
                activity_financial.save()
                return JsonResponse({'success': True, 
                                     'message': 'Financial record created successfully!',
                                     'form_name': 'financial',
                                     'data': {
                                        'site_id': site_id,
                                        'activity_date': activity_date,
                                        'gift_card_purchased': gift_card_purchased,
                                        'gift_card_redeemed': gift_card_redeemed,
                                        'skip_the_dishes': skip_the_dishes,
                                        'just_eat': just_eat,
                                        'amex': amex,
                                        'din_club': din_club,
                                        'discover': discover,
                                        'master_card': master_card,
                                        'visa': visa,
                                        'dr_card': dr_card,
                                        'financial_total': total,
                                        'actual_deposit_amt': actual_deposit_amt,
                                        'required_deposit': required_deposit, 
                                     }
                                     })

            except IntegrityError as e:
                errors["financial-general"] = "Database integrity error."
            except DataError as e:
                errors["financial-general"] = f"Invalid data error: {str(e)}"

            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

        
        elif form_name == 'Purchase':
            errors = {}

            # Extract and validate activity date
            activity_date = request.session.get('activity_date')
            
            if not activity_date:
                errors["purchases-general"] = "Activity date is required."
            else:
                try:
                    activity_date = parser.parse(activity_date).date()
                except (ValueError, TypeError):
                    errors["purchases-general"] = "Invalid date format for activity date."

            # Extract optional fields
            purchases_food = request.POST.get('purchases_food') or 0
            purchases_paper = request.POST.get('purchases_paper') or 0
            purchases_liquor = request.POST.get('purchases_liquor') or 0

            logger.info(f"Received Purchase Data: {request.POST}")

            # Check if the record already exists
            if ActivityPurchase.objects.filter(Q(site_id=site_id) & Q(activity_date=activity_date)).exists():
                return JsonResponse({'success': False, 'message': 'The record already exists!'}, status=400)

            # Return errors if any
            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

            # Save the ActivityPurchase record
            try:
                activity_purchase = ActivityPurchase(
                    site_id=site_id,  # Foreign key
                    activity_date=activity_date,
                    purchases_food=purchases_food,
                    purchases_paper=purchases_paper,
                    purchases_liquor=purchases_liquor,
                )

                activity_purchase.save()
                return JsonResponse({'success': True, 'message': 'Purchase record created successfully!'})

            except IntegrityError as e:
                error_message = str(e)
                if "foreign key" in error_message.lower():
                    errors["purchases-general"] = "Invalid site_id reference. Ensure the site exists."
                else:
                    errors["purchases-general"] = "Database integrity error."

            except DataError as e:
                errors["purchases-general"] = f"Invalid data error: {str(e)}"

            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

            return JsonResponse({'success': False, 'message': 'Invalid form submission!'}, status=400)
                
        elif form_name == 'paid-out':
            errors = {}

            # Extract site ID and activity date
            activity_date = request.session.get('activity_date')

            if not activity_date:
                errors["paidout-general"] = "Activity date is required."
            else:
                try:
                    activity_date = parser.parse(activity_date).date()
                except (ValueError, TypeError):
                    errors["paidout-general"] = "Invalid date format for activity date."

            # Check if record already exists for the given site and date
            if ActivityPaidOut.objects.filter(Q(site_id=site_id) & Q(activity_date=activity_date)).exists():
                return JsonResponse({'success': False, 'message': 'The record already exists!'}, status=400)

            # Extract and validate numeric fields
            try:
                food_paid_out = float(request.POST.get('food_paid_out') or 0)
                liquor_paid_out = float(request.POST.get('liquor_paid_out') or 0)
                supplies_paid_out = float(request.POST.get('supplies_paid_out') or 0)
                repair_maintenance_paid_out = float(request.POST.get('repair_maintenance_paid_out') or 0)
                advertising_paid_out = float(request.POST.get('advertising_paid_out') or 0)
                entertainment_paid_out = float(request.POST.get('entertainment_paid_out') or 0)
                others_paid_out = float(request.POST.get('others_paid_out') or 0)
                hst_gst_paid_out = float(request.POST.get('hst_gst_paid_out') or 0)
                paid_out_total = food_paid_out + liquor_paid_out + supplies_paid_out + repair_maintenance_paid_out + advertising_paid_out + entertainment_paid_out + others_paid_out + hst_gst_paid_out
            except ValueError:
                errors["paidout-general"] = "Invalid number format in paid-out fields."

            # Return errors if any
            if errors:
                return JsonResponse({"success": False, "errors": errors}, status=400)

            # Save the paid-out record
            try:
                activity_paid_out = ActivityPaidOut(
                    site_id=site_id,  # Foreign key
                    activity_date=activity_date,
                    food_paid_out=food_paid_out,
                    liquor_paid_out=liquor_paid_out,
                    supplies_paid_out=supplies_paid_out,
                    repair_maintenance=repair_maintenance_paid_out,
                    advertising=advertising_paid_out,
                    entertainment=entertainment_paid_out,
                    others=others_paid_out,
                    hst_gst=hst_gst_paid_out,
                    paid_out_total=paid_out_total,
                )
                activity_paid_out.save()
                return JsonResponse({'success': True, 
                                     'message': 'Paid-out record created successfully!',
                                     'form_name': 'paid-out',
                                     'data':
                                     {
                                        'site_id': site_id,  # Foreign key
                                        'activity_date': activity_date,
                                        'food_paid_out': food_paid_out,
                                        'liquor_paid_out': liquor_paid_out,
                                        'supplies_paid_out': supplies_paid_out,
                                        'repair_maintenance': repair_maintenance_paid_out,
                                        'advertising': advertising_paid_out,
                                        'entertainment': entertainment_paid_out,
                                        'others': others_paid_out,
                                        'hst_gst': hst_gst_paid_out,
                                        'paid_out_total': paid_out_total,
                                     }})

            except IntegrityError as e:
                return JsonResponse({"success": False, "message": f"Database error: {str(e)}"}, status=400)

            except DataError as e:
                return JsonResponse({"success": False, "message": f"Invalid data error: {str(e)}"}, status=400)
        
    except Exception as e:
        logger.error(f"Error in Promotion Form: {str(e)}", exc_info=True)  # Log the full error traceback
        return JsonResponse({'success': False, 'message': f"Error: {str(e)}"}, status=500)
        
@require_POST
def site_update_activity_data(request):
        global sales_backend_total, sales_tax1, sales_tax2, liquor_tax_1, promotion_total
        try:
            form_name = request.POST.get('form_name')
            
            if form_name == 'sales':
                errors = {}

                try:
                    sales_id = request.POST.get('sales_id')  # Unique ID for updating records
                    if not sales_id:
                        errors["general"] = "Not able ton extract the sales id."

                    site_id = request.POST.get('sales_site_id')
                    if not site_id:
                        errors["general"] = "Not able ton extract the site id."

                    # Required field validation
                    sales_total = request.POST.get('sales_total')
                    if not sales_total:
                        errors["sales_total"] = "Missing sales_total field."

                    activity_date = request.POST.get('activity_date')
                    if not activity_date:
                        errors["activity_date"] = "Activity date is required."

                    customer_count = request.POST.get('customer_count')
                    if customer_count and not customer_count.isdigit():
                        errors["customer_count"] = "Customer count must be a number."

                    # If errors exist, return response
                    if errors:
                        return JsonResponse({"success": False, "errors": errors}, status=400)

                    # Parse date field
                    try:
                        activity_date = parser.parse(activity_date).date()
                    except ValueError:
                        return JsonResponse({"success": False, "message": "Invalid Activity Date format!"}, status=400)
                    
                    food = float(request.POST.get('food_sales', 0) or 0)
                    liquor = float(request.POST.get('liquor_sales', 0) or 0)
                    beer = float(request.POST.get('beer_sales', 0) or 0)
                    wine = float(request.POST.get('wine_sales', 0) or 0)
                    beverages_iced_beverages = float(request.POST.get('beverages_iced_beverages_sales', 0) or 0)
                    draft = float(request.POST.get('draft_sales', 0) or 0)
                    honest_to_goodness_fees = float(request.POST.get('honest_to_goodness_fees_sales', 0) or 0)
                    miscellaneous_income = float(request.POST.get('miscellaneous_income_sales', 0) or 0)

                    # Now perform the addition safely
                    
                    sales_backend_total = (food + liquor + beer + wine + beverages_iced_beverages +
                                        draft + honest_to_goodness_fees + miscellaneous_income)
                    
                    sales_tax1 = float(request.POST.get('sales_tax1', 0) or 0)
                    sales_tax2 = float(request.POST.get('sales_tax2', 0) or 0)
                    liquor_tax_1 = float(request.POST.get('liquor_tax_1', 0) or 0)

                    data = {
                        'activity_date': activity_date,
                        'food': food,
                        'liquor': liquor,
                        'beer': beer,
                        'wine': wine,
                        'beverages_iced_beverages': beverages_iced_beverages,
                        'draft': draft,
                        'honest_to_goodness_fees': honest_to_goodness_fees,
                        'miscellaneous_income': miscellaneous_income,
                        'sales_total': sales_backend_total,
                        'sales_tax1': sales_tax1,
                        'sales_tax2': sales_tax2,
                        'liquor_tax_1': liquor_tax_1,
                        'customer_count': customer_count,
                        'sales_notes': request.POST.get('notes'),
                    }

                    # Save data to the database
                    try:
                        updated_count = ActivitySales.objects.filter(activity_date=activity_date, site_id=site_id).update(**data)
                        if updated_count:
                            return JsonResponse({'success': True, 
                                                 'message': 'Sales record updated!',
                                                 'form_name': 'sales',
                                                 'data': {
                                                    'customer_count': customer_count,
                                                    'food_sales': food,
                                                    'liquor_sales': liquor,
                                                    'beer_sales': beer,
                                                    'wine_sales': wine,
                                                    'beverages_iced_beverages_sales': beverages_iced_beverages,
                                                    'draft_sales': draft,
                                                    'honest_to_goodness_fees_sales': honest_to_goodness_fees,
                                                    'miscellaneous_income_sales': miscellaneous_income,
                                                    'sales_backend_total': sales_backend_total,
                                                    'sales_tax1': sales_tax1,
                                                    'sales_tax2': sales_tax2,
                                                    'liquor_tax_1': liquor_tax_1,
                                                    'sales_notes': request.POST.get('notes', ''),
                                                }
                                                 })

                    except IntegrityError as e:
                        error_message = str(e)

                        if "unique constraint" in error_message.lower():
                            errors["general"] = "A unique constraint error occurred. Please check the data."
                        elif "foreign key" in error_message.lower():
                            errors["general"] = "Foreign key constraint failed. Invalid site ID or sales ID."
                        else:
                            errors["general"] = "Database integrity error occurred."

                    except DataError as e:
                        return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                    return JsonResponse({"success": False, "errors": errors}, status=400)

                except Exception as e:
                    return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)


            elif form_name == 'Promotion':
                errors = {}

                # Get activity_date from request
                activity_date = request.POST.get('promotions_activity_date')  
                if not activity_date:
                    errors["activity_date"] = "Activity date is required."

                # Get site_id from session
                site_id = request.session.get('site_id')
                if not site_id:
                    errors["general"] = "Not able to extract the site ID from the session."

                # Required field validation
                promotion_total = request.POST.get('promotion_total')
                if not promotion_total:
                    errors["promotion_total"] = "Missing promotion_total field."

                # If any errors, return immediately
                if errors:
                    return JsonResponse({"success": False, "errors": errors}, status=400)

                # Extract other fields from the request
                staff_meals=float(request.POST.get('promotions_staff_meals') or 0)
                manager_meals=float(request.POST.get('promotion_manager_meals') or 0)
                coupons_disc=float(request.POST.get('promotion_coupons_disc') or 0)
                chucks_bucks=float(request.POST.get('promotion_chucks_bucks') or 0)
                manager_prom=float(request.POST.get('promotion_manager_prom') or 0)
                qsa_complaints=float(request.POST.get('promotion_qsa_complaints') or 0)
                lsm_1=float(request.POST.get('promotion_lsm_1') or 0)
                lsm_2=float(request.POST.get('promotion_lsm_2') or 0)
                promotion_total = staff_meals + manager_meals + coupons_disc + chucks_bucks + manager_prom + qsa_complaints + lsm_1 + lsm_2

                promotion_data = {
                    'staff_meals': staff_meals,
                    'manager_meals': manager_meals,
                    'coupons_disc': coupons_disc,
                    'chucks_bucks': chucks_bucks,
                    'manager_prom': manager_prom,
                    'qsa_complaints': qsa_complaints,
                    'lsm_1' : lsm_1,
                    'lsm_2': lsm_2,
                    'promotion_total': promotion_total,
                }

                try:
                    # Check if the record exists
                    existing_promotion = ActivityPromotions.objects.filter(activity_date=activity_date, site_id=site_id).first()

                    if not existing_promotion:
                        # If no record exists, create a new one
                        try:
                            ActivityPromotions.objects.create(
                                activity_date=activity_date,
                                site_id=site_id,
                                **promotion_data
                            )
                            return JsonResponse({'success': True, 
                                                 'message': 'Promotions data added successfully!',
                                                 'form_name': 'Promotion',
                                     'data': {
                                         'staff_meals': staff_meals,
                                         'manager_meals': manager_meals,
                                         'coupons_disc': coupons_disc,
                                         'chucks_bucks': chucks_bucks,
                                         'manager_prom': manager_prom,
                                         'qsa_complaints': qsa_complaints,
                                         'lsm_1' : lsm_1,
                                         'lsm_2': lsm_2,
                                         'promotion_total': promotion_total,
                                     }})
                        
                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["general"] = "Foreign key constraint failed. Invalid site ID or promotions ID."
                            else:
                                errors["general"] = "Database integrity error occurred."
                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                    else:
                        # If record exists, update it
                        try:
                            obj, created = ActivityPromotions.objects.update_or_create(
                                    activity_date=activity_date, site_id=site_id,
                                    defaults=promotion_data
                                )

                            return JsonResponse({'success': True, 
                                                 'message': 'Promotions record updated successfully!',
                                                 'form_name': 'Promotion',
                                     'data': {
                                         'staff_meals': staff_meals,
                                         'manager_meals': manager_meals,
                                         'coupons_disc': coupons_disc,
                                         'chucks_bucks': chucks_bucks,
                                         'manager_prom': manager_prom,
                                         'qsa_complaints': qsa_complaints,
                                         'lsm_1' : lsm_1,
                                         'lsm_2': lsm_2,
                                         'promotion_total': promotion_total,
                                     }})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["general"] = "Foreign key constraint failed. Invalid site ID or promotions ID."
                            else:
                                errors["general"] = "Database integrity error occurred."
                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                    return JsonResponse({"success": False, "errors": errors}, status=400)
                
                except Exception as e:
                    return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)
                
            elif form_name == 'financial':
                errors = {}

                # Get activity_date from request
                activity_date = request.POST.get('financial_activity_date')  # Unique ID for updating records
                if not activity_date:
                    errors["financial-general"] = "Activity date is required."
                try:
                    activity_date = parser.parse(activity_date).date()
                except ValueError:
                    return JsonResponse({"success": False, "message": "Invalid Activity Date format!"}, status=400)

                site_id = request.POST.get('financial_site_id')
                if not site_id:
                    errors["financial-general"] = "Not able to extract the site ID."

                # Required field validation
                financial_total = request.POST.get('financial_total')
                if not financial_total:
                    errors["financial_total"] = "Missing financial_total field."

                actual_deposit_amt = request.POST.get('actual_deposit_amt')

                required_deposit = request.POST.get('required_deposit')
                if required_deposit and not required_deposit.lstrip('-').replace('.', '', 1).isdigit():
                    errors["required_deposit"] = "Required deposit must be a valid number."

                # If errors exist, return response
                if errors:
                    return JsonResponse({"success": False, "errors": errors}, status=400)
                
                sales_backend_total = float(request.POST.get('sales_total') or 0)
                sales_tax1 = float(request.POST.get('sales_tax1') or 0)
                sales_tax2 = float(request.POST.get('sales_tax2') or 0)
                liquor_tax_1 = float(request.POST.get('liquor_tax_1') or 0)
                promotion_total = float(request.POST.get('promotion_total') or 0)
                
                gift_card_purchased = float(request.POST.get('financial_gift_card_purchased') or 0)
                gift_card_redeemed = float(request.POST.get('financial_gift_card_redeemed') or 0)
                skip_the_dishes = float(request.POST.get('financial_skip_the_dishes') or 0)
                just_eat = float(request.POST.get('financial_just_eat') or 0)
                amex = float(request.POST.get('financial_amex') or 0)
                din_club = float(request.POST.get('financial_din_club') or 0)
                discover = float(request.POST.get('financial_discover') or 0)
                master_card = float(request.POST.get('financial_master_card') or 0)
                visa = float(request.POST.get('financial_visa') or 0)
                dr_card = float(request.POST.get('financial_dr_card') or 0)
                total = skip_the_dishes + just_eat + amex + din_club + discover + master_card + visa + dr_card
                actual_deposit_amt = float(request.POST.get('actual_deposit_amt') or 0)
                required_deposit = sales_backend_total + sales_tax1 + sales_tax2 + liquor_tax_1 - promotion_total + gift_card_purchased - (gift_card_redeemed + skip_the_dishes + just_eat + amex + din_club + discover + master_card + visa + dr_card)

                data = {
                        'gift_card_purchased': gift_card_purchased,
                        'gift_card_redeemed': gift_card_redeemed,
                        'skip_the_dishes': skip_the_dishes,
                        'just_eat': just_eat,
                        'amex': amex,
                        'din_club': din_club,
                        'discover': discover,
                        'master_card': master_card,
                        'visa': visa,
                        'dr_card': dr_card,
                        'financial_total': total,
                        'actual_deposit_amt': actual_deposit_amt,
                        'required_deposit': required_deposit,
                    }

                try:
                    # Check if the record exists
                    existing_promotion = ActivityFinancial.objects.filter(activity_date=activity_date, site_id=site_id).first()

                    if not existing_promotion:
                        # If no record exists, create a new one
                    
                        try:
                            ActivityFinancial.objects.create(
                                            activity_date=activity_date,
                                            site_id=site_id,
                                            **data
                                        )
                            return JsonResponse({'success': True, 'message': 'Financial record created successfully!',
                                                 'form_name': 'financial',
                                                 'data': {data}})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["financial-general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["financial-general"] = "Foreign key constraint failed. Invalid site ID or financial ID."
                            else:
                                errors["financial-general"] = "Database integrity error occurred."
                            return JsonResponse({"success": False, "errors": errors}, status=400)

                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                        except Exception as e:
                            return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)

                    else:
                        # If record exists, update it
                        try:
                            obj, created = ActivityFinancial.objects.update_or_create(
                                                    activity_date=activity_date, site_id=site_id,
                                                    defaults=data
                                                )
                            return JsonResponse({'success': True, 'message': 'Financial record updated successfully!',
                                                'form_name': 'financial',
                                                 'data': data})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["financial-general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["financial-general"] = "Foreign key constraint failed. Invalid site ID or financial ID."
                            else:
                                errors["financial-general"] = "Database integrity error occurred."
                            return JsonResponse({"success": False, "errors": errors}, status=400)

                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                        except Exception as e:
                            return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)

                except Exception as e:
                    return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)

            
            elif form_name == 'Purchase':
                errors = {}

                # Get activity_date from request
                activity_date = request.POST.get('purchases_activity_date')  
                if not activity_date:
                    errors["purchases-general"] = "Activity date is required."
                try:
                    activity_date = parser.parse(activity_date).date()
                except ValueError:
                    return JsonResponse({"success": False, "message": "Invalid Activity Date format!"}, status=400)

                site_id = request.POST.get('purchase_site_id')
                if not site_id:
                    errors["purchases-general"] = "Not able to extract the site ID."

                # If errors exist, return response
                if errors:
                    return JsonResponse({"success": False, "errors": errors}, status=400)

                data = {
                    'purchases_food': request.POST.get('purchases_food'),
                    'purchases_paper': request.POST.get('purchases_paper'),
                    'purchases_liquor': request.POST.get('purchases_liquor'),
                }

                try:
                    # Check if the record exists
                    existing_purchase = ActivityPurchase.objects.filter(activity_date=activity_date, site_id=site_id).first()

                    if not existing_purchase:
                        # If no record exists, create a new one
                        try:
                            ActivityPurchase.objects.create(
                                activity_date=activity_date,
                                site_id=site_id,
                                **data
                            )
                            return JsonResponse({'success': True, 'message': 'Purchases data added successfully!'})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["purchases-general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["purchases-general"] = "Foreign key constraint failed. Invalid site ID or purchases ID."
                            else:
                                errors["purchases-general"] = "Database integrity error occurred."
                            return JsonResponse({"success": False, "errors": errors}, status=400)

                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                        except Exception as e:
                            return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)

                    else:
                        # If record exists, update it
                        try:
                            for key, value in data.items():
                                setattr(existing_purchase, key, value)  # Update existing fields dynamically
                            existing_purchase.save()

                            return JsonResponse({'success': True, 'message': 'Purchases record updated successfully!'})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["purchases-general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["purchases-general"] = "Foreign key constraint failed. Invalid site ID or purchases ID."
                            else:
                                errors["purchases-general"] = "Database integrity error occurred."
                            return JsonResponse({"success": False, "errors": errors}, status=400)

                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                        except Exception as e:
                            return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)

                except Exception as e:
                    return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)

            
            elif form_name == 'paid-out':
                errors = {}

                # Get activity_date from request
                activity_date = request.POST.get('paid_out_activity_date')  # Unique ID for updating records
                if not activity_date:
                    errors["paidout-general"] = "Activity date is required."
                try:
                    activity_date = parser.parse(activity_date).date()
                except ValueError:
                    return JsonResponse({"success": False, "message": "Invalid Activity Date format!"}, status=400)


                site_id = request.POST.get('paid_out_site_id')
                if not site_id:
                    errors["paidout-general"] = "Not able to extract the site ID."

                paid_out_total = request.POST.get('paid_out_total')

                # If errors exist, return response
                if errors:
                    return JsonResponse({"success": False, "errors": errors}, status=400)
                
                food_paid_out = float(request.POST.get('food_paid_out') or 0)
                liquor_paid_out = float(request.POST.get('liquor_paid_out') or 0)
                supplies_paid_out = float(request.POST.get('supplies_paid_out') or 0)
                repair_maintenance_paid_out = float(request.POST.get('repair_maintenance_paid_out') or 0)
                advertising_paid_out = float(request.POST.get('advertising_paid_out') or 0)
                entertainment_paid_out = float(request.POST.get('entertainment_paid_out') or 0)
                others_paid_out = float(request.POST.get('others_paid_out') or 0)
                hst_gst_paid_out = float(request.POST.get('hst_gst_paid_out') or 0)
                paid_out_total = food_paid_out + liquor_paid_out + supplies_paid_out + repair_maintenance_paid_out + advertising_paid_out + entertainment_paid_out + others_paid_out + hst_gst_paid_out

                data = {
                        'food_paid_out': food_paid_out,
                        'liquor_paid_out': liquor_paid_out,
                        'supplies_paid_out': supplies_paid_out,
                        'repair_maintenance': repair_maintenance_paid_out,
                        'advertising': advertising_paid_out,
                        'entertainment': entertainment_paid_out,
                        'others': others_paid_out,
                        'hst_gst': hst_gst_paid_out,
                        'paid_out_total': paid_out_total,
                    }

                try:
                    # Check if the record exists
                    existing_paidout = ActivityPaidOut.objects.filter(activity_date=activity_date, site_id=site_id).first()

                    if not existing_paidout:
                        # If no record exists, create a new one
                    
                        try:
                            ActivityPaidOut.objects.create(
                                            activity_date=activity_date,
                                            site_id=site_id,
                                            **data
                                        )
                            return JsonResponse({'success': True, 
                                                 'message': 'Paid-out record created successfully!',
                                                 'form_name': 'paid-out',
                                                 'data': data})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["paidout-general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["paidout-general"] = "Foreign key constraint failed. Invalid site ID or paidout ID."
                            else:
                                errors["paidout-general"] = "Database integrity error occurred."
                            return JsonResponse({"success": False, "errors": errors}, status=400)

                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                        except Exception as e:
                            return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)

                    else:
                        # If record exists, update it
                        try:
                            obj, created = ActivityPaidOut.objects.update_or_create(
                                                    activity_date=activity_date, site_id=site_id,
                                                    defaults=data
                                                )
                            return JsonResponse({'success': True, 
                                                 'message': 'Paid-out record updated successfully!',
                                                 'form_name': 'paid-out',
                                                 'data': data})

                        except IntegrityError as e:
                            error_message = str(e)
                            if "unique constraint" in error_message.lower():
                                errors["purchases-general"] = "A unique constraint error occurred. Please check the data."
                            elif "foreign key" in error_message.lower():
                                errors["purchases-general"] = "Foreign key constraint failed. Invalid site ID or purchases ID."
                            else:
                                errors["purchases-general"] = "Database integrity error occurred."
                            return JsonResponse({"success": False, "errors": errors}, status=400)

                        except DataError as e:
                            return JsonResponse({"success": False, "errors": {"general": f"Invalid data error: {str(e)}"}}, status=400)

                        except Exception as e:
                            return JsonResponse({"success": False, "message": f"Unexpected error: {str(e)}"}, status=500)

                except Exception as e:
                    return JsonResponse({"success": False, "message": f"Error: {str(e)}"}, status=500)

            else:
                return JsonResponse({'success': False, 'message': 'Invalid form name.'})
    
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
def download_activity_excel(request):
    # Create a new workbook
    site_username=request.session.get('user_name')
    user = User.objects.get(username=site_username)
    site_id = user.id
    wb = openpyxl.Workbook()
    ws_sales = wb.active
    ws_sales.title = "Sales"

    # Define other sheets
    ws_promotions = wb.create_sheet(title="Promotions")
    ws_purchases = wb.create_sheet(title="Purchases")
    ws_paid_out = wb.create_sheet(title="Paid-Out")
    ws_financial = wb.create_sheet(title="Financial")

    # Define column headers for each table
    sales_headers = ["Date", "Food Sales", "Liquor Sales", "Beer Sales", "Wine Sales", "Iced Beverages", "Draft Sales",
                     "Honest Fees", "Misc Income", "Total Sales", "Sales Tax 1", "Sales Tax 2", "Liquor Tax", "Customer Count", "Notes"]
    
    promotions_headers = ["Date", "Staff Meals", "Manager Meals", "Coupons/Discounts", "Chucks Bucks",
                          "Manager Promotions", "QSA Complaints", "LSM 1", "LSM 2", "Total Promotions"]
    
    purchases_headers = ["Date", "Food Purchases", "Paper Purchases", "Liquor Purchases"]
    
    paid_out_headers = ["Date", "Food Paid Out", "Liquor Paid Out", "Supplies", "Repair & Maintenance",
                        "Advertising", "Entertainment", "Others", "HST/GST", "Total Paid-Out"]
    
    financial_headers = ["Date", "Gift Card Purchased", "Gift Card Redeemed", "Skip The Dishes", "Just Eat",
                         "Amex", "Diners Club", "Discover", "MasterCard", "Visa", "Debit Card", "Total Financial", 
                         "Actual Deposit", "Required Deposit"]

    # Function to write data to sheets
    def write_data(sheet, headers, queryset, field_names):
        sheet.append(headers)
        for obj in queryset:
            row = [getattr(obj, field, '') for field in field_names]
            sheet.append(row)
        # Auto-adjust column width
        for col_num, col_name in enumerate(headers, 1):
            sheet.column_dimensions[get_column_letter(col_num)].width = max(len(col_name) + 2, 15)

    # Fetch data
    sales_data = ActivitySales.objects.filter(site_id=site_id)
    promotions_data = ActivityPromotions.objects.filter(site_id=site_id)
    purchases_data = ActivityPurchase.objects.filter(site_id=site_id)
    paid_out_data = ActivityPaidOut.objects.filter(site_id=site_id)
    financial_data = ActivityFinancial.objects.filter(site_id=site_id)

    # Write data to sheets
    write_data(ws_sales, sales_headers, sales_data, [
        "activity_date", "food", "liquor", "beer", "wine", "beverages_iced_beverages",
        "draft", "honest_to_goodness_fees", "miscellaneous_income", "sales_total",
        "sales_tax1", "sales_tax2", "liquor_tax_1", "customer_count", "sales_notes"
    ])

    write_data(ws_promotions, promotions_headers, promotions_data, [
        "activity_date", "staff_meals", "manager_meals", "coupons_disc", "chucks_bucks",
        "manager_prom", "qsa_complaints", "lsm_1", "lsm_2", "promotion_total"
    ])

    write_data(ws_purchases, purchases_headers, purchases_data, [
        "activity_date", "purchases_food", "purchases_paper", "purchases_liquor"
    ])

    write_data(ws_paid_out, paid_out_headers, paid_out_data, [
        "activity_date", "food_paid_out", "liquor_paid_out", "supplies_paid_out",
        "repair_maintenance", "advertising", "entertainment", "others", "hst_gst",
        "paid_out_total"
    ])

    write_data(ws_financial, financial_headers, financial_data, [
        "activity_date", "gift_card_purchased", "gift_card_redeemed", "skip_the_dishes",
        "just_eat", "amex", "din_club", "discover", "master_card", "visa",
        "dr_card", "financial_total", "actual_deposit_amt", "required_deposit"
    ])

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=site_{site_id}_report.xlsx'
    
    # Save the workbook to response
    wb.save(response)

    return response

def way_to_sales_report(request):
    return render(request, 'site_panel/activity_report.html')

def way_to_biweekly_payroll_report(request):
    return render(request, 'site_panel/biweekly_payroll.html')

def way_to_weekly_flash_report(request):
    return render(request, 'site_panel/weekly_flash_report.html')

def way_to_weekly_sales_report(request):
    return render(request, 'site_panel/weekly_activity_report.html')

def way_to_weekly_wage_cost_report(request):
    return render(request, 'site_panel/weekly_wage_cost_report.html')

def way_to_weekly_promo_report(request):
    return render(request, 'site_panel/weekly_promo_report.html')

def way_to_monthly_rollover_report(request):
    return render(request, 'site_panel/monthly_rollover.html')

def way_to_summary_report(request):
    return render(request, 'site_panel/Summary_report.html')

def genereate_activity_sales_report(request):

    location = request.GET.get('location')
    user = User.objects.get(last_name=location)
    date = request.GET.get('data_date')
    report_type = request.GET.get('report_type')

    location = request.GET.get('location')
    date = request.GET.get('data_date')

    if not location:
        return JsonResponse({"error": "Location is required"}, status=400)
    if not date:
        return JsonResponse({"error": "Date is required"}, status=400)

    try:
        td_date = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD"}, status=400)
    

    try:
        user = User.objects.get(last_name=location)
    except User.DoesNotExist:
        return JsonResponse({"error": "No user found for the given location"}, status=404)

    site_id = user.id

    with connection.cursor() as cursor:
        cursor.execute("SELECT MIN(activity_date) FROM activity_sales;")
        earliest_date = cursor.fetchone()[0]

    if earliest_date and td_date < earliest_date:
        return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=site_{site_id}_report.pdf'
        

    base_query = """
            SELECT 
                s.food as "sales_food",
                s.liquor AS "sales_liquor",
                s.beer AS "sales_beer",
                s.wine AS "sales_wine",
                s.beverages_iced_beverages AS "sales_beverages_iced_beverages",
                s.honest_to_goodness_fees AS "sales_honest_to_goodness_fees",
                s.miscellaneous_income AS "sales_miscellaneous_income",
                s.sales_total AS "sales_total",
                s.sales_tax1 AS "sales_tax1",
                s.sales_tax2 AS "sales_tax2",
                s.liquor_tax_1 AS "sales_liquor_tax_1",
                f.skip_the_dishes AS "financial_skip_the_dishes",
                f.gift_card_purchased AS "financial_gift_card_purchased",
                f.gift_card_redeemed AS "financial_gift_card_redeemed",
                f.just_eat AS "financial_just_eat",
                f.amex AS "financial_amex",
                f.din_club AS "financial_din_club",
                f.discover AS "financial_discover",
                f.master_card AS "financial_master_card",
                f.visa AS "financial_visa",
                f.dr_card AS "financial_dr_card",
                f.financial_total AS "financial_financial_total",
                f.actual_deposit_amt AS "financial_actual_deposit_amt",
                f.required_deposit AS "financial_required_deposit",
                po.food_paid_out AS "paid_out_food",
                po.liquor_paid_out AS "paid_out_liquor",
                po.supplies_paid_out AS "paid_out_supplies",
                po.repair_maintenance AS "paid_out_repair_maintenance",
                po.entertainment AS "paid_out_entertainment",
                po.advertising AS "paid_out_advertising",
                po.others AS "paid_out_others",
                po.hst_gst AS "paid_out_hst_gst",
                po.paid_out_total AS "paid_out_total",
                pr.staff_meals AS "promo_staff_meals",
                pr.manager_meals AS "promo_manager_meals",
                pr.coupons_disc AS "promo_coupons_disc",
                pr.chucks_bucks AS "promo_chucks_bucks",
                pr.lsm_1 AS "promo_lsm_1",
                pr.lsm_2 AS "promo_lsm_2",
                pr.manager_prom AS "promo_manager_prom",
                pr.qsa_complaints AS "promo_qsa_complaints",
                pr.promotion_total AS "promo_promotion_total"
            FROM
                activity_sales s
            LEFT JOIN 
                activity_purchase p ON s.site_id = p.site_id AND s.activity_date = p.activity_date
            LEFT JOIN 
                activity_paid_out po ON s.site_id = po.site_id AND s.activity_date = po.activity_date
            LEFT JOIN 
                activity_financial f ON s.site_id = f.site_id AND s.activity_date = f.activity_date
            LEFT JOIN 
                activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
            LEFT JOIN  
                auth_user u ON s.site_id = u.id
            WHERE s.activity_date = %s AND s.site_id = %s;
            """
    
    base_query2 = """
            SELECT 
                c.e_position_type AS Position_Type, 
                COUNT(*) AS occurrence_count, 
                SUM(p.total_daily_rate) AS total_daily_rate_sum
            FROM site_panel_paydailyentry p
            LEFT JOIN compensation c ON c.c_id = p.c_id_id
            WHERE p.day_date = %s AND c.s_id = %s
            GROUP BY c.e_position_type;
            """
    
    
    query_params = [td_date, site_id]

    # Execute the raw SQL queries
    try:
        with connection.cursor() as cursor:
            cursor.execute(base_query, query_params)
            columns = [col[0] for col in cursor.description]
            sales_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
            sales_data = sales_data[0] if sales_data else {}

            cursor.execute(base_query2, query_params)
            columns = [col[0] for col in cursor.description]
            position_data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    except DatabaseError as db_error:
        raise ValueError(f"Database error: {str(db_error)}")

    if report_type == "PDF":
        pdf = canvas.Canvas(response, pagesize=letter)  # Initialize canvas only once

        # Draw the image
        static_file_path = finders.find('site_panel/images/restaurant_icon.png')
        if static_file_path:
            try:
                pdf.drawImage(static_file_path, 50, 710, width=80, height=60)
            except Exception as e:
                print(f"Error drawing image: {e}")
        else:
            print("Static file not found.")

        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(200, 750, f"Daily Sales & Activity Report")
        pdf.drawString(230, 730, f"Date: {date}")
        pdf.setFont("Helvetica", 12)

        # Cash Deposits Table
        cash_deposits_data = [
            ["Cash Deposits", f"${float(sales_data.get('financial_required_deposit', 0) or 0):,.2f}"],
            ["Just Eat", f"${float(sales_data.get('financial_just_eat', 0)or 0):,.2f}"],
            ["Skip The Dishes", f"${float(sales_data.get('financial_skip_the_dishes', 0)or 0):,.2f}"],
            ["Order Partner Total", f"${float(sales_data.get('financial_skip_the_dishes', 0)or 0) + float(sales_data.get('financial_just_eat', 0)or 0):,.2f}"],
            ["Visa", f"${float(sales_data.get('financial_visa', 0)or 0):,.2f}"],
            ["Mastercard", f"${float(sales_data.get('financial_master_card', 0)or 0):,.2f}"],
            ["American Express", f"${float(sales_data.get('financial_amex', 0)or 0):,.2f}"],
            ["Diners Club", f"${float(sales_data.get('financial_din_club', 0)or 0):,.2f}"],
            ["Discover", f"${float(sales_data.get('financial_discover', 0)or 0):,.2f}"],
            ["Debit Card", f"${float(sales_data.get('financial_dr_card', 0)or 0):,.2f}"],
            ["Card Total", f"${float(sales_data.get('financial_visa', 0)or 0) + float(sales_data.get('financial_master_card', 0)or 0) + float(sales_data.get('financial_amex', 0)or 0) + float(sales_data.get('financial_din_club', 0)or 0) + float(sales_data.get('financial_discover', 0)or 0) + float(sales_data.get('financial_dr_card', 0)or 0):,.2f}"],
            ["Total Deposit", f"${float(sales_data.get('financial_required_deposit', 0)or 0) + float(sales_data.get('financial_skip_the_dishes', 0)or 0) + float(sales_data.get('financial_just_eat', 0)or 0) + float(sales_data.get('financial_visa', 0)or 0) + float(sales_data.get('financial_master_card', 0)or 0) + float(sales_data.get('financial_amex', 0)or 0) + float(sales_data.get('financial_din_club', 0)or 0) + float(sales_data.get('financial_discover', 0)or 0) + float(sales_data.get('financial_dr_card', 0)or 0):,.2f}"]
        ]

        y = 700  # Set the starting Y position

        cash_deposits_table = Table(cash_deposits_data)
        cash_deposits_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        table_width, table_height = cash_deposits_table.wrapOn(pdf, 50, y)  # Get table dimensions
        cash_deposits_table.drawOn(pdf, 50, y - table_height)  # Draw the table at y - table height

        y -= table_height + 20
        
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, 440, "Cash Payouts")

        Cash_Payouts=[
            ["Food", f"${sales_data.get('paid_out_food', 0) or 0:,.2f}"],
            ["Liquor", f"${sales_data.get('paid_out_liquor', 0) or 0:,.2f}"],
            ["Supplies", f"${sales_data.get('paid_out_supplies', 0) or 0:,.2f}"],
            ["Repair Maintenance", f"${sales_data.get('paid_out_repair_maintenance', 0) or 0:,.2f}"],
            ["Advertising", f"${sales_data.get('paid_out_advertising', 0) or 0:,.2f}"],
            ["Entertainment", f"${sales_data.get('paid_out_entertainment', 0) or 0:,.2f}"],
            ["Others", f"${sales_data.get('paid_out_others', 0) or 0:,.2f}"],
            ["HST", f"${sales_data.get('paid_out_hst_gst', 0) or 0:,.2f}"],
            ["Total Payouts", f"${sales_data.get('paid_out_total', 0) or 0:,.2f}"]
        ]

        y = 420  # Set the starting Y position

        cash_payouts_table = Table(Cash_Payouts)
        cash_payouts_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        table_width, table_height = cash_payouts_table.wrapOn(pdf, 50, y)  # Get table dimensions
        cash_payouts_table.drawOn(pdf, 50, y - table_height)  # Draw the table at y - table height

        y -= table_height + 20

        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(50, 160, "Position Type Breakdown")

        
        position_type = [["Position Type","Percentage of Total Deposit (%)"]]

        total_deposit = sales_data.get('financial_required_deposit')
        total_deposit = float(total_deposit) if total_deposit is not None else 0

        # Starting position for drawing text

        labour_total= 0

        for position in position_data:
            if isinstance(position, dict):
                e_position_type = position.get('Position_Type', 'Unknown')
            else:
                e_position_type = 'Invalid Position'
            
            l1=[]
            l1.append(e_position_type)
            
            if total_deposit > 0:
                if isinstance(position, dict):
                    total_daily_rate_sum = position.get('total_daily_rate_sum', 0)
                    labour_total += total_daily_rate_sum
                else:
                    total_daily_rate_sum = 0  # Default value if position is not a dictionary

                percentage = (total_daily_rate_sum / Decimal(total_deposit)) * 100

                print(percentage)

                l1.append(f"${percentage:.2f}%")
                position_type.append(l1)
            else:
                l1.append(f"N/A")
                position_type.append(l1)

        labour_total = Decimal(labour_total)  # Ensure it's a Decimal
        total_deposit = Decimal(total_deposit)  # Convert float to Decimal

        labour_percentage = (labour_total / total_deposit) * 100 if total_deposit > 0 else 0
        position_type.append(['Labour_Total', f"{labour_percentage:.2f}%"])


        y = 140  # Set the starting Y position

        position_type_table = Table(position_type)
        position_type_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        table_width, table_height = position_type_table.wrapOn(pdf, 50, y)  # Get table dimensions
        position_type_table.drawOn(pdf, 50, y - table_height)  # Draw the table at y - table height

        y -= table_height + 20

        pdf.setFont("Helvetica", 12)

        net_sales_data = [
            ["NET Sales", f"${sales_data.get('sales_total', 0) or 0:,.2f}"],
            ["Total Tax", f"${(sales_data.get('sales_tax1', 0) or 0) + (sales_data.get('sales_tax2', 0) or 0) + (sales_data.get('sales_liquor_tax_1', 0) or 0):,.2f}"],
            ["GROSS Sales", f"${(sales_data.get('sales_total', 0) or 0) + (sales_data.get('sales_tax1', 0) or 0) + (sales_data.get('sales_tax2', 0) or 0) + (sales_data.get('sales_liquor_tax_1', 0) or 0):,.2f}"],
            [],
            ["Gift Card Sold", f"${sales_data.get('financial_gift_card_purchased', 0) or 0:,.2f}"],
            ["Misc Income", f"${sales_data.get('sales_miscellaneous_income', 0) or 0:,.2f}"],
            ["Promo & Charges", f"${sales_data.get('promo_promotion_total', 0) or 0:,.2f}"],
            ["Cash Payouts", f"${sales_data.get('paid_out_total', 0) or 0:,.2f}"],
            ["NET Total", f"${Decimal(sales_data.get('sales_total', 0) or 0) + Decimal(sales_data.get('sales_tax1', 0) or 0) + Decimal(sales_data.get('sales_tax2', 0) or 0) + Decimal(sales_data.get('sales_liquor_tax_1', 0) or 0) - Decimal(sales_data.get('financial_gift_card_purchased', 0) or 0) - Decimal(sales_data.get('promo_promotion_total', 0) or 0) - Decimal(sales_data.get('paid_out_total', 0) or 0):,.2f}"],
            ["Cash Over/short", f"$({(sales_data.get('financial_skip_the_dishes', 0) or 0) + (sales_data.get('financial_just_eat', 0) or 0):,.2f})"]
        ]

        y = 700

        net_sales_table = Table(net_sales_data)
        net_sales_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        table_width, table_height = net_sales_table.wrapOn(pdf, 350, y)
        net_sales_table.drawOn(pdf, 350, y - table_height)

        y -= table_height + 20

        financial_gift_card_redeemed = sales_data.get('financial_gift_card_redeemed') or Decimal(0)
        promo_promotion_total = sales_data.get('promo_promotion_total') or Decimal(0)

        promo_data = [
            ["Gift Card Redeemed", f"${sales_data.get('financial_gift_card_redeemed', 0) or 0:,.2f}"],
            [],
            ["Promo Meals", f"${sales_data.get('promo_staff_meals', 0) or 0:,.2f}"],
            ["Manager Meals", f"${sales_data.get('promo_manager_meals', 0) or 0:,.2f}"],
            ["Manager Promo", f"${sales_data.get('promo_manager_prom', 0) or 0:,.2f}"],
            ["Coupons & Discounts", f"${sales_data.get('promo_coupons_disc', 0) or 0:,.2f}"],
            ["QSA Complaints", f"${sales_data.get('promo_qsa_complaints', 0) or 0:,.2f}"],
            ["Chucks Bucks", f"${sales_data.get('promo_chucks_bucks', 0) or 0:,.2f}"],
            ["LSM 1", f"${sales_data.get('promo_lsm_1', 0) or 0:,.2f}"],
            ["LSM 2", f"${sales_data.get('promo_lsm_2', 0) or 0:,.2f}"],
            ["Total Discounts", f"${(sales_data.get('promo_promotion_total', 0) or 0) + (sales_data.get('promo_chucks_bucks', 0) or 0) + (sales_data.get('promo_lsm_1', 0) or 0) + (sales_data.get('promo_lsm_2', 0) or 0):,.2f}"],
            [],
            ["Promo + Charges", f"${financial_gift_card_redeemed + promo_promotion_total:,.2f}"]
        ]

        y = 460  # Starting Y position for the table

        promo_table = Table(promo_data)
        promo_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Left align labels
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'), # Right align values
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        table_width, table_height = promo_table.wrapOn(pdf, 350, y)
        promo_table.drawOn(pdf, 350, y - table_height)

        # Update y position for subsequent elements
        y -= table_height + 20

        pdf.showPage()
        pdf.save()

        return response
    elif report_type == "Excel":
            excel_response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_response['Content-Disposition'] = f'attachment; filename=site_{site_id}_report.xlsx'

            # Create a workbook and select the active sheet
            wb = Workbook()
            ws = wb.active

            # --- ADDED CODE START ---
            # Add Image
            static_file_path = finders.find('site_panel/images/restaurant_icon.png')
            if static_file_path:
                try:
                    img = ExcelImage(static_file_path)
                    # --- ADDED CODE TO DECREASE IMAGE SIZE ---
                    img.width = int(img.width * 0.06)  # Reduce width by 50%
                    img.height = int(img.height * 0.06) # Reduce height by 50%
                    # --- END OF ADDED CODE ---
                    ws.add_image(img, 'A1')  # Adjust cell as needed
                except Exception as e:
                    print(f"Error adding image to Excel: {e}")

            # Add Title
            ws.cell(row=1, column=3, value='Daily Sales & Activity Report').font = openpyxl.styles.Font(bold=True, size=16)

            # Add Date
            ws.cell(row=2, column=3, value=f'Date: {date}').font = openpyxl.styles.Font(size=12)

            ws.append([])

            # Define border style
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            # --- ADDED CODE END ---

            # Cash Deposits Table Data
            cash_deposits_data = [
                ["Cash Deposits", float(sales_data.get('financial_required_deposit', 0) or 0)],
                ["Just Eat", float(sales_data.get('financial_just_eat', 0) or 0)],
                ["Skip The Dishes", float(sales_data.get('financial_skip_the_dishes', 0) or 0)],
                ["Order Partner Total", float(sales_data.get('financial_skip_the_dishes', 0) or 0) + float(sales_data.get('financial_just_eat', 0) or 0)],
                ["Visa", float(sales_data.get('financial_visa', 0) or 0)],
                ["Mastercard", float(sales_data.get('financial_master_card', 0) or 0)],
                ["American Express", float(sales_data.get('financial_amex', 0) or 0)],
                ["Diners Club", float(sales_data.get('financial_din_club', 0) or 0)],
                ["Discover", float(sales_data.get('financial_discover', 0) or 0)],
                ["Debit Card", float(sales_data.get('financial_dr_card', 0) or 0)],
                ["Card Total", float(sales_data.get('financial_visa', 0) or 0) + float(sales_data.get('financial_master_card', 0) or 0) + float(sales_data.get('financial_amex', 0) or 0) + float(sales_data.get('financial_din_club', 0) or 0) + float(sales_data.get('financial_discover', 0) or 0) + float(sales_data.get('financial_dr_card', 0) or 0)],
                ["Total Deposit", float(sales_data.get('financial_required_deposit', 0) or 0) + float(sales_data.get('financial_skip_the_dishes', 0) or 0) + float(sales_data.get('financial_just_eat', 0) or 0) + float(sales_data.get('financial_visa', 0) or 0) + float(sales_data.get('financial_master_card', 0) or 0) + float(sales_data.get('financial_amex', 0) or 0) + float(sales_data.get('financial_din_club', 0) or 0) + float(sales_data.get('financial_discover', 0) or 0) + float(sales_data.get('financial_dr_card', 0) or 0)]
            ]

            # Write Cash Deposits Table
            start_row = 4
            start_col = 1
            for row_index, row_data in enumerate(cash_deposits_data):
                ws.append(row_data)
                # Apply border
                # Apply border to all cells EXCEPT the second column
                for col_index, cell in enumerate(ws[start_row + row_index]):
                    if col_index != 2:  # Exclude the second column (index 1)
                        cell.border = thin_border
                    else:
                        cell.border = Border() # Set an empty border (effectively removing it)

            # Cash Payouts Table Data
            cash_payouts_data = [
                ["Food", float(sales_data.get('paid_out_food', 0 ) or 0)],
                ["Liquor", float(sales_data.get('paid_out_liquor', 0) or 0)],
                ["Supplies", float(sales_data.get('paid_out_supplies', 0) or 0)],
                ["Repair Maintenance", float(sales_data.get('paid_out_repair_maintenance', 0) or 0)],
                ["Advertising", float(sales_data.get('paid_out_advertising', 0) or 0)],
                ["Entertainment", float(sales_data.get('paid_out_entertainment', 0) or 0)],
                ["Others", float(sales_data.get('paid_out_others', 0) or 0)],
                ["HST", float(sales_data.get('paid_out_hst_gst', 0) or 0)],
                ["Total Payouts", float(sales_data.get('paid_out_total', 0) or 0)]
            ]

            # Write Cash Payouts Table
            start_row = ws.max_row + 2
            for row_index, row_data in enumerate(cash_payouts_data):
                ws.append(row_data)
                # Apply border
                for col_index, cell in enumerate(ws[start_row + row_index]):
                    if col_index != 2:  # Exclude the second column (index 1)
                        cell.border = thin_border
                    else:
                        cell.border = Border() # Set an empty border (effectively removing it)

            # Position Type Breakdown Table Data
            position_type_data = [["Position Type", "Percentage of Total Deposit (%)"]]
            total_deposit = sales_data.get('financial_required_deposit', 0)
            labour_total=0
            for position in position_data:
                if isinstance(position, dict):
                    e_position_type = position.get('Position_Type', 'Unknown')
                    total_daily_rate_sum = position.get('total_daily_rate_sum', 0)
                    labour_total += total_daily_rate_sum
                    if total_deposit > 0:
                        percentage = (total_daily_rate_sum / total_deposit) * 100
                        position_type_data.append([e_position_type, f"{percentage:.2f}%"])
                    else:
                        position_type_data.append([e_position_type, "N/A"])
            labour_percentage = (labour_total / total_deposit) * 100 if total_deposit > 0 else 0
            position_type_data.append(['Labour_Total', f"{labour_percentage:.2f}%"])

            # Write Position Type Breakdown Table
            start_row = ws.max_row + 2
            for row_index, row_data in enumerate(position_type_data):
                ws.append(row_data)
                # Apply border
                for cell in ws[start_row + row_index]:
                    cell.border = thin_border

            # Net Sales Data
            net_sales_data = [
                ["NET Sales", f"${sales_data.get('sales_total', 0) or 0:,.2f}"],
                ["Total Tax", f"${(sales_data.get('sales_tax1', 0) or 0) + (sales_data.get('sales_tax2', 0) or 0) + (sales_data.get('sales_liquor_tax_1', 0) or 0):,.2f}"],
                ["GROSS Sales", f"${(sales_data.get('sales_total', 0) or 0) + (sales_data.get('sales_tax1', 0) or 0) + (sales_data.get('sales_tax2', 0) or 0) + (sales_data.get('sales_liquor_tax_1', 0) or 0):,.2f}"],
                [],
                ["Gift Card Sold", f"${sales_data.get('financial_gift_card_purchased', 0) or 0:,.2f}"],
                ["Promo & Charges", f"${sales_data.get('promo_promotion_total', 0) or 0:,.2f}"],
                ["Cash Payouts", f"${sales_data.get('paid_out_total', 0) or 0:,.2f}"],
                ["NET Total", f"${Decimal(sales_data.get('sales_total', 0) or 0) + Decimal(sales_data.get('sales_tax1', 0) or 0) + Decimal(sales_data.get('sales_tax2', 0) or 0) + Decimal(sales_data.get('sales_liquor_tax_1', 0) or 0) - Decimal(sales_data.get('financial_gift_card_purchased', 0) or 0) - Decimal(sales_data.get('promo_promotion_total', 0) or 0) - Decimal(sales_data.get('paid_out_total', 0) or 0):,.2f}"],
                ["Cash Over/short", f"$({(sales_data.get('financial_skip_the_dishes', 0) or 0) + (sales_data.get('financial_just_eat', 0) or 0):,.2f})"]
            ]

            # Write Net Sales Data
            start_row = 4
            start_col = 4 # Start from column D
            for row_index, row_data in enumerate(net_sales_data):
                for col_index, cell_value in enumerate(row_data):
                    ws.cell(row=start_row + row_index, column=start_col + col_index, value=cell_value)
                    ws.cell(row=start_row + row_index, column=start_col + col_index).border = thin_border

            # Promo Data
            promo_data = [
                ["Gift Card Redeemed", f"${sales_data.get('financial_gift_card_redeemed', 0) or 0:,.2f}"],
                [],
                ["Promo Meals", f"${sales_data.get('promo_staff_meals', 0) or 0:,.2f}"],
                ["Manager Meals", f"${sales_data.get('promo_manager_meals', 0) or 0:,.2f}"],
                ["Coupons & Discounts", f"${sales_data.get('promo_coupons_disc', 0) or 0:,.2f}"],
                ["QSA Complaints", f"${sales_data.get('promo_qsa_complaints', 0) or 0:,.2f}"],
                ["Chucks Bucks", f"${sales_data.get('promo_chucks_bucks', 0) or 0:,.2f}"],
                ["LSM 1", f"${sales_data.get('promo_lsm_1', 0) or 0:,.2f}"],
                ["LSM 2", f"${sales_data.get('promo_lsm_2', 0) or 0:,.2f}"],
                ["Total Discounts", f"${(sales_data.get('promo_qsa_complaints', 0) or 0) + (sales_data.get('promo_chucks_bucks', 0) or 0) + (sales_data.get('promo_lsm_1', 0) or 0) + (sales_data.get('promo_lsm_2', 0) or 0):,.2f}"],
                [],
                ["Promo + Charges", f"${(sales_data.get('financial_gift_card_redeemed', 0) or 0) + (sales_data.get('promo_promotion_total', 0) or 0):,.2f}"]
            ]

            # Write Promo Data
            start_row = ws.max_row - len(promo_data) + 1 # Adjust start row to be near Net Sales
            start_col = 4 # Start from column F
            for row_index, row_data in enumerate(promo_data):
                for col_index, cell_value in enumerate(row_data):
                    ws.cell(row=start_row + row_index, column=start_col + col_index, value=cell_value)
                    ws.cell(row=start_row + row_index, column=start_col + col_index).border = thin_border

            # Center align all cells with data
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Save the workbook to the response
            wb.save(excel_response)
            return excel_response
    return JsonResponse({"error": "Invalid report type"}, status=400)

def payroll(request):
    location = request.GET.get('location')
    end_date_str = request.GET.get('data_date')
    report_type = request.GET.get('report_type')

    if not location:
        return JsonResponse({"error": "Location is required"}, status=400)
    if not end_date_str:
        return JsonResponse({"error": "Date is required"}, status=400)

    try:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD"}, status=400)

        # Get user by location
    try:
        user = User.objects.get(last_name=location)
    except User.DoesNotExist:
        return JsonResponse({"error": "No user found for the given location"}, status=404)
    restaurant_name = user.first_name
    site_id = user.id
    start_date = end_date - timedelta(days=13)

        # Get the earliest employee record date
    with connection.cursor() as cursor:
        cursor.execute("SELECT MIN(day_date) FROM site_panel_paydailyentry;")
        earliest_date = cursor.fetchone()[0]

    if earliest_date and start_date < earliest_date:
        return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=site_{site_id}_report.pdf'

    base_query = """
        SELECT
            e.employee_name AS Employee_Name,
            c.e_position AS Position_Type,
            p.day_date AS Date,
            p.regular_hrs_worked AS Hours_Worked,
            p.daily_hourly_rate AS Daily_Hourly_Rate,
            p.total_daily_rate AS Total_Daily_Rate
        FROM
            site_panel_paydailyentry p
        JOIN
            compensation c ON c.c_id = p.c_id_id
        JOIN
            employee e ON e.e_id = c.e_id
        WHERE
            p.day_date BETWEEN %s AND %s
            
            AND c.s_id = %s
        ORDER BY
            e.employee_name, c.e_position, p.day_date;
        """

    query_params = [start_date, end_date, site_id]

    try:
        with connection.cursor() as cursor:
            cursor.execute(base_query, query_params)
            results = cursor.fetchall()
            column_names = ['Employee_Name', 'Position_Type', 'Date', 'Hours_Worked', 'Daily_Hourly_Rate', 'Total_Daily_Rate']
            df = pd.DataFrame(results, columns=column_names)

            df['Date'] = pd.to_datetime(df['Date'])
            df['Day_Name'] = df['Date'].dt.day_name()
            df['Days_Since_Start'] = (df['Date'] - pd.Timestamp(start_date)).dt.days
            df['Days_Since_Start'] = (df['Date'] - pd.Timestamp(start_date)).dt.days
            df['Week_Number'] = (df['Days_Since_Start'] // 7) + 1  # Divide by 7 to group into weeks

            # Pivot the DataFrame
            pivot_hours = df.pivot_table(index=['Employee_Name', 'Position_Type'], columns=['Date', 'Day_Name'], values='Hours_Worked', aggfunc='first', fill_value='')
            pivot_rate = df.pivot_table(index=['Employee_Name', 'Position_Type'], columns=['Date', 'Day_Name'], values='Daily_Hourly_Rate', aggfunc='first', fill_value='')
            pivot_total = df.pivot_table(index=['Employee_Name', 'Position_Type'], columns=['Date', 'Day_Name'], values='Total_Daily_Rate', aggfunc='first', fill_value='')

            weekly_totals = df.groupby(['Employee_Name', 'Position_Type', 'Week_Number'])[['Hours_Worked', 'Total_Daily_Rate']].sum().reset_index()

            grand_totals = df.groupby(['Employee_Name', 'Position_Type'])[['Hours_Worked', 'Total_Daily_Rate']].sum().reset_index()
            grand_totals = grand_totals.rename(columns={'Hours_Worked': 'Grand_Total_Hours', 'Total_Daily_Rate': 'Grand_Total_Rate'})

            # Create PDF
            doc = SimpleDocTemplate(response, pagesize=landscape(letter))
            elements = []
            elements.append(Spacer(1, 70))
            
            first_7_dates = pivot_hours.columns[:7]
            remaining_dates = pivot_hours.columns[7:]
            bold_font_style = ParagraphStyle(name='BoldFont', fontSize=6, fontName='Helvetica-Bold', alignment=1)

            def header_footer(canvas, doc, restaurant_name, start_date, end_date):
                canvas.saveState()

                # Load logo
                static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                image_path = os.path.join(static_dir, "restaurant_icon.png")

                try:
                    logo = Image(image_path, width=70, height=70)
                    logo.drawOn(canvas, 60, doc.pagesize[1] - 120)
                except Exception as e:
                    print(f"Error loading image: {e}")

                canvas.setFont('Helvetica', 25)
                canvas.drawCentredString(landscape(letter)[0] -400, landscape(letter)[1] - 80, "Biweekly Payroll")
                canvas.setFont('Helvetica', 18)
                canvas.drawCentredString(landscape(letter)[0] -400, landscape(letter)[1] - 110, f"{restaurant_name} - Week Ending {end_date.strftime('%Y-%m-%d')}")

                canvas.setFont('Helvetica', 8)
                canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
                canvas.drawRightString(landscape(letter)[0] - 100, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                canvas.restoreState()


            # Rebuild the header to insert Week 1 Total after 7 days
            header_data = [
                [Paragraph("Employee", bold_font_style), Paragraph("Position", bold_font_style), Paragraph("", bold_font_style)] + 
                [Paragraph(date.strftime('%Y-%m-%d'), bold_font_style) for date, day_name in first_7_dates] + [Paragraph("Week 1 Total", bold_font_style)] +
                [Paragraph(date.strftime('%Y-%m-%d'), bold_font_style) for date, day_name in remaining_dates] + [Paragraph("Week 2 Total", bold_font_style), Paragraph("Grand Total", bold_font_style)],

                [Paragraph("", bold_font_style), Paragraph("", bold_font_style), Paragraph("", bold_font_style)] + 
                [Paragraph(day_name, bold_font_style) for date, day_name in first_7_dates] + [Paragraph("", bold_font_style)] +
                [Paragraph(day_name, bold_font_style) for date, day_name in remaining_dates] + [Paragraph("", bold_font_style), Paragraph("", bold_font_style)]
            ]


            table_data = []

            for index, row in pivot_hours.iterrows():
                employee_name, position_type = index
                
                # Rate Row
                small_font_style = ParagraphStyle(name='SmallFont', fontSize=6, alignment=1)  # alignment=1 for middle alignment
                

                rate_row = [Paragraph(employee_name, small_font_style), position_type, Paragraph("Rate", small_font_style)] + \
                    [Paragraph(str(val), small_font_style) for val in list(pivot_rate.loc[index][:7])] + \
                    [Paragraph("", small_font_style)] + \
                    [Paragraph(str(val), small_font_style) for val in list(pivot_rate.loc[index][7:])] + \
                    [Paragraph("", small_font_style), Paragraph("", small_font_style)]


                # Get weekly totals
                employee_weekly_totals = weekly_totals[
                    (weekly_totals['Employee_Name'] == employee_name) &
                    (weekly_totals['Position_Type'] == position_type)
                ]

                week1_hours = employee_weekly_totals.loc[employee_weekly_totals['Week_Number'] == 1, 'Hours_Worked'].values
                week2_hours = employee_weekly_totals.loc[employee_weekly_totals['Week_Number'] == 2, 'Hours_Worked'].values
                week1_total = employee_weekly_totals.loc[employee_weekly_totals['Week_Number'] == 1, 'Total_Daily_Rate'].values
                week2_total = employee_weekly_totals.loc[employee_weekly_totals['Week_Number'] == 2, 'Total_Daily_Rate'].values

                # Insert weekly totals at correct positions (after first 7 days)
                employee_grand_totals = grand_totals[
                (grand_totals['Employee_Name'] == employee_name) &
                (grand_totals['Position_Type'] == position_type)
                ]
                grand_total_hours = employee_grand_totals['Grand_Total_Hours'].values[0] if not employee_grand_totals.empty else ""
                grand_total_rate = employee_grand_totals['Grand_Total_Rate'].values[0] if not employee_grand_totals.empty else ""

                # Insert weekly and grand totals at correct positions
                

                # Hours Row
                hours_row = [Paragraph("", small_font_style), Paragraph("", small_font_style), Paragraph("Hours", small_font_style)] + \
                [Paragraph(str(val), small_font_style) for val in list(row[:7])] + \
                [Paragraph(str(week1_hours[0]), bold_font_style) if week1_hours.size > 0 else Paragraph("", small_font_style)] + \
                [Paragraph(str(val), small_font_style) for val in list(row[7:])] + \
                [Paragraph(str(week2_hours[0]), bold_font_style) if week2_hours.size > 0 else Paragraph("", small_font_style), Paragraph(str(grand_total_hours), bold_font_style)]

                # Total Row
                total_row = [Paragraph("", small_font_style), Paragraph("", small_font_style), Paragraph("Total", bold_font_style)] + \
                [Paragraph(str(val), bold_font_style) for val in list(pivot_total.loc[index][:7])] + \
                [Paragraph(str(week1_total[0]), bold_font_style) if week1_total.size > 0 else Paragraph("", small_font_style)] + \
                [Paragraph(str(val), bold_font_style) for val in list(pivot_total.loc[index][7:])] + \
                [Paragraph(str(week2_total[0]), bold_font_style) if week2_total.size > 0 else Paragraph("", small_font_style), Paragraph(str(grand_total_rate), bold_font_style)]



                table_data.extend([rate_row, hours_row, total_row])

            # Create table with adjusted column widths
            max_rows_per_page = 21  # Adjust based on font size & page size
            combined_data = header_data.copy()
            for i in range(0, len(table_data), max_rows_per_page):
                page_data = header_data + table_data[i:i + max_rows_per_page]
                combined_data.extend(table_data[i:i + max_rows_per_page])
                # Dynamically calculate column widths based on content
                colWidths =[37, 32] + [25] + [41.5] * len(first_7_dates) + [28.5]+ [41.5] * len(remaining_dates) + [28.5]

                # Create the table with dynamic column widths
                table = Table(page_data, colWidths=colWidths)
                
                # Table styling
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),  # General font size for table
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 15),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ]))

                elements.append(table)
                elements.append(PageBreak())

            # Build PDF
            doc.build(elements, 
                onFirstPage=lambda canvas, doc: header_footer(canvas, doc, restaurant_name, start_date, end_date), 
                onLaterPages=lambda canvas, doc: header_footer(canvas, doc, restaurant_name, start_date, end_date)
            )
    except DatabaseError as db_error:
        raise ValueError(f"Database error: {str(db_error)}")
    if report_type.lower() == 'excel':
        try:
            # Clean combined_data (remove Paragraphs)
            cleaned_data = []
            for row in combined_data:
                cleaned_row = []
                for cell in row:
                    if hasattr(cell, 'text'):
                        cleaned_row.append(cell.text)
                    else:
                        cleaned_row.append(str(cell))  # fallback
                cleaned_data.append(cleaned_row)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Payroll_{location}"

            # Write cleaned data to sheet
            for row in cleaned_data:
                ws.append(row)

            # Define styles
            header_fill = PatternFill(start_color='91AAD6', end_color='91AAD6', fill_type='solid')
            bold_font = Font(bold=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

            # Apply styling to first two rows (headers)
            for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = bold_font
                    cell.border = thin_border

            # Apply styling to rows containing "Total"
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    # Check if any cell in the row contains "Total" (case insensitive)
                    if any(isinstance(cell.value, str) and 'total' in cell.value.lower() for cell in row):
                        cell.font = bold_font

            # Auto-size columns for better readability
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # Create HTTP response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=site_{site_id}_report.xlsx'
            wb.save(response)

            return response

        except Exception as e:
            return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)
    return response

def generate_weekly_activity_sales(request):
    try:
        location = request.GET.get('location')
        end_date_str = request.GET.get('data_date')
        report_type = request.GET.get('report_type')

        if not location:
            return JsonResponse({"error": "Location is required"}, status=400)
        if not end_date_str:
            return JsonResponse({"error": "Date is required"}, status=400)

        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD"}, status=400)

        try:
            user = User.objects.get(last_name=location)
        except User.DoesNotExist:
            return JsonResponse({"error": "No user found for the given location"}, status=404)

        site_id = user.id
        restaurant_name=user.first_name
        start_date = end_date - timedelta(days=6)

        with connection.cursor() as cursor:
            cursor.execute("SELECT MIN(activity_date) FROM activity_sales;")
            earliest_date = cursor.fetchone()[0]

        if earliest_date and start_date < earliest_date:
            return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=weekly_activity_sales_{site_id}_report.pdf'

        base_query = """
                    SELECT
                        s.activity_date AS Date,
                        s.food as "sales_food",
                        s.liquor AS "sales_liquor",
                        s.beer AS "sales_beer",
                        s.wine AS "sales_wine",
                        s.draft AS "sales_draft",
                        s.beverages_iced_beverages AS "sales_beverages_iced_beverages",
                        s.honest_to_goodness_fees AS "sales_honest_to_goodness_fees",
                        s.miscellaneous_income AS "sales_miscellaneous_income",
                        s.sales_total AS "sales_total",
                        s.sales_tax1 AS "sales_tax1",
                        s.sales_tax2 AS "sales_tax2",
                        s.customer_count AS "customer_count",
                        s.liquor_tax_1 AS "sales_liquor_tax_1",
                        f.skip_the_dishes AS "financial_skip_the_dishes",
                        f.gift_card_purchased AS "financial_gift_card_purchased",
                        f.gift_card_redeemed AS "financial_gift_card_redeemed",
                        f.just_eat AS "financial_just_eat",
                        f.amex AS "financial_amex",
                        f.din_club AS "financial_din_club",
                        f.discover AS "financial_discover",
                        f.master_card AS "financial_master_card",
                        f.visa AS "financial_visa",
                        f.dr_card AS "financial_dr_card",
                        f.financial_total AS "financial_financial_total",
                        f.actual_deposit_amt AS "financial_actual_deposit_amt",
                        f.required_deposit AS "financial_required_deposit",
                        po.food_paid_out AS "paid_out_food",
                        po.liquor_paid_out AS "paid_out_liquor",
                        po.supplies_paid_out AS "paid_out_supplies",
                        po.repair_maintenance AS "paid_out_repair_maintenance",
                        po.entertainment AS "paid_out_entertainment",
                        po.advertising AS "paid_out_advertising",
                        po.others AS "paid_out_others",
                        po.hst_gst AS "paid_out_hst_gst",
                        po.paid_out_total AS "paid_out_total",
                        pr.staff_meals AS "promo_staff_meals",
                        pr.manager_meals AS "promo_manager_meals",
                        pr.coupons_disc AS "promo_coupons_disc",
                        pr.chucks_bucks AS "promo_chucks_bucks",
                        pr.lsm_1 AS "promo_lsm_1",
                        pr.lsm_2 AS "promo_lsm_2",
                        pr.manager_prom AS "promo_manager_prom",
                        pr.qsa_complaints AS "promo_qsa_complaints",
                        pr.promotion_total AS "promo_promotion_total"
                    FROM
                        activity_sales s
                    LEFT JOIN
                        activity_purchase p ON s.site_id = p.site_id AND s.activity_date = p.activity_date
                    LEFT JOIN
                        activity_paid_out po ON s.site_id = po.site_id AND s.activity_date = po.activity_date
                    LEFT JOIN
                        activity_financial f ON s.site_id = f.site_id AND s.activity_date = f.activity_date
                    LEFT JOIN
                        activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
                    LEFT JOIN
                        auth_user u ON s.site_id = u.id
                    WHERE s.activity_date BETWEEN %s AND %s AND s.site_id = %s;
                    """

        query_params = [start_date, end_date, site_id]
        date_range = pd.date_range(start=start_date, end=end_date)
        all_dates = [date.strftime('%Y-%m-%d') for date in date_range]
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                results = cursor.fetchall()
                column_names_from_db = [col[0] for col in cursor.description]

                # Create DataFrame with all dates to ensure completeness
                df = pd.DataFrame(index=all_dates, columns=column_names_from_db)
                df = df.fillna(0)  # Fill missing values with 0
                
                # Populate with actual data
                for row in results:
                    date_str = row[column_names_from_db.index('Date')].strftime('%Y-%m-%d')
                    df.loc[date_str] = [0 if value is None else value for value in row]

                df['Date'] = pd.to_datetime(df.index)
                df = df.sort_values(by='Date')

                # --- SALES TABLE ---
                cash_sales_cols = [
                    'sales_food', 'sales_beverages_iced_beverages', 'sales_liquor',
                    'sales_beer', 'sales_wine', 'sales_draft',
                    'sales_honest_to_goodness_fees', 'sales_miscellaneous_income',
                    'sales_total', 'sales_tax1', 'sales_tax2'
                ]

                # Prepare Sales Data
                sales_df = df[['Date'] + cash_sales_cols].copy()
                sales_df['Date'] = sales_df['Date'].dt.strftime('%Y-%m-%d')
                sales_by_date = sales_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                sales_by_date['Total'] = sales_by_date.sum(axis=1)  # Add Total column

                # Build the sales table data
                sales_deposits_data = [["Category"] + all_dates + ["Total"]]  # Header with all dates + Total

                category_map = {
                    'sales_food': 'Food',
                    'sales_beverages_iced_beverages': 'Beverages',
                    'sales_liquor': 'Liquor',
                    'sales_beer': 'Beer',
                    'sales_wine': 'Wine',
                    'sales_draft': 'Draft',
                    'sales_honest_to_goodness_fees': 'Honest To Goodness Fee',
                    'sales_miscellaneous_income': 'Miscellaneous Income',
                    'sales_total': 'Net Sales',
                }

                def format_currency(value):
                    try:
                        if isinstance(value, (str, bytes)):
                            value = value.replace('$', '').replace(',', '')
                            return f"${float(value):,.2f}"
                        elif isinstance(value, Decimal):
                            return f"${float(value):,.2f}"
                        else:
                            return f"${float(value or 0):,.2f}"
                    except (ValueError, TypeError):
                        return "$0.00"

                # Add data rows with totals
                for col, display_name in category_map.items():
                    values = [format_currency(sales_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(sales_by_date.loc[col, 'Total'])
                    sales_deposits_data.append([display_name] + values + [total])

                sales_deposits_data.append([""] + ["" for _ in all_dates])  # Blank line, optional

                sales_tax_map = {
                    'sales_tax1': 'Sales Tax 1',
                    'sales_tax2': 'Sales Tax 2',
                }

                for col, display_name in sales_tax_map.items():
                    values = [format_currency(sales_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(sales_by_date.loc[col, 'Total'])
                    sales_deposits_data.append([display_name] + values + [total])

                # Calculate and add totals
                total_tax = sales_by_date.loc['sales_tax1'] + sales_by_date.loc['sales_tax2']
                total_tax_total = total_tax.sum()/2
                formatted_total_tax = ["Total Tax"] + [format_currency(total_tax[date]) for date in all_dates] + [format_currency(total_tax_total)]
                sales_deposits_data.append(formatted_total_tax)

                gross_sales_values = sales_by_date.loc['sales_total'] + total_tax
                gross_sales_total = gross_sales_values.sum()/2
                formatted_total_gross = ["Gross Sales"] + [format_currency(gross_sales_values[date]) for date in all_dates] + [format_currency(gross_sales_total)]
                sales_deposits_data.append(formatted_total_gross)

                # --- FINANCIAL DATA ---
                financial_cols = [
                    'financial_visa', 'financial_master_card', 'financial_amex',
                    'financial_din_club', 'financial_discover', 'financial_dr_card',
                    'financial_just_eat', 'financial_skip_the_dishes'
                ]

                financial_df = df[['Date'] + financial_cols].copy()
                financial_by_date = financial_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                financial_by_date['Total'] = financial_by_date.sum(axis=1)  # Add Total column

                financial_deposits_data = []

                payment_methods = {
                    'financial_visa': 'Visa',
                    'financial_master_card': 'Master Card',
                    'financial_amex': 'American Express',
                    'financial_din_club': 'Diners Club',
                    'financial_discover': 'Discover',
                    'financial_dr_card': 'Debit Card'
                }

                for col, display_name in payment_methods.items():
                    values = [format_currency(financial_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(financial_by_date.loc[col, 'Total'])
                    financial_deposits_data.append([display_name] + values + [total])

                # Calculate totals with Total column
                card_totals = financial_by_date.loc[list(payment_methods.keys())].sum()
                total_Credit_Debit_Card_row = ["Credit & Debit Card Total"] + [format_currency(card_totals[date]) for date in all_dates] + [format_currency(card_totals.sum()/2)]
                financial_deposits_data.append(total_Credit_Debit_Card_row)
                financial_deposits_data.append([""] + ["" for _ in all_dates] + [""])

                partner_totals = financial_by_date.loc[['financial_just_eat', 'financial_skip_the_dishes']].sum()
                Order_Partner_Total_row = ["Order Partner Total"] + [format_currency(partner_totals[date]) for date in all_dates] + [format_currency(partner_totals.sum()/2)]
                financial_deposits_data.append(Order_Partner_Total_row)

                # --- CASH PAYOUT DATA ---
                payout_cols = [
                    'promo_staff_meals', 'promo_manager_meals', 'promo_coupons_disc',
                    'promo_manager_prom', 'promo_qsa_complaints', 'promo_chucks_bucks',
                    'promo_lsm_1', 'promo_lsm_2', 'financial_gift_card_purchased',
                    'financial_gift_card_redeemed', 'paid_out_total', 'financial_actual_deposit_amt',
                    'customer_count'
                ]

                payout_df = df[['Date'] + payout_cols].copy()
                payout_by_date = payout_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                payout_by_date['Total'] = payout_by_date.sum(axis=1)  # Add Total column

                cash_payout_financial_data = []

                payout_categories = {
                    'promo_staff_meals': 'Staff Meals',
                    'promo_manager_meals': 'Manager Meals',
                    'promo_coupons_disc': 'Coupons And Discounts',
                    'promo_manager_prom': 'Manager Promotions',
                    'promo_qsa_complaints': 'QSA Complaints',
                    'promo_chucks_bucks': "Chuck's Bucks",
                    'promo_lsm_1': 'LSM 1',
                    'promo_lsm_2': 'LSM 2'
                }

                for col, display_name in payout_categories.items():
                    values = [format_currency(payout_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(payout_by_date.loc[col, 'Total'])
                    cash_payout_financial_data.append([display_name] + values + [total])

                # Calculate Total Discounts with Total column
                total_discounts = payout_by_date.loc[list(payout_categories.keys())].sum()
                total_discounts_total = total_discounts.sum()/2
                cash_payout_financial_data.append(["Total Discounts"] + [format_currency(total_discounts[date]) for date in all_dates] + [format_currency(total_discounts_total)])

                # Calculate Net Cash with Total column
                net_cash = []
                for date in all_dates:
                    gross = float(sales_by_date.loc['sales_total', date] or 0) + \
                            float(sales_by_date.loc['sales_tax1', date] or 0) + \
                            float(sales_by_date.loc['sales_tax2', date] or 0)
                    
                    cards = float(financial_by_date.loc[list(payment_methods.keys()), date].sum() or 0)
                    partners = float(financial_by_date.loc[['financial_just_eat', 'financial_skip_the_dishes'], date].sum() or 0)
                    discounts = float(payout_by_date.loc[list(payout_categories.keys()), date].sum() or 0)
                    
                    net_cash.append(gross - cards - partners - discounts)

                net_cash_total = sum(net_cash)
                net_cash_row = ["Net Cash"] + [format_currency(v) for v in net_cash] + [format_currency(net_cash_total)]
                cash_payout_financial_data.append(net_cash_row)
                cash_payout_financial_data.append([""] + ["" for _ in all_dates])

                # Add other items with totals
                for col, display_name in {
                    'financial_gift_card_purchased': 'Gift Card Amount Purchased',
                    'financial_gift_card_redeemed': 'Gift Card Amount Redeemed',
                    'paid_out_total': 'Cash Paid Out'
                }.items():
                    values = [format_currency(payout_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(payout_by_date.loc[col, 'Total'])
                    cash_payout_financial_data.append([display_name] + values + [total])

                # Calculate Cash Adjustments with Total column
                cash_adjustments = payout_by_date.loc[['financial_gift_card_purchased', 'financial_gift_card_redeemed', 'paid_out_total']].sum()
                cash_adjustments_total = cash_adjustments.sum()/2
                cash_adjustments_row = ["Cash Adjustments"] + [format_currency(cash_adjustments[date]) for date in all_dates] + [format_currency(cash_adjustments_total)]
                cash_payout_financial_data.append(cash_adjustments_row)
                cash_payout_financial_data.append([""] + ["" for _ in all_dates] + [""])

                # Calculate final amounts with totals
                calculated_deposit = [net_cash[i] - float(cash_adjustments[all_dates[i]] or 0) for i in range(len(all_dates))]
                calculated_deposit_total = sum(calculated_deposit)
                
                actual_deposit = [float(payout_by_date.loc['financial_actual_deposit_amt', date] or 0) for date in all_dates]
                actual_deposit_total = sum(actual_deposit)
                
                deposit_diff = [calculated_deposit[i] - actual_deposit[i] for i in range(len(all_dates))]
                deposit_diff_total = sum(deposit_diff)

                cash_payout_financial_data.append(["Calculated Deposit Amount"] + [format_currency(v) for v in calculated_deposit] + [format_currency(calculated_deposit_total)])
                cash_payout_financial_data.append(["Actual Deposit Amount"] + [format_currency(v) for v in actual_deposit] + [format_currency(actual_deposit_total)])
                cash_payout_financial_data.append(["Deposit Over/Under"] + [format_currency(v) for v in deposit_diff] + [format_currency(deposit_diff_total)])

                cash_payout_financial_data.append([""] + ["" for _ in all_dates])  # Blank line, optional
                customer_counts = [int(payout_by_date.loc['customer_count', date]) for date in all_dates]
                cash_payout_financial_data.append(["Customer Count"] + customer_counts)
                # --- PDF GENERATION ---
                
                doc = SimpleDocTemplate(response, pagesize=landscape(letter))
                elements = []
                elements.append(Spacer(1, 20))
                styles = getSampleStyleSheet()

                def header_footer(canvas, doc, restaurant_name, start_date, end_date):
                    canvas.saveState()
                            
                    try:
                        # Import Image at the function level
                        from reportlab.platypus import Image
                                
                                # Use Django's static file finder to locate the image
                        from django.contrib.staticfiles.finders import find
                        image_path = find('site_panel/images/restaurant_icon.png')
                                
                        if image_path:
                            try:
                                img = Image(image_path, width=70, height=70)
                                # Draw the image at the top-left corner
                                img.drawOn(canvas, 50, landscape(letter)[1] - 80)
                            except Exception as img_error:
                                print(f"Error drawing image: {str(img_error)}")
                                # Fallback text if image fails to draw
                                canvas.setFont('Helvetica', 10)
                                canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                        else:
                            print("Image not found in static files")
                            # Fallback text if image not found
                            canvas.setFont('Helvetica', 10)
                            canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                    except Exception as e:
                        print(f"Error in image handling: {str(e)}")
                        # Final fallback if anything else goes wrong
                        canvas.setFont('Helvetica', 10)
                        canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")

                    canvas.setFont('Helvetica', 25)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 50, "Weekly Activity Sales")
                    canvas.setFont('Helvetica', 18)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 80, f"{restaurant_name} - {start_date} to {end_date}")

                    canvas.setFont('Helvetica', 8)
                    canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
                    canvas.drawRightString(landscape(letter)[0] - 30, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    canvas.restoreState()

                my_light_blue = HexColor('#91aad6')

                # Calculate number of columns (should be same for all tables now)
                num_columns = len(all_dates) + 2  # Category + dates + Total

                # Adjust column widths
                available_width = landscape(letter)[0] - 2 * inch
                first_col_width = 1.8 * inch
                date_col_width = (available_width - first_col_width - 1 * inch) / len(all_dates)  # Space for Total column
                col_widths = [first_col_width] + [date_col_width] * len(all_dates) + [1 * inch]  # Total column width

                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), my_light_blue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),  # Highlight Total column
                    ('FONTNAME', (0, len(sales_deposits_data)-6), 
                    (-1, len(sales_deposits_data)-6), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)-6), 
                    (-1, len(sales_deposits_data)-6), my_light_blue),
                    ('FONTNAME', (0, len(sales_deposits_data)-1), 
                    (-1, len(sales_deposits_data)-1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)-1), 
                    (-1, len(sales_deposits_data)-1), my_light_blue),
                    ('FONTNAME', (0, len(sales_deposits_data)+len(financial_deposits_data)-2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)-2), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)+len(financial_deposits_data)-2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)-2), my_light_blue),
                    ('FONTNAME', (0, len(sales_deposits_data)+len(financial_deposits_data)), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)+len(financial_deposits_data)), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)), my_light_blue),
                    
                    # Net Cash row
                    ('FONTNAME', (0, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(net_cash_row)+2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(net_cash_row)+2), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(net_cash_row)+2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(net_cash_row)+2), my_light_blue),
                    
                    # Cash Adjustments row
                    ('FONTNAME', (0, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), my_light_blue),
                                        ('FONTNAME', (0, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), 
                    (-1, len(sales_deposits_data)+len(financial_deposits_data)+cash_payout_financial_data.index(cash_adjustments_row)+2), my_light_blue),
                ])

                # Combine all data ensuring consistent column counts
                combined_data = []
                
                # Add sales data
                combined_data.extend(sales_deposits_data)
                combined_data.append([""] * num_columns)  # Blank row
                
                # Add financial data
                combined_data.extend(financial_deposits_data)
                combined_data.append([""] * num_columns)  # Blank row
                
                # Add payout data
                combined_data.extend(cash_payout_financial_data)

                # Create the table
                combined_table = Table(combined_data, colWidths=col_widths)
                combined_table.setStyle(table_style)
                elements.append(combined_table)

                doc.build(elements, 
                onFirstPage=lambda canvas, doc: header_footer(canvas, doc, restaurant_name, start_date, end_date), 
                onLaterPages=lambda canvas, doc: header_footer(canvas, doc, restaurant_name, start_date, end_date)
            )

        except Exception as e:
            return JsonResponse({"error": f"Error generating financial report: {str(e)}"}, status=500)
        if report_type.lower() == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                wb = Workbook()
                ws = wb.active
                
                # Define styling
                highlight_style = PatternFill(
                    start_color='91AAD6',
                    end_color='91AAD6',
                    fill_type='solid'
                )
                bold_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                try:
                    from openpyxl.drawing.image import Image
                    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                    image_path = os.path.join(static_dir, "restaurant_icon.png")
                    
                    # Verify image exists before trying to insert
                    if os.path.exists(image_path):
                        img = Image(image_path)
                        img.width = 70
                        img.height = 70
                        # Add image after text to avoid positioning issues
                        ws.add_image(img, 'A1')  # Positioned to the right of the title
                except Exception as img_error:
                    print(f"Image insertion skipped: {str(img_error)}")

                # Add title and date first (we'll add image after to avoid issues)
                ws['C1'] = "Weekly Activity Sales"
                ws['C1'].font = Font(size=20, bold=True)
                
                ws['C2'] = f"{restaurant_name} - {start_date} to {end_date}"
                ws['C2'].font = Font(size=14)
                
                # Try to add image after setting up basic structure

                # Start data from row 4
                start_row = 5
                
                # Terms that trigger row highlighting
                highlight_terms = {
                    'Category',
                    'Credit & Debit Card Total',
                    'Gross Sales',
                    'Net Sales',
                    'Order Partner Total',
                    'Total Discounts',
                    'Cash Adjustments',
                    'Customer Count'
                }

                # Add the data table with proper borders
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # Apply highlighting to important rows
                        if col_idx == 1 and str(value).strip() in highlight_terms:
                            for col in range(1, len(row) + 1):
                                ws.cell(row=row_idx, column=col).fill = highlight_style
                                ws.cell(row=row_idx, column=col).font = bold_font

                # Safe column auto-sizing
                for col in ws.columns:
                    try:
                        column_letter = col[0].column_letter
                        max_length = 0
                        for cell in col:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                continue
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
                    except:
                        continue

                wb.save(response)
                return response
                
            except Exception as e:
                return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
def generate_weekly_promo(request):
    try:
        location = request.GET.get('location')
        end_date_str = request.GET.get('data_date')
        report_type = request.GET.get('report_type')

        if not location:
            return JsonResponse({"error": "Location is required"}, status=400)
        if not end_date_str:
            return JsonResponse({"error": "Date is required"}, status=400)

        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD"}, status=400)
        

        try:
            user = User.objects.get(last_name=location)
        except User.DoesNotExist:
            return JsonResponse({"error": "No user found for the given location"}, status=404)

        site_id = user.id
        restaurant_name = user.first_name
        start_date = end_date - timedelta(days=6)

        with connection.cursor() as cursor:
            cursor.execute("SELECT MIN(activity_date) FROM activity_sales;")
            earliest_date = cursor.fetchone()[0]

        if earliest_date and start_date < earliest_date:
            return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=weekly_activity_sales_{site_id}_report.pdf'

        base_query = """
                    SELECT
                        s.activity_date AS "Date",
                        s.sales_total AS "sales_total",
                        pr.staff_meals AS "promo_staff_meals",
                        pr.manager_meals AS "promo_manager_meals",
                        pr.coupons_disc AS "promo_coupons_disc",
                        pr.chucks_bucks AS "promo_chucks_bucks",
                        pr.lsm_1 AS "promo_lsm_1",
                        pr.lsm_2 AS "promo_lsm_2",
                        pr.manager_prom AS "promo_manager_prom",
                        pr.qsa_complaints AS "promo_qsa_complaints",
                        pr.promotion_total AS "promo_promotion_total"
                    FROM
                        activity_sales s
                    LEFT JOIN
                        activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
                    WHERE s.activity_date BETWEEN %s AND %s AND s.site_id = %s;
                    """

        query_params = [start_date, end_date, site_id]
        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                results = cursor.fetchall()
                results = [tuple(0 if value is None else value for value in row) for row in results]
                column_names_from_db = [col[0] for col in cursor.description]

                if 'Date' not in column_names_from_db:
                    raise ValueError("Your base_query must return a 'Date' column.")

                df = pd.DataFrame(results, columns=column_names_from_db)
                df = df.fillna(0)
                df['Date'] = pd.to_datetime(df['Date'])
                
                # Generate complete date range
                date_range = pd.date_range(start=start_date, end=end_date)
                all_dates = [date.strftime('%Y-%m-%d') for date in date_range]

                payout_cols = [
                    'promo_staff_meals', 'promo_manager_meals', 'promo_coupons_disc',
                    'promo_manager_prom', 'promo_chucks_bucks', 'promo_qsa_complaints',
                    'promo_lsm_1', 'promo_lsm_2', 'sales_total',
                ]

                # Create a complete date-indexed DataFrame first
                payout_df = df[['Date'] + payout_cols].copy()
                payout_df['Date'] = payout_df['Date'].dt.strftime('%Y-%m-%d')

                # Reindex with all dates to ensure completeness
                payout_by_date = payout_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                payout_by_date['Total'] = payout_by_date.sum(axis=1)

            
                # Mapping of column names to display names
                payout_categories = {
                    'promo_staff_meals': 'Staff Meals',
                    'promo_manager_meals': 'Manager Meals',
                    'promo_coupons_disc': 'Coupons And Discounts',
                    'promo_manager_prom': 'Manager Promotions',
                    'promo_chucks_bucks': "Chuck's Bucks",
                    'promo_qsa_complaints': 'QSA Complaints',
                    'promo_lsm_1': 'LSM 1',
                    'promo_lsm_2': 'LSM 2'
                }

                cash_payout_financial_data = [["Category"] + all_dates + ["Total"]]

                def format_currency(value):
                    try:
                        return f"${float(value):,.2f}"
                    except (ValueError, TypeError):
                        return "$0.00"

                # For each promotion category, format each value
                for col, display_name in payout_categories.items():
                    # Get raw values from the pivot table (as a list)
                    values = payout_by_date.loc[col].tolist()
                    # Format each value with a dollar sign and commas
                    formatted_values = [format_currency(v) for v in values]
                    cash_payout_financial_data.append([display_name] + formatted_values)


                # Calculate and add totals
                total_discounts = payout_by_date.loc[list(payout_categories.keys())].sum()
                formatted_totals = [format_currency(v) for v in total_discounts.tolist()]
                total_sales = payout_by_date.loc['sales_total']
                formatted_totals = [format_currency(v) for v in total_sales.tolist()]

                percentages = []
                for disc, sales in zip(total_discounts, total_sales):
                    if sales != 0:
                        percentage = (disc / sales) * 100
                    else:
                        percentage = 0
                    percentages.append(f"{percentage:.2f}%")

                
                cash_payout_financial_data.append(["Total Discounts"] + formatted_totals)
                
                cash_payout_financial_data.append([""] + ["" for _ in all_dates])
                cash_payout_financial_data.append(["Total Sales"] + formatted_totals)
                cash_payout_financial_data.append(["Discount % of Sales"] + percentages)

                def header_footer(canvas, doc):
                    canvas.saveState()
                    try:
                            # Use Django's static file finder to locate the image
                        from reportlab.platypus import Image
                                
                                # Use Django's static file finder to locate the image
                        from django.contrib.staticfiles.finders import find
                        image_path = find('site_panel/images/restaurant_icon.png')
                                
                        if image_path:
                            try:
                                img = Image(image_path, width=70, height=70)
                                # Draw the image at the top-left corner
                                img.drawOn(canvas, 50, landscape(letter)[1] - 80)
                            except Exception as img_error:
                                print(f"Error drawing image: {str(img_error)}")
                                # Fallback text if image fails to draw
                                canvas.setFont('Helvetica', 10)
                                canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                        else:
                            print("Image not found in static files")
                            # Fallback text if image not found
                            canvas.setFont('Helvetica', 10)
                            canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                    except Exception as e:
                        print(f"Error in image handling: {str(e)}")
                        # Final fallback if anything else goes wrong
                        canvas.setFont('Helvetica', 10)
                        canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")

                    # Set font and draw the report header text
                    canvas.setFont('Helvetica', 25)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 50, "Weekly Promo Report")
                    canvas.setFont('Helvetica', 18)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 80, f"{restaurant_name} - {start_date} to {end_date}")

                    # Draw footer information (page number, timestamp)
                    canvas.setFont('Helvetica', 8)
                    canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
                    canvas.drawRightString(landscape(letter)[0] - 30, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    canvas.restoreState()


                doc = SimpleDocTemplate(response, pagesize=landscape(letter))
                elements = []
                elements.append(Spacer(1, 20))
                styles = getSampleStyleSheet()

                # Add the date header just once before the first table
                elements.append(Paragraph(f"", styles['Heading2']))\
                
                my_light_blue = HexColor('#91aad6')

                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), my_light_blue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),  # Increased font size for the header
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ])

                # Calculate column widths - make all tables the same width
                num_columns = len(cash_payout_financial_data[0])
                available_width = landscape(letter)[0] - 2 * inch  # Leave 1 inch margins on each side
                first_col_width = 1.8 * inch  # Wider first column for labels
                other_col_width = (available_width - first_col_width) / (num_columns - 1)
                col_widths = [first_col_width] + [other_col_width] * (num_columns - 1)

                # Combine all tables into one continuous table
                combined_data = []

                # Add first table
                combined_data.extend(cash_payout_financial_data)

                # Create single combined table
                combined_table = Table(combined_data, colWidths=col_widths)

                # Apply style to the entire combined table
                combined_table.setStyle(table_style)

                # Add the combined table to elements
                elements.append(combined_table)

                doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)

        except Exception as e:
            return JsonResponse({"error": f"Error generating financial report: {str(e)}"}, status=500)
        if report_type.lower() == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                wb = Workbook()
                ws = wb.active
                
                # Define styling
                highlight_style = PatternFill(
                    start_color='91AAD6',  # Your light blue color
                    end_color='91AAD6',
                    fill_type='solid'
                )
                bold_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                try:
                    from openpyxl.drawing.image import Image
                    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                    image_path = os.path.join(static_dir, "restaurant_icon.png")
                    
                    # Verify image exists before trying to insert
                    if os.path.exists(image_path):
                        img = Image(image_path)
                        img.width = 70
                        img.height = 70
                        # Add image after text to avoid positioning issues
                        ws.add_image(img, 'A1')  # Positioned to the right of the title
                except Exception as img_error:
                    print(f"Image insertion skipped: {str(img_error)}")

                # Add title and date first (we'll add image after to avoid issues)
                ws['C1'] = "Weekly Promo Report"
                ws['C1'].font = Font(size=20, bold=True)
                
                ws['C2'] = f"{restaurant_name} - {start_date} to {end_date}"
                ws['C2'].font = Font(size=14)
                
                # Try to add image after setting up basic structure

                # Start data from row 4
                start_row = 5
                
                # Terms that trigger row highlighting
                highlight_terms = {
                    'Category',
                    'Total Sales(A)',
                    'Total Promotions(B)',
                }
                
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # Check first column for highlight terms
                        if col_idx == 1 and str(value).strip() in highlight_terms:
                            # Apply to entire row
                            for col in range(1, len(row) + 1):
                                ws.cell(row=row_idx, column=col).fill = highlight_style
                                ws.cell(row=row_idx, column=col).font = bold_font
                
                wb.save(response)
                return response
                
            except Exception as e:
                return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)

        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
def generate_weekly_flash_report(request):
    try:
        location = request.GET.get('location')
        end_date_str = request.GET.get('data_date')
        report_type = request.GET.get('report_type')

        if not location:
            return JsonResponse({"error": "Location is required"}, status=400)
        if not end_date_str:
            return JsonResponse({"error": "Date is required"}, status=400)

        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD"}, status=400)

        try:
            user = User.objects.get(last_name=location)
        except User.DoesNotExist:
            return JsonResponse({"error": "No user found for the given location"}, status=404)

        site_id = user.id
        restaurant_name=user.first_name
        start_date = end_date - timedelta(days=6)

        with connection.cursor() as cursor:
            cursor.execute("SELECT MIN(activity_date) FROM activity_sales;")
            earliest_date = cursor.fetchone()[0]

        if earliest_date and start_date < earliest_date:
            return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=weekly_activity_sales_{site_id}_report.pdf'

        base_query = """
                    SELECT
                        s.activity_date AS Date,
                        s.food as "sales_food",
                        s.liquor AS "sales_liquor",
                        s.beer AS "sales_beer",
                        s.wine AS "sales_wine",
                        s.draft AS "sales_draft",
                        s.beverages_iced_beverages AS "sales_beverages_iced_beverages",
                        s.honest_to_goodness_fees AS "sales_honest_to_goodness_fees",
                        s.miscellaneous_income AS "sales_miscellaneous_income",
                        s.sales_total AS "sales_total",
                        s.sales_tax1 AS "sales_tax1",
                        s.sales_tax2 AS "sales_tax2",
                        s.customer_count AS "customer_count",
                        s.liquor_tax_1 AS "sales_liquor_tax_1",
                        pr.staff_meals AS "promo_staff_meals",
                        pr.manager_meals AS "promo_manager_meals",
                        pr.coupons_disc AS "promo_coupons_disc",
                        pr.chucks_bucks AS "promo_chucks_bucks",
                        pr.lsm_1 AS "promo_lsm_1",
                        pr.lsm_2 AS "promo_lsm_2",
                        pr.manager_prom AS "promo_manager_prom",
                        pr.qsa_complaints AS "promo_qsa_complaints",
                        pr.promotion_total AS "promo_promotion_total"
                    FROM
                        activity_sales s
                    LEFT JOIN
                        activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
                    LEFT JOIN
                        auth_user u ON s.site_id = u.id
                    WHERE s.activity_date BETWEEN %s AND %s AND s.site_id = %s;
                    """

        query_params = [start_date, end_date, site_id]
        date_range = pd.date_range(start=start_date, end=end_date)
        all_dates = [date.strftime('%Y-%m-%d') for date in date_range]

        all_dates_dt = pd.to_datetime(all_dates)
        date_headers = [d.strftime('%Y-%m-%d') for d in all_dates_dt]
        day_headers = [d.strftime('%A') for d in all_dates_dt]
        
        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                results = cursor.fetchall()
                column_names_from_db = [col[0] for col in cursor.description]

                # Create DataFrame with all dates to ensure completeness
                df = pd.DataFrame(index=all_dates, columns=column_names_from_db)
                
                # Convert all numeric columns to float, keeping Date as string
                for col in column_names_from_db:
                    if col != 'Date':
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
                
                # Populate with actual data
                for row in results:
                    date_str = row[column_names_from_db.index('Date')].strftime('%Y-%m-%d')
                    row_data = []
                    for value in row:
                        if value is None:
                            row_data.append(0.0)
                        elif isinstance(value, (Decimal, float, int)):
                            row_data.append(float(value))
                        else:
                            row_data.append(value)  # For non-numeric values like dates
                    df.loc[date_str] = row_data
                
                # Convert Date column to datetime
                df['Date'] = pd.to_datetime(df.index)
                df = df.sort_values(by='Date')

                # --- SALES TABLE ---
                cash_sales_cols = [
                    'sales_food', 'sales_beverages_iced_beverages', 'sales_liquor',
                    'sales_beer', 'sales_wine', 'sales_draft',
                    'sales_honest_to_goodness_fees', 'sales_miscellaneous_income',
                    'sales_total', 'customer_count'
                ]

                # Prepare Sales Data
                sales_df = df[['Date'] + cash_sales_cols].copy()
                sales_df['Date'] = sales_df['Date'].dt.strftime('%Y-%m-%d')
                sales_by_date = sales_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                sales_by_date['Total'] = sales_by_date.sum(axis=1)  # Add Total column

                # Build the sales table data
                sales_deposits_data = [
                    ["Category"] + date_headers + ["Total"],
                    [""] + day_headers + [""]  # Empty first cell, day names for each date, empty Total column
                ]

                category_map = {
                    'sales_food': 'Food',
                    'sales_beverages_iced_beverages': 'Beverages',
                    'sales_liquor': 'Liquor',
                    'sales_beer': 'Beer',
                    'sales_wine': 'Wine',
                    'sales_draft': 'Draft',
                    'sales_honest_to_goodness_fees': 'Honest To Goodness Fee',
                    'sales_miscellaneous_income': 'Miscellaneous Income',
                    'sales_total': 'Total Sales(A)',
                }

                def format_currency(value):
                    try:
                        if isinstance(value, (str, bytes)):
                            value = value.replace('$', '').replace(',', '')
                            return f"${float(value):,.2f}"
                        elif isinstance(value, Decimal):
                            return f"${float(value):,.2f}"
                        else:
                            return f"${float(value or 0):,.2f}"
                    except (ValueError, TypeError):
                        return "$0.00"

                # Add data rows with totals
                for col, display_name in category_map.items():
                    values = [format_currency(sales_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(sales_by_date.loc[col, 'Total'])
                    sales_deposits_data.append([display_name] + values + [total])

                sales_deposits_data.append([""] + ["" for _ in all_dates])  # Blank line, optional

                # --- CASH PAYOUT DATA ---
                payout_cols = [
                    'promo_staff_meals', 'promo_manager_meals', 'promo_coupons_disc',
                    'promo_manager_prom', 'promo_qsa_complaints', 'promo_chucks_bucks',
                    'promo_lsm_1', 'promo_lsm_2'
                ]

                payout_df = df[['Date'] + payout_cols].copy()
                payout_by_date = payout_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                payout_by_date['Total'] = payout_by_date.sum(axis=1)  # Add Total column

                cash_payout_financial_data = []

                payout_categories = {
                    'promo_staff_meals': 'Staff Meals',
                    'promo_manager_meals': 'Manager Meals',
                    'promo_coupons_disc': 'Coupons And Discounts',
                    'promo_manager_prom': 'Manager Promotions',
                    'promo_qsa_complaints': 'QSA Complaints',
                    'promo_chucks_bucks': "Chuck's Bucks",
                    'promo_lsm_1': 'LSM 1',
                    'promo_lsm_2': 'LSM 2'
                }

                for col, display_name in payout_categories.items():
                    values = [format_currency(payout_by_date.loc[col, date]) for date in all_dates]
                    total = format_currency(payout_by_date.loc[col, 'Total'])
                    cash_payout_financial_data.append([display_name] + values + [total])

                # Calculate Total Discounts with Total column
                total_discounts = payout_by_date.loc[list(payout_categories.keys())].sum()
                total_discounts_total = total_discounts.sum()/2
                cash_payout_financial_data.append(["Total Promotions(B)"] + [format_currency(total_discounts[date]) for date in all_dates] + [format_currency(total_discounts_total)])
                # Calculate Net Sales (A-B) by subtracting Total Promotions from Total Sales
                cash_payout_financial_data.append([""] + ["" for _ in all_dates])  
                # Calculate Net Sales (A-B) by subtracting Total Promotions from Total Sales
                net_sales_values = [
                    format_currency(
                        float(sales_by_date.loc['sales_total', date]) - 
                        float(total_discounts[date])
                    )
                    for date in all_dates
                ]
                net_sales_total = format_currency(
                    float(sales_by_date.loc['sales_total', 'Total']) - 
                    float(total_discounts_total)
                )

                # Add Net Sales row to the data
                cash_payout_financial_data.append(["Net Sales (A-B)"] + net_sales_values + [net_sales_total])
                cash_payout_financial_data.append([""] + ["" for _ in all_dates]) 
                # Add boundary row
                cash_payout_financial_data.append(["Customers"] + [""] * len(all_dates))

                # Customer Count row
                customer_total = sales_by_date.loc["customer_count", 'Total']
                customer_row = ["Customer Count"] + [sales_by_date.loc["customer_count", date] for date in all_dates] + [customer_total]
                cash_payout_financial_data.append(customer_row)

                # Average Cheque / Person row
                avg_cheque_row = ["Average Cheque/ Person"]
                for date in all_dates:
                    sales = float(sales_by_date.loc['sales_total', date])
                    customers = sales_by_date.loc['customer_count', date]
                    avg = sales / customers if customers != 0 else 0
                    avg_cheque_row.append(format_currency(avg))

                # Add the total column
                sales_total = float(sales_by_date.loc['sales_total', 'Total'])
                avg_total = sales_total / customer_total if customer_total != 0 else 0
                avg_cheque_row.append(format_currency(avg_total))

                cash_payout_financial_data.append(avg_cheque_row)
                cash_payout_financial_data.append([""] + ["" for _ in all_dates]) 
                cash_payout_financial_data.append(["Royalties"] + [""] * len(all_dates))
                cash_payout_financial_data.append(
                    ["Royalty Fee Description"] + 
                    [f"3.75% on \nweekly gross \nsales up to \n$38,000\n 7% on portion \nof sales \nabove $38,000"] * len(all_dates)
                )
                net_sales = float(net_sales_total.replace('$', '').replace(',', ''))
                if net_sales > 38000:
                    above_base_value = net_sales-38000
                    royalty_amt = (38000*0.0375) + (above_base_value *0.07)
                else:
                    royalty_amt = net_sales * 0.0375
                royalty_row = ["Amount"] + ["$0.00" for _ in all_dates] + [format_currency(royalty_amt)]
                cash_payout_financial_data.append(royalty_row)
                ontario_hst = royalty_amt * 0.13
                royalty_tax1_row = ["Tax 1: Ontario HST "] + ["$0.00" for _ in all_dates] + [format_currency(ontario_hst)]
                royalty_tax2_row = ["Tax 2:"] + ["$0.00" for _ in all_dates]
                cash_payout_financial_data.append(royalty_tax1_row)
                cash_payout_financial_data.append(royalty_tax2_row)
                total_value = royalty_amt + ontario_hst
                royalty_tax1_row = ["Total Royalties Due"] + ["$0.00" for _ in all_dates] + [format_currency(total_value)]
                cash_payout_financial_data.append(royalty_tax1_row)
                cash_payout_financial_data.append([" "] + ["" for _ in all_dates]) 
                cash_payout_financial_data.append(["NATIONAL MARKETING"] + [""] * len(all_dates))
                cash_payout_financial_data.append(["Marketing Fee Description"] + ["Standard 1.25%"] * len(all_dates))
                Marketing_due_fees_values = [
                    format_currency(
                        (float(sales_by_date.loc['sales_total', date]) - 
                        float(total_discounts[date]))
                    *0.0125)
                    for date in all_dates
                ]
                marketing_due_fees_total = format_currency(
                    format_currency((float(sales_by_date.loc['sales_total', 'Total']) - 
                    float(total_discounts_total))*0.01245)
                )
                cash_payout_financial_data.append(["National Marketing"] + Marketing_due_fees_values + [marketing_due_fees_total])
                Marketing_tax1_values = [
                    format_currency(
                        (float(sales_by_date.loc['sales_total', date]) - 
                        float(total_discounts[date]))
                    *0.0125*0.13)
                    for date in all_dates
                ]
                marketing_tax1_total = format_currency(
                    format_currency((float(sales_by_date.loc['sales_total', 'Total']) - 
                    float(total_discounts_total))*0.0125*0.13)
                )
                cash_payout_financial_data.append(["Tax 1: Ontario HST"] + Marketing_tax1_values + [marketing_tax1_total])
                cash_payout_financial_data.append(["Tax 2:"] + ["$0.0" for _ in all_dates] + ["$0.0"])
                Total_marketing_due_fees_values = [
                    format_currency((
                        (float(sales_by_date.loc['sales_total', date]) - 
                        float(total_discounts[date]))
                     + (float(sales_by_date.loc['sales_total', date]) - 
                        float(total_discounts[date]))*0.13)*0.0125)
                    for date in all_dates
                ]
                Grand_total_marketing_due_fees = format_currency(
                    format_currency(((float(sales_by_date.loc['sales_total', 'Total']) - 
                    float(total_discounts_total)) + (float(sales_by_date.loc['sales_total', 'Total']) - 
                    float(total_discounts_total))*0.13)*0.0125)
                )

                cash_payout_financial_data.append(["Total Marketin Due:"] + Total_marketing_due_fees_values + [Grand_total_marketing_due_fees])
                def header_footer(canvas, doc):
                    canvas.saveState()
                    try:
                        # Import Image at the function level
                        from reportlab.platypus import Image
                        
                        # Use Django's static file finder to locate the image
                        from django.contrib.staticfiles.finders import find
                        image_path = find('site_panel/images/restaurant_icon.png')
                        
                        if image_path:
                            try:
                                img = Image(image_path, width=70, height=70)
                                # Draw the image at the top-left corner
                                img.drawOn(canvas, 50, landscape(letter)[1] - 80)
                            except Exception as img_error:
                                print(f"Error drawing image: {str(img_error)}")
                                # Fallback text if image fails to draw
                                canvas.setFont('Helvetica', 10)
                                canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                        else:
                            print("Image not found in static files")
                            # Fallback text if image not found
                            canvas.setFont('Helvetica', 10)
                            canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                    except Exception as e:
                        print(f"Error in image handling: {str(e)}")
                        # Final fallback if anything else goes wrong
                        canvas.setFont('Helvetica', 10)
                        canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")

                    # Rest of your header content
                    canvas.setFont('Helvetica', 25)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 50, "Weekly Flash Report")
                    canvas.setFont('Helvetica', 18)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 80, f"{restaurant_name} - {start_date} to {end_date}")

                    canvas.setFont('Helvetica', 8)
                    canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
                    canvas.drawRightString(landscape(letter)[0] - 30, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    canvas.restoreState()
                doc = SimpleDocTemplate(response, pagesize=landscape(letter))
                elements = []
                elements.append(Spacer(1, 20))
                styles = getSampleStyleSheet()

                my_light_blue = HexColor('#91aad6')

                # Calculate number of columns (should be same for all tables now)
                num_columns = len(all_dates) + 2  # Category + dates + Total

                # Adjust column widths
                available_width = landscape(letter)[0] - 2 * inch
                first_col_width = 1.8 * inch
                date_col_width = (available_width+(0.6 * inch)  - first_col_width - 1 * inch) / len(all_dates)  # Space for Total column
                col_widths = [first_col_width] + [date_col_width] * len(all_dates) + [1 * inch]  # Total column width

                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), my_light_blue),
                    ('BACKGROUND', (0, 0), (-1, 1), my_light_blue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),  # Highlight Total colum
                ])

                # Combine all data ensuring consistent column counts
                combined_data = []
                
                # Add sales data
                combined_data.extend(sales_deposits_data)
                combined_data.append([""] * num_columns)  # Blank row
                
                # Add payout data
                combined_data.extend(cash_payout_financial_data)

                # Create the table
                combined_table = Table(combined_data, colWidths=col_widths)
                combined_table.setStyle(table_style)
                elements.append(combined_table)

                doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)

        except Exception as e:
            return JsonResponse({"error": f"Error generating financial report: {str(e)}"}, status=500)
        if report_type.lower() == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                wb = Workbook()
                ws = wb.active
                
                # Define styling
                highlight_style = PatternFill(
                    start_color='91AAD6',  # Your light blue color
                    end_color='91AAD6',
                    fill_type='solid'
                )
                bold_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                try:
                    from openpyxl.drawing.image import Image
                    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                    image_path = os.path.join(static_dir, "restaurant_icon.png")
                    
                    # Verify image exists before trying to insert
                    if os.path.exists(image_path):
                        img = Image(image_path)
                        img.width = 70
                        img.height = 70
                        # Add image after text to avoid positioning issues
                        ws.add_image(img, 'A1')  # Positioned to the right of the title
                except Exception as img_error:
                    print(f"Image insertion skipped: {str(img_error)}")

                # Add title and date first (we'll add image after to avoid issues)
                ws['C1'] = "Weekly Flash Report"
                ws['C1'].font = Font(size=20, bold=True)
                
                ws['C2'] = f"{restaurant_name} - {start_date} to {end_date}"
                ws['C2'].font = Font(size=14)
                
                # Try to add image after setting up basic structure

                # Start data from row 4
                start_row = 5
                
                # Terms that trigger row highlighting
                highlight_terms = {
                    'Category',
                    'Total Sales(A)',
                    'Total Promotions(B)',
                    'Net Sales (A-B)',
                    'Total Royalties Due',
                    'Total Marketin Due'
                }
                
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # Check first column for highlight terms
                        if col_idx == 1 and str(value).strip() in highlight_terms:
                            # Apply to entire row
                            for col in range(1, len(row) + 1):
                                ws.cell(row=row_idx, column=col).fill = highlight_style
                                ws.cell(row=row_idx, column=col).font = bold_font
                
                wb.save(response)
                return response
                
            except Exception as e:
                return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)
        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
def generate_weekly_wage_cost_report(request):
    try:
        location = request.GET.get('location')
        end_date_str = request.GET.get('data_date')
        report_type = request.GET.get('report_type')

        if not location:
            return JsonResponse({"error": "Location is required"}, status=400)
        if not end_date_str:
            return JsonResponse({"error": "Date is required"}, status=400)

        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected %Y-%m-%d"}, status=400)

        try:
            user = User.objects.get(last_name=location)
        except User.DoesNotExist:
            return JsonResponse({"error": "No user found for the given location"}, status=404)

        site_id = user.id
        restaurant_name = user.first_name
        start_date = end_date - timedelta(days=6)

        with connection.cursor() as cursor:
            cursor.execute("SELECT MIN(activity_date) FROM activity_sales;")
            earliest_date = cursor.fetchone()[0]
        if earliest_date and start_date < earliest_date:
            return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=weekly_activity_sales_{site_id}_report.pdf'

        base_query = """
            SELECT
                s.activity_date AS Date,
                s.sales_total AS sales_total
            FROM
                activity_sales s
            LEFT JOIN
                auth_user u ON s.site_id = u.id
            WHERE s.activity_date BETWEEN %s AND %s AND s.site_id = %s;
        """

        query2 = """
            SELECT p.day_date AS Date, c.e_position_type, SUM(p.total_daily_rate)
            FROM social99.site_panel_paydailyentry AS p
            LEFT JOIN social99.compensation AS c
            ON p.c_id_id = c.c_id
            WHERE p.day_date BETWEEN %s AND %s AND c.s_id = %s
            GROUP BY p.day_date, c.e_position_type;
        """

        query_params = [start_date, end_date, site_id]
        date_range = pd.date_range(start=start_date, end=end_date)
        all_dates = [date.strftime('%Y-%m-%d') for date in date_range]

        all_dates_dt = pd.to_datetime(all_dates)
        date_headers = [d.strftime('%Y-%m-%d') for d in all_dates_dt]
        day_headers = [d.strftime('%A') for d in all_dates_dt]

        # Initialize df and df2 outside the inner try block
        df = pd.DataFrame(index=all_dates)
        df2 = pd.DataFrame(index=all_dates, columns=['Date', 'e_position_type', 'total_daily_rate'])

        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                results = cursor.fetchall()
                column_names_from_db = [col[0] for col in cursor.description]
                df = pd.DataFrame(results, columns=column_names_from_db)
                if 'Date' in df.columns:
                    df['Date'] = pd.to_datetime(df['Date'])
                    df.set_index('Date', inplace=True)
                    df = df.reindex(all_dates_dt, fill_value=0).reset_index()
                    df.rename(columns={'index': 'Date'}, inplace=True)
                for col in df.columns:
                    if col != 'Date':
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

                cursor.execute(query2, query_params)
                result_wages = cursor.fetchall()
                wages_column_names_from_db = [col[0] for col in cursor.description]

                for row in result_wages:
                    try:
                        date_obj = row[wages_column_names_from_db.index('Date')]
                        date_str = date_obj.strftime('%Y-%m-%d') if date_obj else None
                        position_type = row[wages_column_names_from_db.index('e_position_type')]
                        daily_rate_index = wages_column_names_from_db.index('SUM(p.total_daily_rate)')
                        daily_rate = row[daily_rate_index] if len(row) > daily_rate_index else None

                        if date_str:
                            df2.loc[date_str, 'e_position_type'] = position_type
                            df2.loc[date_str, 'total_daily_rate'] = float(daily_rate) if daily_rate is not None else 0.0
                    except ValueError as ve:
                        print(f"ValueError while processing row: {row}. Error: {ve}")
                    except IndexError as ie:
                        print(f"IndexError while processing row: {row}. Error: {ie}. Columns: {wages_column_names_from_db}")

            # Convert 'total_daily_rate' to numeric after populating
            df2['total_daily_rate'] = pd.to_numeric(df2['total_daily_rate'], errors='coerce').fillna(0.0)
            df2['Date'] = pd.to_datetime(df2.index)
            df2 = df2.sort_values(by='Date').reset_index(drop=True)

        except Exception as e:
            return JsonResponse({"error": f"Error generating financial report: {str(e)}"}, status=500)

        cash_sales_cols = [
            'sales_total'
        ]

        # Prepare Sales Data
        sales_df = df[['Date'] + cash_sales_cols].copy()
        sales_df['Date'] = sales_df['Date'].dt.strftime('%Y-%m-%d')
        sales_by_date = sales_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
        sales_by_date['Total'] = sales_by_date.sum(axis=1)  # Add Total column

        # Build the sales table data
        sales_deposits_data = [
            [""] + date_headers + ["Total"],
            [""] + day_headers + [""]  # Empty first cell, day names for each date, empty Total column
        ]

        category_map = {
            'sales_total': 'Total Sales',
        }

        from decimal import Decimal

        def format_currency(value):
            try:
                if isinstance(value, (str, bytes)):
                    # Clean potential HTML and commas from string inputs
                    cleaned_value = value.replace('<span class="math-inline">', '').replace('</span>', '').replace(',', '')
                    return f"${float(cleaned_value):,.2f}"
                elif isinstance(value, Decimal):
                    return f"${float(value):,.2f}"
                else:
                    return f"${float(value or 0):,.2f}"
            except (ValueError, TypeError):
                return "$0.00"

        # Add data rows with totals
        for col, display_name in category_map.items():
            values = [format_currency(sales_by_date.loc[col, date]) for date in all_dates]
            total = format_currency(sales_by_date.loc[col, 'Total'])
            sales_deposits_data.append([display_name] + values + [total])

        position_types = ['Front of the House', 'Back of the House', 'Staff/ Supervisor', 'Management/Owner']

        # Dictionary to store individual DataFrames
        position_dfs = {}

        # Loop through each type and filter
        for position in position_types:
            filtered_df = df2[df2['e_position_type'] == position][['Date', 'total_daily_rate']]
            position_dfs[position] = filtered_df.reset_index(drop=True)

        # Initialize summary DataFrame
        summary_data = {'Date': all_dates, 'Base Wages': [0.0] * len(all_dates), 'Benefits (15%)': [0.0] * len(all_dates), 'Total Wages': [0.0] * len(all_dates)}
        summary_df = pd.DataFrame(summary_data)
        summary_df['Date'] = pd.to_datetime(summary_df['Date'])

        # Iterate through position types and compute merged data
        position_summaries = {}  # Dictionary to hold the summary DataFrames for each position
        for position, position_df in position_dfs.items():
            if not position_df.empty:
                # Merge summary_df with the current position's DataFrame
                merged_df = pd.merge(summary_df, position_df, on='Date', how='left')
                merged_df['Base Wages'] = merged_df['total_daily_rate'].fillna(0)
                merged_df['Calculate Wages'] = merged_df['Base Wages'] / 1.15
                merged_df['Benefits (15%)'] = merged_df['Base Wages'] - merged_df['Calculate Wages']
                merged_df['Total Wages'] = merged_df['Base Wages'] + merged_df['Benefits (15%)']
                # Reorganize columns and save to the dictionary
                merged_df = merged_df[['Date', 'Base Wages', 'Benefits (15%)', 'Total Wages']]
                position_summaries[position] = merged_df
            else:
                position_summaries[position] = summary_df

        front_summary = position_summaries['Front of the House']
        back_summary = position_summaries['Back of the House']
        staff_summary = position_summaries['Staff/ Supervisor']
        management_summary = position_summaries['Management/Owner']

        summary_dfs = {
            'Front of the House': front_summary,
            'Back of the House': back_summary,
            'Staff/ Supervisor': staff_summary,
            'Management/Owner': management_summary,
        }

        for name, df in summary_dfs.items():
            total_row = pd.DataFrame({
                'Date': ['Total'],
                'Base Wages': [df['Base Wages'].sum()],
                'Benefits (15%)': [df['Benefits (15%)'].sum()],
                'Total Wages': [df['Total Wages'].sum()],
                # Add any other columns specific to each df and their totals here
            })
            summary_dfs[name] = pd.concat([df, total_row], ignore_index=True)

        # Now the updated DataFrames with the 'Total' row are back in the summary_dfs dictionary
        front_summary = summary_dfs['Front of the House']
        back_summary = summary_dfs['Back of the House']
        staff_summary = summary_dfs['Staff/ Supervisor']
        management_summary = summary_dfs['Management/Owner']

        def header_footer(canvas, doc):
            canvas.saveState()
            try:
                # Import Image at the function level
                from reportlab.platypus import Image
                        
                        # Use Django's static file finder to locate the image
                from django.contrib.staticfiles.finders import find
                image_path = find('site_panel/images/restaurant_icon.png')
                        
                if image_path:
                    try:
                        img = Image(image_path, width=70, height=70)
                        # Draw the image at the top-left corner
                        img.drawOn(canvas, 50, landscape(letter)[1] - 80)
                    except Exception as img_error:
                        print(f"Error drawing image: {str(img_error)}")
                        # Fallback text if image fails to draw
                        canvas.setFont('Helvetica', 10)
                        canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                else:
                    print("Image not found in static files")
                    # Fallback text if image not found
                    canvas.setFont('Helvetica', 10)
                    canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
            except Exception as e:
                print(f"Error in image handling: {str(e)}")
                # Final fallback if anything else goes wrong
                canvas.setFont('Helvetica', 10)
                canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")


            canvas.setFont('Helvetica', 25)
            canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 50, "Weekly Wage Costs")
            canvas.setFont('Helvetica', 18)
            canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 80, f"{restaurant_name} - Week Ending {end_date.strftime('%Y-%m-%d')}")

            canvas.setFont('Helvetica', 8)
            canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
            canvas.drawRightString(landscape(letter)[0] - 30, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            canvas.restoreState()

        doc = SimpleDocTemplate(response, pagesize=landscape(letter))
        elements = []

        my_light_blue = HexColor('#91aad6')

        # Calculate number of columns (should be same for all tables now)
        num_columns = len(all_dates) + 2  # Category + dates + Total

        # Adjust column widths
        available_width = landscape(letter)[0] - 2 * inch
        first_col_width = 1.8 * inch
        date_col_width = (available_width + (0.6 * inch) - first_col_width - 1 * inch) / len(all_dates)  # Space for Total column
        col_widths = [first_col_width] + [date_col_width] * len(all_dates) + [1 * inch]  # Total column width

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), my_light_blue),
            ('BACKGROUND', (0, 0), (-1, 1), my_light_blue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),  # Highlight Total column
        ])

        # Combine all data ensuring consistent column counts
        combined_data = []

        # Add sales data
        combined_data.extend(sales_deposits_data)
        combined_data.append([""] * num_columns)

        

        def format_currency_report(value):
            """Format numeric values with $ prefix and 2 decimal places"""
            try:
                return f"${float(value or 0):,.2f}"
            except (ValueError, TypeError):
                return "$0.00"

        def build_wage_report_data(summary_df, position_name, date_headers):
            data = [[position_name] + [""] * (len(date_headers) + 1)]  # Position header row
            data.append([""] + [d.strftime('%a') for d in pd.to_datetime(date_headers)] + [""])  # Day names row

            # Format all currency values with $ prefix
            base_wages = [format_currency_report(val) for val in summary_df.iloc[:-1]['Base Wages']]
            benefits = [format_currency_report(val) for val in summary_df.iloc[:-1]['Benefits (15%)']]
            total_wages = [format_currency_report(val) for val in summary_df.iloc[:-1]['Total Wages']]
            percent_sales = []

            data.append(["Base Wages"] + base_wages + [format_currency_report(summary_df.iloc[-1]['Base Wages'])])
            data.append(["Benefits (15%)"] + benefits + [format_currency_report(summary_df.iloc[-1]['Benefits (15%)'])])
            data.append(["Total Wages"] + total_wages + [format_currency_report(summary_df.iloc[-1]['Total Wages'])])

            # Calculate % Total Sales (no $ prefix needed here)
            total_sales_values = sales_by_date.loc['sales_total', all_dates].values
            total_wages_values = summary_df.iloc[:-1]['Total Wages'].values
            for i in range(len(all_dates)):
                sale = total_sales_values[i]
                wage = total_wages_values[i]
                percent = (wage / sale * 100) if sale else 0
                percent_sales.append(f"{percent:.2f}%")
            
            total_sales_total = sales_by_date.loc['sales_total', 'Total']
            total_wages_total = summary_df.iloc[-1]['Total Wages']
            total_percent_sales = (total_wages_total / total_sales_total * 100) if total_sales_total else 0
            data.append(["% Total Sales"] + percent_sales + [f"{total_percent_sales:.2f}%"])
            data.append([""] * len(data[0]))  # Add an empty row for spacing
            return data

        wage_report_data = []
        wage_report_data.extend(build_wage_report_data(front_summary, 'Front of House', date_headers))
        wage_report_data.extend(build_wage_report_data(back_summary, 'Back of House', date_headers))

        # Staff/ Supervisor might be empty, so add a placeholder if needed
        if not staff_summary.empty:
            wage_report_data.extend(build_wage_report_data(staff_summary, 'Staff / Supervisor', date_headers))
        else:
            wage_report_data.append(['Staff / Supervisor'] + [''] * (len(date_headers) + 1))
            wage_report_data.append(['Base Wages'] + [''] * (len(date_headers) + 1))
            wage_report_data.append(['Benefits (15%)'] + [''] * (len(date_headers) + 1))
            wage_report_data.append(['Total Wages'] + [''] * (len(date_headers) + 1))
            wage_report_data.append(['% Total Sales'] + [''] * (len(date_headers) + 1))
            wage_report_data.append([""] * (len(date_headers) + 2)) # Add an empty row for spacing

        wage_report_data.extend(build_wage_report_data(management_summary, 'Manager / Owner', date_headers))
        
        combined_data.extend(wage_report_data)
                # Add total wages table
        total_wages_data = [["Total Wages"] + [""] * (len(date_headers) + 1)]
        total_wages_data.append([""] + [d.strftime('%a') for d in pd.to_datetime(date_headers)] + [""])
        
        # Calculate sums across all position types
        base_wages_total = [front_summary.iloc[:-1]['Base Wages'].values[i] + back_summary.iloc[:-1]['Base Wages'].values[i] + 
                          staff_summary.iloc[:-1]['Base Wages'].values[i] + management_summary.iloc[:-1]['Base Wages'].values[i] 
                          for i in range(len(all_dates))]
        benefits_total = [front_summary.iloc[:-1]['Benefits (15%)'].values[i] + back_summary.iloc[:-1]['Benefits (15%)'].values[i] + 
                        staff_summary.iloc[:-1]['Benefits (15%)'].values[i] + management_summary.iloc[:-1]['Benefits (15%)'].values[i] 
                        for i in range(len(all_dates))]
        total_wages = [front_summary.iloc[:-1]['Total Wages'].values[i] + back_summary.iloc[:-1]['Total Wages'].values[i] + 
                     staff_summary.iloc[:-1]['Total Wages'].values[i] + management_summary.iloc[:-1]['Total Wages'].values[i] 
                     for i in range(len(all_dates))]
        
        # Format values
        # Format values

        total_wages_data.append(["Total Wages"] + [format_currency_report(val) for val in total_wages] + 
                       [format_currency_report(
                           front_summary.iloc[-1]['Total Wages'] + 
                           back_summary.iloc[-1]['Total Wages'] +
                           staff_summary.iloc[-1]['Total Wages'] + 
                           management_summary.iloc[-1]['Total Wages']
                       )])
        # Calculate % of Total Sales
        total_sales_values = sales_by_date.loc['sales_total', all_dates].values
        percent_sales_total = [(total_wages[i] / total_sales_values[i] * 100) if total_sales_values[i] else 0 for i in range(len(all_dates))]
        total_percent_sales = (sum(total_wages) / sales_by_date.loc['sales_total', 'Total'] * 100) if sales_by_date.loc['sales_total', 'Total'] else 0
        
        total_wages_data.append(["% Total Sales"] + [f"{val:.2f}%" for val in percent_sales_total] + [f"{total_percent_sales:.2f}%"])
        combined_data.extend(total_wages_data)
        combined_data.append([""] * num_columns)

        styles = getSampleStyleSheet()
        elements.append(Spacer(1, 12))
        wage_table = Table(combined_data, colWidths=col_widths)
        wage_table.setStyle(table_style)
        elements.append(wage_table)
        elements.append(Spacer(1, 24))

        doc.build(elements, onFirstPage=header_footer, onLaterPages=header_footer)
        if report_type.lower() == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                wb = Workbook()
                ws = wb.active
                
                # Define styling
                highlight_style = PatternFill(
                    start_color='91AAD6',
                    end_color='91AAD6',
                    fill_type='solid'
                )
                bold_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                try:
                    from openpyxl.drawing.image import Image
                    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                    image_path = os.path.join(static_dir, "restaurant_icon.png")
                    
                    # Verify image exists before trying to insert
                    if os.path.exists(image_path):
                        img = Image(image_path)
                        img.width = 70
                        img.height = 70
                        # Add image after text to avoid positioning issues
                        ws.add_image(img, 'A1')  # Positioned to the right of the title
                except Exception as img_error:
                    print(f"Image insertion skipped: {str(img_error)}")

                # Add title and date first (we'll add image after to avoid issues)
                ws['C1'] = "Weekly Wage Cost Report"
                ws['C1'].font = Font(size=20, bold=True)
                
                ws['C2'] = f"{restaurant_name} - {start_date} to {end_date}"
                ws['C2'].font = Font(size=14)
                
                # Try to add image after setting up basic structure

                # Start data from row 4
                start_row = 5
                
                # Terms that trigger row highlighting
                highlight_cells = [
                    (9, 1),  # A9 (Row 9, Column 1)
                    (16, 1), # A16 (Row 16, Column 1)
                    (23, 1), # A23 (Row 23, Column 1)
                    (30, 1),  # A30 (Row 30, Column 1)
                    (37, 1)  # A30 (Row 30, Column 1)
                ]

                # Loop through each row and column in the combined data
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # If the current cell is one of the specific ones to highlight, apply the styles
                        if (row_idx, col_idx) in highlight_cells:
                            cell.fill = highlight_style
                            cell.font = bold_font

                # List of specific rows you want to highlight (rows 5, 6, 39, and 40)
                highlight_rows = [5, 6, 38, 39, 40]

                # Loop through each row and column in the combined data
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # Highlight specific cells (A9, A16, A23, A30)
                        if (row_idx, col_idx) in highlight_cells:
                            cell.fill = highlight_style
                            cell.font = bold_font
                        
                        # Highlight rows 5, 6, 39, and 40, from column 1 to column 9
                        if row_idx in highlight_rows and col_idx <= 9:  # Only columns 1 to 9
                            cell.fill = highlight_style
                            cell.font = bold_font

                # Safe column auto-sizing
                for col in ws.columns:
                    try:
                        column_letter = col[0].column_letter
                        max_length = 0
                        for cell in col:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                continue
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
                    except:
                        continue

                wb.save(response)
                return response
                
            except Exception as e:
                return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)

        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def generate_monthly_roll_over(request):
    try:
        # Get request parameters
        location = request.GET.get('location')
        end_date_str = request.GET.get('data_date')
        report_type = request.GET.get('report_type')

        # Validate parameters
        if not location:
            return JsonResponse({"error": "Location is required", "status": "error"}, status=400)
        if not end_date_str:
            return JsonResponse({"error": "Date is required", "status": "error"}, status=400)

        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD", "status": "error"}, status=400)

        # Get user/site information
        try:
            user = User.objects.get(last_name=location)
        except User.DoesNotExist:
            return JsonResponse({"error": "No user found for the given location", "status": "error"}, status=404)
        restaurant_name = user.first_name

        site_id = user.id
        start_date = end_date - timedelta(days=27)

        # Validate date range
        with connection.cursor() as cursor:
            cursor.execute("SELECT MIN(day_date) FROM site_panel_paydailyentry;")
            earliest_date = cursor.fetchone()[0]

        if earliest_date and start_date < earliest_date:
            return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}", "status": "error"}, status=400)

        # Execute base query
        base_query = """
            SELECT
                s.activity_date AS "Date",
                s.food as "sales_food",
                s.liquor AS "sales_liquor",
                s.beer AS "sales_beer",
                s.wine AS "sales_wine",
                s.draft AS "sales_draft",
                s.beverages_iced_beverages AS "sales_beverages_iced_beverages",
                s.honest_to_goodness_fees AS "sales_honest_to_goodness_fees",
                s.miscellaneous_income AS "sales_miscellaneous_income",
                s.sales_total AS "sales_total",
                s.sales_tax1 AS "sales_tax1",
                s.sales_tax2 AS "sales_tax2",
                s.customer_count AS "customer_count",
                s.liquor_tax_1 AS "sales_liquor_tax_1",
                pr.staff_meals AS "promo_staff_meals",
                pr.manager_meals AS "promo_manager_meals",
                pr.coupons_disc AS "promo_coupons_disc",
                pr.chucks_bucks AS "promo_chucks_bucks",
                pr.lsm_1 AS "promo_lsm_1",
                pr.lsm_2 AS "promo_lsm_2",
                pr.manager_prom AS "promo_manager_prom",
                pr.qsa_complaints AS "promo_qsa_complaints",
                pr.promotion_total AS "promo_promotion_total"
            FROM
                activity_sales s
            LEFT JOIN
                activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
            LEFT JOIN
                auth_user u ON s.site_id = u.id
            WHERE s.activity_date BETWEEN %s AND %s AND s.site_id = %s;
        """

        query_params = [start_date, end_date, site_id]
        
        with connection.cursor() as cursor:
            cursor.execute(base_query, query_params)
            results = cursor.fetchall()
            results = [tuple(0 if value is None else value for value in row) for row in results]
            column_names_from_db = [col[0] for col in cursor.description]

            if 'Date' not in column_names_from_db:
                raise ValueError("Your base_query must return a 'Date' column.")

            # Create and process DataFrame
            df = pd.DataFrame(results, columns=column_names_from_db)
            df = df.fillna(0)
            df['Date'] = pd.to_datetime(df['Date'])
            df = df.sort_values(by='Date')

            # Calculate weeks
            df['Week'] = ((df['Date'] - pd.to_datetime(start_date)).dt.days // 7)
            
            # Group by week and calculate metrics
            weekly_sales = df.groupby('Week').agg({
                'sales_total': 'sum',
                'sales_liquor': 'sum',
                'sales_beer': 'sum',
                'sales_wine': 'sum',
                'sales_beverages_iced_beverages': 'sum',
                'promo_promotion_total': 'sum',
                'customer_count': 'sum'
            }).fillna(0) 

            weekly_sales['Food Sales'] = weekly_sales['sales_total'] - (
                weekly_sales['sales_liquor'] + 
                weekly_sales['sales_beer'] + 
                weekly_sales['sales_wine'] + 
                weekly_sales['sales_beverages_iced_beverages']
            )

            weekly_sales['Cold Beverages'] = (
                weekly_sales['sales_liquor'] + 
                weekly_sales['sales_beer'] + 
                weekly_sales['sales_wine'] + 
                weekly_sales['sales_beverages_iced_beverages']
            )

            weekly_sales['Total Discounts'] = weekly_sales['promo_promotion_total']
            weekly_sales['Net Sales'] = weekly_sales['sales_total'] - weekly_sales['Total Discounts']

            # Calculate percentages
            weekly_sales['Food %'] = (weekly_sales['Food Sales'] / weekly_sales['sales_total']) * 100
            weekly_sales['Cold Beverages %'] = (weekly_sales['Cold Beverages'] / weekly_sales['sales_total']) * 100
            weekly_sales['Discounts %'] = (weekly_sales['Total Discounts'] / weekly_sales['sales_total']) * 100

            # Get date ranges for each week
            date_ranges = df.groupby('Week')['Date'].agg(['min', 'max'])
            weekly_sales = weekly_sales.join(date_ranges).reset_index()

            # Create week dictionary
            week_dict = {}
            for week_num in range(4):  # Assuming 4 weeks in monthly report
                week_data = weekly_sales.loc[week_num] if week_num in weekly_sales.index else {
                    'sales_total': 0,
                    'Food Sales': 0,
                    'Cold Beverages': 0,
                    'Total Discounts': 0,
                    'Net Sales': 0,
                    'Food %': 0,
                    'Cold Beverages %': 0,
                    'Discounts %': 0,
                    'min': start_date + timedelta(days=7*week_num),
                    'max': start_date + timedelta(days=7*(week_num+1)-1)
                }
                
                week_dict[f"Week {week_num + 1}"] = {
                    "week_number": week_num + 1,
                    "date_range": f"{(start_date + timedelta(days=7*week_num)).strftime('%Y-%m-%d')} to {(start_date + timedelta(days=7*(week_num+1)-1)).strftime('%Y-%m-%d')}",
                    "sales_data": {
                        "Food Sales": {
                            "amount": float(week_data.get('Food Sales', 0)),
                            "percentage": float(week_data.get('Food %', 0))
                        },
                        "Cold Beverages": {
                            "amount": float(week_data.get('Cold Beverages', 0)),
                            "percentage": float(week_data.get('Cold Beverages %', 0))
                        },
                        "Total Sales": float(week_data.get('sales_total', 0)),
                        "Total Discounts": {
                            "amount": float(week_data.get('Total Discounts', 0)),
                            "percentage": float(week_data.get('Discounts %', 0))
                        },
                        "Net Sales": float(week_data.get('Net Sales', 0)),
                        "Customer Count": float(week_data.get('customer_count', 0))
                    }
                }

            # Create PDF in memory buffer
            buffer = BytesIO()
            
            # Create PDF document
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()

            centered_heading1 = ParagraphStyle(
                'CenteredHeading1',
                parent=styles['Heading1'],
                alignment=1  # 1 = TA_CENTER
            )

            centered_heading2 = ParagraphStyle(
                'CenteredHeading2',
                parent=styles['Heading2'],
                alignment=1  # 1 = TA_CENTER
            )
            
            # Build PDF content
            story = []
 
            story.append(Spacer(1, 100))

            combined_data = [
                ["", "", ""]
            ]
            def header_with_logo(canvas, doc, restaurant_name, start_date, end_date):
                canvas.saveState()

                # Load logo
                static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                image_path = os.path.join(static_dir, "restaurant_icon.png")

                try:
                    logo = Image(image_path, width=70, height=70)
                    logo.drawOn(canvas, 60, doc.pagesize[1] - 150)
                except Exception as e:
                    print(f"Error loading image: {e}")

                canvas.setFont('Helvetica', 25)
                canvas.drawCentredString(landscape(letter)[0] -450, landscape(letter)[1] + 75, "Monthly Rollover Report")
                canvas.setFont('Helvetica', 18)
                canvas.drawCentredString(landscape(letter)[0] -450, landscape(letter)[1] + 45, f"{restaurant_name} - Week Ending {end_date.strftime('%Y-%m-%d')}")

                canvas.setFont('Helvetica', 8)
                canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
                canvas.drawRightString(landscape(letter)[0] - 200, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                canvas.restoreState()

            for week_num, week_data in week_dict.items():
                
                # Create table data
                table_data = [
                    [f"{week_data['date_range']}", "Percentage", "Total ($)"],
                    [
                        "Food Sales", 
                        f"{week_data['sales_data']['Food Sales']['percentage']:.2f}%", 
                        f"${week_data['sales_data']['Food Sales']['amount']:,.2f}"
                    ],
                    [
                        "Cold Beverages", 
                        f"{week_data['sales_data']['Cold Beverages']['percentage']:.2f}%", 
                        f"${week_data['sales_data']['Cold Beverages']['amount']:,.2f}"
                    ],
                    ["Total Sales", "-", f"${week_data['sales_data']['Total Sales']:,.2f}"],
                    [
                        "Total Discounts", 
                        f"{week_data['sales_data']['Total Discounts']['percentage']:.2f}%", 
                        f"${week_data['sales_data']['Total Discounts']['amount']:,.2f}"
                    ],
                    ["Net Sales", "-", f"${week_data['sales_data']['Net Sales']:,.2f}"],
                ]
                my_light_blue = HexColor('#91aad6')

                week_name = week_data['date_range']
                combined_data.append([
                week_name, f"Percentage", f"Total ($)"
                ])
                # Add Food Sales row
                combined_data.append([
                    "Food Sales",
                    f"{week_data['sales_data']['Food Sales']['percentage']:.2f}%",
                    f"{week_data['sales_data']['Food Sales']['amount']:,.2f}"
                ])
                
                # Add Cold Beverages row
                combined_data.append([
                    "Cold Beverages", 
                    f"{week_data['sales_data']['Cold Beverages']['percentage']:.2f}%",
                    f"{week_data['sales_data']['Cold Beverages']['amount']:,.2f}"
                ])
                
                # Add Total Sales row
                combined_data.append([
                    "Total Sales",
                    "-",
                    f"{week_data['sales_data']['Total Sales']:,.2f}"
                ])
                
                # Add Total Discounts row
                combined_data.append([
                    "Total Discounts",
                    f"{week_data['sales_data']['Total Discounts']['percentage']:.2f}%",
                    f"{week_data['sales_data']['Total Discounts']['amount']:,.2f}"
                ])
                
                # Add Net Sales row
                combined_data.append([
                    "Net Sales",
                    "-",
                    f"{week_data['sales_data']['Net Sales']:,.2f}"
                ])
                
                # Add empty row between weeks
                combined_data.append([ "", "", ""])
                
                # Create table
                t = Table(table_data, colWidths=[2*inch, 2.5*inch, 2.5*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), my_light_blue),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold'),
                    ('LINEBELOW', (0,0), (-1,0), 1, colors.black),
                    ('LINEBELOW', (0,4), (-1,4), 1, colors.black),
                ]))
                
                story.append(t)
                story.append(Spacer(1, 24))
            
            # Generate PDF
            doc.build(story, 
                onFirstPage=lambda canvas, doc: header_with_logo(canvas, doc, restaurant_name, start_date, end_date), 
                onLaterPages=lambda canvas, doc: header_with_logo(canvas, doc, restaurant_name, start_date, end_date)
            )

            
            # Get PDF content and close buffer
            pdf = buffer.getvalue()
            buffer.close()
            
            # Create response
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="monthly_rollover_{restaurant_name}_{end_date}.pdf"'
            response.write(pdf)
            if report_type.lower() == 'excel':
                try:
                    # Generate Excel response
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="monthly_rollover_{restaurant_name}_{end_date}.xlsx"'
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Monthly Rollover"
                    
                    # Define styles
                    header_fill = PatternFill(start_color='91AAD6', end_color='91AAD6', fill_type='solid')
                    bold_font = Font(bold=True)
                    thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'),
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                    center_aligned = Alignment(horizontal='center', vertical='center')
                    title_font = Font(bold=True, size=14)

                    try:
                        from openpyxl.drawing.image import Image
                        static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                        image_path = os.path.join(static_dir, "restaurant_icon.png")
                        
                        # Verify image exists before trying to insert
                        if os.path.exists(image_path):
                            img = Image(image_path)
                            img.width = 70
                            img.height = 70
                            # Add image after text to avoid positioning issues
                            ws.add_image(img, 'A1')  # Positioned to the right of the title
                    except Exception as img_error:
                        print(f"Image insertion skipped: {str(img_error)}")

                    
                    # Add title row with restaurant name
                    ws.append([f"Monthly Rollover"])
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
                    title_cell = ws.cell(row=1, column=1)
                    title_cell.font = title_font
                    title_cell.alignment = center_aligned
                    
                    # Add date range
                    ws.append([f"Date Range: {start_date} to {end_date}"])
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
                    date_cell = ws.cell(row=2, column=1)
                    date_cell.alignment = center_aligned
                    
                    # Add empty row
                    ws.append([])
                    ws.append([])
                    ws.append([])
                    
                    # Write the data rows
                    for row in combined_data[1:]:
                        ws.append(row)
                    
                    # Apply styling
                    for row in ws.iter_rows():
                        # Check if this is a header row (contains "Percentage" or "Total ($)")
                        is_header_row = any(cell.value in ["Percentage", "Total ($)"] for cell in row)
                        
                        for cell in row:
                            cell.border = thin_border
                            if is_header_row:
                                cell.fill = header_fill
                                cell.font = bold_font
                    
                    # Auto-size columns
                    for col_idx in range(1, ws.max_column + 1):
                        max_length = 0
                        column_letter = get_column_letter(col_idx)
                        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                try:
                                    if not isinstance(cell, MergedCell):  # Skip merged cells
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                except:
                                    pass
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    wb.save(response)
                    return response
                    
                except Exception as e:
                    return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)
            return response

    except Exception as e:
        return JsonResponse({"error": str(e), "status": "error"}, status=500)
    
def generate_summary_report(request):
    try:
        # Get request parameters
        location = request.GET.get('location')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        report_type = request.GET.get('report_type')

        # Validate parameters
        if not location:
            return JsonResponse({"error": "Location is required", "status": "error"}, status=400)
        if not end_date_str:
            return JsonResponse({"error": "End date is required", "status": "error"}, status=400)
        if not start_date_str:
            return JsonResponse({"error": "Start date is required", "status": "error"}, status=400)
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD", "status": "error"}, status=400)
        
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            return JsonResponse({"error": "Invalid date format, expected YYYY-MM-DD", "status": "error"}, status=400)
        
        # Get user/site information
        try:
            user = User.objects.get(last_name=location)
        except User.DoesNotExist:
            return JsonResponse({"error": "No user found for the given location", "status": "error"}, status=404)

        site_id = user.id
        restaurant_name = user.first_name

        with connection.cursor() as cursor:
            cursor.execute("SELECT MIN(activity_date) FROM activity_sales;")
            earliest_date = cursor.fetchone()[0]

        if earliest_date and start_date < earliest_date:
            return JsonResponse({"error": f"Date cannot be earlier than {earliest_date}"}, status=400)

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=weekly_activity_sales_{site_id}_report.pdf'

        base_query = """
                    SELECT
                        s.activity_date AS Date,
                        s.food as "sales_food",
                        s.liquor AS "sales_liquor",
                        s.beer AS "sales_beer",
                        s.wine AS "sales_wine",
                        s.draft AS "sales_draft",
                        s.beverages_iced_beverages AS "sales_beverages_iced_beverages",
                        s.honest_to_goodness_fees AS "sales_honest_to_goodness_fees",
                        s.miscellaneous_income AS "sales_miscellaneous_income",
                        s.sales_total AS "sales_total",
                        s.sales_tax1 AS "sales_tax1",
                        s.sales_tax2 AS "sales_tax2",
                        s.liquor_tax_1 AS "sales_liquor_tax_1",
                        s.customer_count AS "customer_count",
                        pr.staff_meals AS "promo_staff_meals",
                        pr.manager_meals AS "promo_manager_meals",
                        pr.coupons_disc AS "promo_coupons_disc",
                        pr.chucks_bucks AS "promo_chucks_bucks",
                        pr.lsm_1 AS "promo_lsm_1",
                        pr.lsm_2 AS "promo_lsm_2",
                        pr.manager_prom AS "promo_manager_prom",
                        pr.qsa_complaints AS "promo_qsa_complaints",
                        pr.promotion_total AS "promo_promotion_total"
                    FROM
                        activity_sales s
                    LEFT JOIN
                        activity_promotions pr ON s.site_id = pr.site_id AND s.activity_date = pr.activity_date
                    LEFT JOIN
                        auth_user u ON s.site_id = u.id
                    WHERE s.activity_date BETWEEN %s AND %s AND s.site_id = %s;
                    """

        date_range = pd.date_range(start=start_date, end=end_date)
        all_dates = [date.strftime('%Y-%m-%d') for date in date_range]
    
        query_params = [start_date, end_date, site_id]
        try:
            with connection.cursor() as cursor:
                cursor.execute(base_query, query_params)
                results = cursor.fetchall()
                column_names_from_db = [col[0] for col in cursor.description]

                # Create DataFrame with all dates to ensure completeness
                df = pd.DataFrame(index=all_dates, columns=column_names_from_db)
                df = df.fillna(0)  # Fill missing values with 0
                
                # Populate with actual data
                for row in results:
                    date_str = row[column_names_from_db.index('Date')].strftime('%Y-%m-%d')
                    df.loc[date_str] = [0 if value is None else value for value in row]
                    
                df['Date'] = pd.to_datetime(df.index)
                df = df.sort_values(by='Date')

                # --- SALES TABLE ---
                cash_sales_cols = [
                    'sales_food', 'sales_beverages_iced_beverages', 'sales_liquor',
                    'sales_beer', 'sales_wine', 'sales_draft',
                    'sales_honest_to_goodness_fees', 'sales_miscellaneous_income',
                    'sales_total', 'customer_count',
                ]

                # Mapping of column names to display names
                category_map = {
                    'sales_food': 'Food',
                    'sales_beverages_iced_beverages': 'Beverages',
                    'sales_liquor': 'Liquor',
                    'sales_beer': 'Beer',
                    'sales_wine': 'Wine',
                    'sales_draft': 'Draft',
                    'sales_honest_to_goodness_fees': 'Honest To Goodness Fee',
                    'sales_miscellaneous_income': 'Miscellaneous Income',
                    'sales_total': 'Total',
                    'customer_count': 'Customer Count',
                }

                payout_cols = [
                    'promo_staff_meals', 'promo_manager_meals', 'promo_coupons_disc',
                    'promo_manager_prom', 'promo_qsa_complaints', 'promo_chucks_bucks',
                    'promo_lsm_1', 'promo_lsm_2', 'promo_promotion_total',
                ]

                payout_categories = {
                    'promo_staff_meals': 'Staff Meals',
                    'promo_manager_meals': 'Manager Meals',
                    'promo_coupons_disc': 'Coupons And Discounts',
                    'promo_manager_prom': 'Manager Promotions',
                    'promo_qsa_complaints': 'QSA Complaints',
                    'promo_chucks_bucks': "Chuck's Bucks",
                    'promo_lsm_1': 'LSM 1',
                    'promo_lsm_2': 'LSM 2',
                    'promo_promotion_total': 'Total',
                }

                # Prepare Sales Data with friendly names
                sales_df = df[['Date'] + cash_sales_cols].copy()
                sales_df['Date'] = sales_df['Date'].dt.strftime('%Y-%m-%d')
                sales_by_date = sales_df.groupby('Date').sum().reindex(all_dates, fill_value=0).T
                sales_by_date['Total'] = sales_by_date.sum(axis=1)
                sales_by_date.index = sales_by_date.index.map(lambda x: category_map.get(x, x))

                # Prepare Payout Data with friendly names
                payout_df = df[['Date'] + payout_cols].copy()
                payout_by_date = payout_df.groupby('Date').sum().reindex(pd.to_datetime(all_dates), fill_value=0).T
                payout_by_date.columns = [pd.to_datetime(col).strftime('%Y-%m-%d') for col in payout_by_date.columns]
                payout_by_date['Total'] = payout_by_date.sum(axis=1)
                payout_by_date.index = payout_by_date.index.map(lambda x: payout_categories.get(x, x))

                def header_footer(canvas, doc, date_chunk):
                    canvas.saveState()
                    try:
                        # Import Image at the function level
                        from reportlab.platypus import Image
                        # Use Django's static file finder to locate the image
                        from django.contrib.staticfiles.finders import find
                        image_path = find('site_panel/images/restaurant_icon.png')
                                
                        if image_path:
                            try:
                                img = Image(image_path, width=70, height=70)
                                # Draw the image at the top-left corner
                                img.drawOn(canvas, 50, landscape(letter)[1] - 80)
                            except Exception as img_error:
                                print(f"Error drawing image: {str(img_error)}")
                                # Fallback text if image fails to draw
                                canvas.setFont('Helvetica', 10)
                                canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                        else:
                            print("Image not found in static files")
                            # Fallback text if image not found
                            canvas.setFont('Helvetica', 10)
                            canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                    except Exception as e:
                        print(f"Error in image handling: {str(e)}")
                        # Final fallback if anything else goes wrong
                        canvas.setFont('Helvetica', 10)
                        canvas.drawString(50, landscape(letter)[1] - 80, "[Logo]")
                    
                    page_start = date_chunk[0]
                    page_end = date_chunk[-1]
                    
                    canvas.setFont('Helvetica', 25)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 50, "Weekly Activity Sales")
                    canvas.setFont('Helvetica', 18)
                    canvas.drawCentredString(landscape(letter)[0] / 2, landscape(letter)[1] - 80, f"{restaurant_name} - {page_start} to {page_end}")
                    
                    canvas.setFont('Helvetica', 8)
                    canvas.drawString(30, 20, f"Page {canvas.getPageNumber()}")
                    canvas.drawRightString(landscape(letter)[0] - 30, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    canvas.restoreState()
                    
                doc = SimpleDocTemplate(response, pagesize=landscape(letter))
                elements = []
                elements.append(Spacer(1, 20))
                styles = getSampleStyleSheet()
                my_light_blue = HexColor('#91aad6')
                
                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), my_light_blue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ])
            
                date_chunks = [all_dates[i:i+5] for i in range(0, len(all_dates), 5)]

                # Custom day name abbreviations
                day_abbreviations = {
                    'Sunday': 'Sun',
                    'Monday': 'Mon',
                    'Tuesday': 'Tues',
                    'Wednesday': 'Weds',
                    'Thursday': 'Thurs',
                    'Friday': 'Fri',
                    'Saturday': 'Sat'
                }

                for date_chunk in date_chunks:
                    sales_by_date_chunk = sales_by_date[date_chunk + ['Total']]
                    payout_by_date_chunk = payout_by_date[date_chunk + ['Total']]
                    
                    num_dates = len(date_chunk)
                    num_columns = 1 + (num_dates * 2) + 2  # 1 category + (dates * 2) + 2 total columns
                    
                    # Create header structure
                    header_row1 = [""] + [date for date in date_chunk for _ in (0, 1)] + ["Total", ""]
                    header_row2 = ["Sales"] + [day_abbreviations.get(datetime.strptime(date, '%Y-%m-%d').strftime('%A'), "") 
                                            for date in date_chunk for _ in (0, 1)] + ["", "%"]
                    
                    combined_data = [header_row1, header_row2]
                    
                    # Add sales data with correct percentage calculation
                    for col in sales_by_date_chunk.index:
                        row_data = [col]
                        
                        for date_str in date_chunk:
                            sales_value = sales_by_date_chunk.loc[col, date_str]
                            date_total = sales_by_date_chunk.loc['Total', date_str]  # Get total for this specific date
                            percentage = (sales_value / date_total * 100) if date_total != 0 else 0
                            row_data.extend([
                                f"${sales_value:,.2f}",
                                f"{percentage:.1f}%" if col != 'Total' else ""
                            ])
                        
                        # Add Total column data
                        total_value = sales_by_date_chunk.loc[col, 'Total']
                        grand_total = sales_by_date_chunk.loc['Total', 'Total']
                        row_data.extend([
                            f"${total_value:,.2f}",
                            f"{(total_value / grand_total * 100):.1f}%" if col != 'Total' and grand_total != 0 else ""
                        ])
                        
                        combined_data.append(row_data)
                    
                    combined_data.append([""] * num_columns)  # Empty row
                    
                    # Add Promotions section with correct percentage calculation
                    combined_data.append(["Promotions"] + [""] * (num_columns - 1))
                    combined_data.append(header_row2)
                    
                    for col in payout_by_date_chunk.index:
                        row_data = [col]
                        
                        for date_str in date_chunk:
                            payout_value = payout_by_date_chunk.loc[col, date_str]
                            date_total = sales_by_date_chunk.loc['Total', date_str]  # Use sales total for the date
                            percentage = (abs(payout_value) / date_total * 100) if date_total != 0 else 0
                            row_data.extend([
                                f"${payout_value:,.2f}",
                                f"{percentage:.1f}%" if col != 'Total' else ""
                            ])
                        
                        # Add Total column data
                        total_value = payout_by_date_chunk.loc[col, 'Total']
                        grand_total = sales_by_date_chunk.loc['Total', 'Total']
                        row_data.extend([
                            f"${total_value:,.2f}",
                            f"{(abs(total_value) / grand_total * 100):.1f}%" if col != 'Total' and grand_total != 0 else ""
                        ])
                        
                        combined_data.append(row_data)                   
                    # Column widths
                    col_widths = [1.5 * inch]  # Category column
                    col_widths += [0.8 * inch, 0.5 * inch] * num_dates  # Date value + percentage columns
                    col_widths += [1 * inch, 0.5 * inch]  # Total columns
                    
                    combined_table = Table(combined_data, colWidths=col_widths)
                    
                    # Table styling - now fully dynamic
                    style_commands = [
                        ('BACKGROUND', (0,0), (-1,1), my_light_blue),
                        ('BACKGROUND', (0,len(combined_data)-len(payout_by_date_chunk)-3), 
                        (-1,len(combined_data)-len(payout_by_date_chunk)-3), my_light_blue),
                        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ('GRID', (0,0), (-1,-1), 1, colors.black),
                        ('FONTNAME', (0,0), (-1,1), 'Helvetica-Bold'),
                        ('FONTNAME', (0,len(combined_data)-len(payout_by_date_chunk)-3), 
                        (-1,len(combined_data)-len(payout_by_date_chunk)-3), 'Helvetica-Bold'),
                        ('FONTSIZE', (0,0), (-1,1), 10),
                        ('FONTSIZE', (0,2), (-1,-1), 9),
                    ]
                    
                    # Add date header spans
                    for i in range(num_dates):
                        start_col = 1 + (i * 2)
                        style_commands.append(('SPAN', (start_col,0), (start_col+1,0)))
                    
                    # Add Total header span
                    style_commands.append(('SPAN', (num_columns-2,0), (num_columns-1,0)))
                    
                    # Add Promotions header span
                    promotions_row = len(combined_data)-len(payout_by_date_chunk)-3
                    style_commands.append(('SPAN', (0,promotions_row), (num_columns-1,promotions_row)))
                    
                    combined_table.setStyle(TableStyle(style_commands))
                    elements.append(combined_table)
                    elements.append(Spacer(1, 0.5 * inch))

                doc.build(elements, onFirstPage=lambda c, d: header_footer(c, d, date_chunks[0]), 
                        onLaterPages=lambda c, d: header_footer(c, d, date_chunks[0]))

        except Exception as e:
            return JsonResponse({"error": f"Error generating financial report: {str(e)}"}, status=500)
        if report_type.lower() == 'excel':
            try:
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                wb = Workbook()
                ws = wb.active
                
                # Define styling
                highlight_style = PatternFill(
                    start_color='91AAD6',
                    end_color='91AAD6',
                    fill_type='solid'
                )
                bold_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )

                try:
                    from openpyxl.drawing.image import Image
                    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static/site_panel/images')
                    image_path = os.path.join(static_dir, "restaurant_icon.png")
                    
                    # Verify image exists before trying to insert
                    if os.path.exists(image_path):
                        img = Image(image_path)
                        img.width = 70
                        img.height = 70
                        # Add image after text to avoid positioning issues
                        ws.add_image(img, 'A1')  # Positioned to the right of the title
                except Exception as img_error:
                    print(f"Image insertion skipped: {str(img_error)}")

                # Add title and date first (we'll add image after to avoid issues)
                ws['C1'] = "Summary Report"
                ws['C1'].font = Font(size=20, bold=True)
                
                ws['C2'] = f"{restaurant_name} - {start_date} to {end_date}"
                ws['C2'].font = Font(size=14)
                
                # Try to add image after setting up basic structure

                # Start data from row 4
                start_row = 5
                
                # Terms that trigger row highlighting
                highlight_cells = [
                    (9, 1),  # A9 (Row 9, Column 1)
                    (16, 1), # A16 (Row 16, Column 1)
                    (23, 1), # A23 (Row 23, Column 1)
                    (30, 1),  # A30 (Row 30, Column 1)
                    (37, 1)  # A30 (Row 30, Column 1)
                ]

                # Loop through each row and column in the combined data
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # If the current cell is one of the specific ones to highlight, apply the styles
                        if (row_idx, col_idx) in highlight_cells:
                            cell.fill = highlight_style
                            cell.font = bold_font

                # List of specific rows you want to highlight (rows 5, 6, 39, and 40)
                highlight_rows = [5, 6, 38, 39, 40]

                # Loop through each row and column in the combined data
                for row_idx, row in enumerate(combined_data, start_row):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        
                        # Highlight specific cells (A9, A16, A23, A30)
                        if (row_idx, col_idx) in highlight_cells:
                            cell.fill = highlight_style
                            cell.font = bold_font
                        
                        # Highlight rows 5, 6, 39, and 40, from column 1 to column 9
                        if row_idx in highlight_rows and col_idx <= 9:  # Only columns 1 to 9
                            cell.fill = highlight_style
                            cell.font = bold_font

                # Safe column auto-sizing
                for col in ws.columns:
                    try:
                        column_letter = col[0].column_letter
                        max_length = 0
                        for cell in col:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                continue
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column_letter].width = adjusted_width
                    except:
                        continue

                wb.save(response)
                return response
                
            except Exception as e:
                return JsonResponse({"error": f"Excel generation failed: {str(e)}"}, status=500)


        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def calculate_column_widths(data, font_name="Helvetica", font_size=6):
    """Calculate column widths based on cell content."""
    column_widths = []
    for column in zip(*data):  # Transpose data to iterate over columns
        max_width = max([stringWidth(str(cell), font_name, font_size) for cell in column])
        column_widths.append(max_width + 10)  # Add padding for readability
    return column_widths